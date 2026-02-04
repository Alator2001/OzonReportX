# -*- coding: utf-8 -*-
"""
–ú–æ–¥—É–ª—å —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è —Ü–µ–Ω–∞–º–∏ - —Ä–∞–∑–¥–µ–ª–µ–Ω–∏–µ –ª–æ–≥–∏–∫–∏ –Ω–∞ –æ—Ç–¥–µ–ª—å–Ω—ã–µ –¥–µ–π—Å—Ç–≤–∏—è.
"""

import json
import sys
from pathlib import Path
from typing import Dict, Optional, Tuple

# –î–æ–±–∞–≤–ª—è–µ–º –ø—É—Ç—å –∫ scripts –¥–ª—è –∏–º–ø–æ—Ä—Ç–∞
script_dir = Path(__file__).resolve().parent
if str(script_dir) not in sys.path:
    sys.path.insert(0, str(script_dir))

from recommended_prices import (
    load_margin_settings,
    save_margin_settings,
    MIN_MARGIN_DEFAULT,
    DESIRED_MARGIN_DEFAULT,
    load_costs_df,
    COSTS_FILENAME,
    compute_prices,
    get_product_prices_from_ozon,
    get_actions_for_products,
    COL_MIN_PRICE,
    COL_DESIRED_PRICE,
    COL_CURRENT_PRICE,
    COL_MARKETING_PRICE,
    COL_CURRENT_MARGIN,
    compute_current_margin,
    get_report_path,
    get_prev_month_year,
    load_rates_from_report,
    generate_monthly_report,
    MONTHS_RU,
    collect_deactivation_candidates_from_sheet,
    deactivate_products_in_action,
    get_action_candidates,
    activate_products_in_action,
    _artikul_normalize,
    _normalize_offer_id,
    get_discount_requests,
    approve_discount_requests,
    decline_discount_requests,
    get_sku_to_offer_id_mapping,
)

try:
    from utils import prompt_yes_no, print_step, log_verbose
except ImportError:
    def prompt_yes_no(prompt: str, default_yes: bool = True) -> bool:
        default_str = "Y/n" if default_yes else "y/N"
        response = input(f"{prompt} ({default_str}): ").strip().lower()
        if not response:
            return default_yes
        return response in ("y", "yes", "–¥–∞", "–¥")
    
    def print_step(text: str):
        print(f"\n¬∑ {text}")
    
    def log_verbose(_msg: str) -> None:
        pass

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter


def action_set_margin_range(repo_root: Path) -> Tuple[float, float]:
    """
    –î–µ–π—Å—Ç–≤–∏–µ 1: –î–∏–∞–ø–∞–∑–æ–Ω —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç–∏.
    –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –∑–∞–¥–∞—ë—Ç –¥–∏–∞–ø–∞–∑–æ–Ω –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π –∏ –∂–µ–ª–∞—Ç–µ–ª—å–Ω–æ–π —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç–∏.
    """
    print_step("–î–∏–∞–ø–∞–∑–æ–Ω —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç–∏")
    
    # –ó–∞–≥—Ä—É–∂–∞–µ–º —Å–æ—Ö—Ä–∞–Ω—ë–Ω–Ω—ã–µ –Ω–∞—Å—Ç—Ä–æ–π–∫–∏
    saved_min, saved_desired = load_margin_settings(repo_root)
    
    if saved_min is not None and saved_desired is not None:
        print(f"–¢–µ–∫—É—â–∏–µ –Ω–∞—Å—Ç—Ä–æ–π–∫–∏:")
        print(f"  –ú–∏–Ω–∏–º–∞–ª—å–Ω–∞—è —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç—å: {saved_min*100:.1f}%")
        print(f"  –ñ–µ–ª–∞—Ç–µ–ª—å–Ω–∞—è —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç—å: {saved_desired*100:.1f}%")
        
        if not prompt_yes_no("–ò–∑–º–µ–Ω–∏—Ç—å –Ω–∞—Å—Ç—Ä–æ–π–∫–∏?", default_yes=False):
            return saved_min, saved_desired
    
    # –ó–∞–ø—Ä–∞—à–∏–≤–∞–µ–º –Ω–æ–≤—ã–µ –∑–Ω–∞—á–µ–Ω–∏—è
    while True:
        try:
            min_input = input(f"–ú–∏–Ω–∏–º–∞–ª—å–Ω–∞—è —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç—å (–¥–æ–ª—è 0-1, –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é {MIN_MARGIN_DEFAULT}): ").strip()
            if not min_input:
                min_margin = MIN_MARGIN_DEFAULT
            else:
                min_margin = float(min_input.replace(",", "."))
                if not (0 < min_margin < 1):
                    print(f"–†–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç—å –¥–æ–ª–∂–Ω–∞ –±—ã—Ç—å –º–µ–∂–¥—É 0 –∏ 1. –ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –∑–Ω–∞—á–µ–Ω–∏–µ –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é {MIN_MARGIN_DEFAULT}.")
                    min_margin = MIN_MARGIN_DEFAULT
            
            desired_input = input(f"–ñ–µ–ª–∞—Ç–µ–ª—å–Ω–∞—è —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç—å (–¥–æ–ª—è 0-1, –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é {DESIRED_MARGIN_DEFAULT}): ").strip()
            if not desired_input:
                desired_margin = DESIRED_MARGIN_DEFAULT
            else:
                desired_margin = float(desired_input.replace(",", "."))
                if not (0 < desired_margin < 1):
                    print(f"–†–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç—å –¥–æ–ª–∂–Ω–∞ –±—ã—Ç—å –º–µ–∂–¥—É 0 –∏ 1. –ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –∑–Ω–∞—á–µ–Ω–∏–µ –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é {DESIRED_MARGIN_DEFAULT}.")
                    desired_margin = DESIRED_MARGIN_DEFAULT
            
            if min_margin >= desired_margin:
                print("‚ö†Ô∏è –ú–∏–Ω–∏–º–∞–ª—å–Ω–∞—è —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç—å –¥–æ–ª–∂–Ω–∞ –±—ã—Ç—å –º–µ–Ω—å—à–µ –∂–µ–ª–∞—Ç–µ–ª—å–Ω–æ–π. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
                continue
            
            break
        except (ValueError, KeyboardInterrupt):
            print("‚ö†Ô∏è –ù–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã–π –≤–≤–æ–¥. –ò—Å–ø–æ–ª—å–∑—É—é—Ç—Å—è –∑–Ω–∞—á–µ–Ω–∏—è –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é.")
            min_margin = MIN_MARGIN_DEFAULT
            desired_margin = DESIRED_MARGIN_DEFAULT
            break
    
    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –Ω–∞—Å—Ç—Ä–æ–π–∫–∏
    save_margin_settings(repo_root, min_margin, desired_margin)
    
    return min_margin, desired_margin


def action_calculate_optimal_prices(repo_root: Path) -> bool:
    """
    –î–µ–π—Å—Ç–≤–∏–µ 2: –†–∞—Å—Å—á–∏—Ç–∞—Ç—å –æ–ø—Ç–∏–º–∞–ª—å–Ω—É—é —Ü–µ–Ω—É.
    –†–∞—Å—á—ë—Ç –∫–æ–ª–æ–Ω–æ–∫ –ú–∏–Ω–∏–º–∞–ª—å–Ω–∞—è —Ü–µ–Ω–∞ –ø—Ä–æ–¥–∞–∂–∏ –∏ –ñ–µ–ª–∞—Ç–µ–ª—å–Ω–∞—è —Ü–µ–Ω–∞ –ø—Ä–æ–¥–∞–∂–∏.
    """
    print_step("–†–∞—Å—Å—á–∏—Ç–∞—Ç—å –æ–ø—Ç–∏–º–∞–ª—å–Ω—É—é —Ü–µ–Ω—É")
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –Ω–∞—Å—Ç—Ä–æ–µ–∫ —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç–∏
    min_margin, desired_margin = load_margin_settings(repo_root)
    
    if min_margin is None or desired_margin is None:
        print("‚ö†Ô∏è –î–∏–∞–ø–∞–∑–æ–Ω —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç–∏ –Ω–µ –∑–∞–¥–∞–Ω.")
        if prompt_yes_no("–ó–∞–¥–∞—Ç—å –¥–∏–∞–ø–∞–∑–æ–Ω —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç–∏ —Å–µ–π—á–∞—Å?", default_yes=True):
            min_margin, desired_margin = action_set_margin_range(repo_root)
        else:
            print("‚ùå –ù–µ–≤–æ–∑–º–æ–∂–Ω–æ —Ä–∞—Å—Å—á–∏—Ç–∞—Ç—å —Ü–µ–Ω—ã –±–µ–∑ –¥–∏–∞–ø–∞–∑–æ–Ω–∞ —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç–∏.")
            return False
    
    # –ü–æ–ª—É—á–∞–µ–º –æ—Ç—á—ë—Ç –¥–ª—è —Ä–∞—Å—á—ë—Ç–∞ –∫–æ–º–∏—Å—Å–∏–∏
    prev_year, prev_month = get_prev_month_year()
    report_path = get_report_path(repo_root, prev_year, prev_month)
    costs_path = repo_root / COSTS_FILENAME
    
    print(f"–ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –æ—Ç—á—ë—Ç –∑–∞ –ø—Ä–µ–¥—ã–¥—É—â–∏–π –º–µ—Å—è—Ü: {MONTHS_RU[prev_month - 1]} {prev_year}")
    
    if not report_path.exists():
        print(f"‚ö† –§–∞–π–ª –æ—Ç—á—ë—Ç–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω: {report_path.name}")
        if prompt_yes_no("–°–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞—Ç—å –æ—Ç—á—ë—Ç –∑–∞ –ø—Ä–µ–¥—ã–¥—É—â–∏–π –º–µ—Å—è—Ü?", default_yes=True):
            try:
                report_path = generate_monthly_report(repo_root, prev_month, prev_year)
            except Exception as e:
                print(f"‚ùå –ù–µ —É–¥–∞–ª–æ—Å—å —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞—Ç—å –æ—Ç—á—ë—Ç: {e}")
                return False
        else:
            print("‚ùå –ù–µ–≤–æ–∑–º–æ–∂–Ω–æ —Ä–∞—Å—Å—á–∏—Ç–∞—Ç—å —Ü–µ–Ω—ã –±–µ–∑ –æ—Ç—á—ë—Ç–∞.")
            return False
    
    log_verbose(f"–§–∞–π–ª –æ—Ç—á—ë—Ç–∞: {report_path}")
    total_rate = load_rates_from_report(report_path)
    log_verbose(f"–ö–æ–º–∏—Å—Å–∏—è+–ª–æ–≥–∏—Å—Ç–∏–∫–∞: {total_rate*100:.2f}%")
    df, key_col, cost_col = load_costs_df(costs_path)
    log_verbose(f"–ó–∞–≥—Ä—É–∂–µ–Ω–æ –∑–∞–ø–∏—Å–µ–π: {len(df)}")
    
    # –†–∞—Å—Å—á–∏—Ç—ã–≤–∞–µ–º —Ü–µ–Ω—ã
    min_prices = []
    desired_prices = []
    
    for _, row in df.iterrows():
        try:
            cost_val = float(row.get(cost_col, 0) or 0)
        except (TypeError, ValueError):
            cost_val = 0.0
        min_p, des_p = compute_prices(cost_val, total_rate, min_margin, desired_margin)
        min_prices.append(min_p)
        desired_prices.append(des_p)
    
    # –û–±–Ω–æ–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫–∏
    if COL_MIN_PRICE in df.columns:
        df = df.drop(columns=[COL_MIN_PRICE])
    if COL_DESIRED_PRICE in df.columns:
        df = df.drop(columns=[COL_DESIRED_PRICE])
    
    df[COL_MIN_PRICE] = min_prices
    df[COL_DESIRED_PRICE] = desired_prices
    
    df.to_excel(costs_path, index=False)
    print(f"‚úÖ –†–∞—Å—Å—á–∏—Ç–∞–Ω—ã –æ–ø—Ç–∏–º–∞–ª—å–Ω—ã–µ —Ü–µ–Ω—ã –¥–ª—è {len(df)} —Ç–æ–≤–∞—Ä–æ–≤ (–º–∞—Ä–∂–∞ {min_margin*100:.0f}% / {desired_margin*100:.0f}%).")
    return True


def action_get_current_prices(repo_root: Path) -> bool:
    """
    –î–µ–π—Å—Ç–≤–∏–µ 3: –£–∑–Ω–∞—Ç—å —Ç–µ–∫—É—â—É—é —Ü–µ–Ω—É –ø—Ä–æ–¥–∞–∂–∏.
    –†–∞—Å—á—ë—Ç –∫–æ–ª–æ–Ω–æ–∫ –¢–µ–∫—É—â–∞—è —Ü–µ–Ω–∞ –Ω–∞ Ozon, –¶–µ–Ω–∞ —Å —É—á—ë—Ç–æ–º –∞–∫—Ü–∏–π –∏ —Å–∫–∏–¥–æ–∫, –û–∂–∏–¥–∞–µ–º–∞—è —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç—å.
    """
    print_step("–£–∑–Ω–∞—Ç—å —Ç–µ–∫—É—â—É—é —Ü–µ–Ω—É –ø—Ä–æ–¥–∞–∂–∏")
    
    costs_path = repo_root / COSTS_FILENAME
    
    if not costs_path.exists():
        print(f"‚ùå –§–∞–π–ª {COSTS_FILENAME} –Ω–µ –Ω–∞–π–¥–µ–Ω.")
        return False
    
    df, key_col, cost_col = load_costs_df(costs_path)
    log_verbose(f"–ó–∞–≥—Ä—É–∂–µ–Ω–æ –∑–∞–ø–∏—Å–µ–π: {len(df)}")
    if COL_MIN_PRICE not in df.columns:
        print(f"‚ö†Ô∏è –ö–æ–ª–æ–Ω–∫–∞ ¬´{COL_MIN_PRICE}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω–∞.")
        if prompt_yes_no("–†–∞—Å—Å—á–∏—Ç–∞—Ç—å –æ–ø—Ç–∏–º–∞–ª—å–Ω—ã–µ —Ü–µ–Ω—ã —Å–µ–π—á–∞—Å?", default_yes=True):
            if not action_calculate_optimal_prices(repo_root):
                return False
            # –ü–µ—Ä–µ–∑–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ
            df, key_col, cost_col = load_costs_df(costs_path)
        else:
            print("‚ùå –ù–µ–≤–æ–∑–º–æ–∂–Ω–æ —Ä–∞—Å—Å—á–∏—Ç–∞—Ç—å —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç—å –±–µ–∑ –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π —Ü–µ–Ω—ã.")
            return False
    
    # –ü–æ–ª—É—á–∞–µ–º –æ—Ç—á—ë—Ç –¥–ª—è —Ä–∞—Å—á—ë—Ç–∞ –∫–æ–º–∏—Å—Å–∏–∏
    prev_year, prev_month = get_prev_month_year()
    report_path = get_report_path(repo_root, prev_year, prev_month)
    
    if not report_path.exists():
        print(f"‚ö† –§–∞–π–ª –æ—Ç—á—ë—Ç–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω: {report_path.name}")
        if prompt_yes_no("–°–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞—Ç—å –æ—Ç—á—ë—Ç –∑–∞ –ø—Ä–µ–¥—ã–¥—É—â–∏–π –º–µ—Å—è—Ü?", default_yes=True):
            try:
                report_path = generate_monthly_report(repo_root, prev_month, prev_year)
            except Exception as e:
                print(f"‚ùå –ù–µ —É–¥–∞–ª–æ—Å—å —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞—Ç—å –æ—Ç—á—ë—Ç: {e}")
                return False
        else:
            print("‚ùå –ù–µ–≤–æ–∑–º–æ–∂–Ω–æ —Ä–∞—Å—Å—á–∏—Ç–∞—Ç—å —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç—å –±–µ–∑ –æ—Ç—á—ë—Ç–∞.")
            return False
    
    total_rate = load_rates_from_report(report_path)
    log_verbose(f"–ö–æ–º–∏—Å—Å–∏—è+–ª–æ–≥–∏—Å—Ç–∏–∫–∞: {total_rate*100:.2f}%")
    log_verbose("–ü–æ–ª—É—á–µ–Ω–∏–µ —Ü–µ–Ω —Å Ozon...")
    offer_ids_list = []
    for _, row in df.iterrows():
        art = _artikul_normalize(row.get(key_col))
        if art:
            offer_ids_list.append(art)
    
    prices_map, marketing_prices_map = get_product_prices_from_ozon(offer_ids_list)
    if prices_map:
        print(f"‚úÖ –ü–æ–ª—É—á–µ–Ω–æ —Ü–µ–Ω –¥–ª—è {len(prices_map)} –∞—Ä—Ç–∏–∫—É–ª–æ–≤")
    else:
        print("‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å —Ü–µ–Ω—ã —Å Ozon (–≤–æ–∑–º–æ–∂–Ω–æ, –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã API –∫–ª—é—á–∏).")
    
    # –†–∞—Å—Å—á–∏—Ç—ã–≤–∞–µ–º —Ç–µ–∫—É—â–∏–µ —Ü–µ–Ω—ã –∏ —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç—å
    current_prices = []
    marketing_prices = []
    current_margins = []
    
    for _, row in df.iterrows():
        try:
            cost_val = float(row.get(cost_col, 0) or 0)
        except (TypeError, ValueError):
            cost_val = 0.0
        
        art = _artikul_normalize(row.get(key_col))
        if art:
            art_normalized = _normalize_offer_id(art)
            current_price = prices_map.get(art) or prices_map.get(art_normalized)
            marketing_price = marketing_prices_map.get(art) or marketing_prices_map.get(art_normalized)
        else:
            current_price = None
            marketing_price = None
        
        current_prices.append(round(current_price) if current_price is not None else None)
        marketing_prices.append(round(marketing_price) if marketing_price is not None else None)
        
        price_for_margin = marketing_price if marketing_price is not None else current_price
        margin = compute_current_margin(price_for_margin, cost_val, total_rate)
        current_margins.append(round(margin * 100, 2) if margin is not None else None)
    
    # –û–±–Ω–æ–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫–∏
    for c in [COL_CURRENT_PRICE, COL_MARKETING_PRICE, COL_CURRENT_MARGIN]:
        if c in df.columns:
            df = df.drop(columns=[c])
    
    df[COL_CURRENT_PRICE] = current_prices
    df[COL_MARKETING_PRICE] = marketing_prices
    df[COL_CURRENT_MARGIN] = current_margins
    
    df.to_excel(costs_path, index=False)
    
    # –ü—Ä–∏–º–µ–Ω—è–µ–º —É—Å–ª–æ–≤–Ω–æ–µ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ
    try:
        wb = load_workbook(costs_path)
        ws = wb.active
        
        # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –¥–ª—è —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç–∏
        min_margin, desired_margin = load_margin_settings(repo_root)
        if min_margin is None:
            min_margin = MIN_MARGIN_DEFAULT
        if desired_margin is None:
            desired_margin = DESIRED_MARGIN_DEFAULT
        
        margin_col_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == COL_CURRENT_MARGIN:
                margin_col_idx = col_idx
                break
        
        if margin_col_idx:
            min_margin_pct = min_margin * 100
            desired_margin_pct = desired_margin * 100
            
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            green_rule = CellIsRule(
                operator="between",
                formula=[min_margin_pct, desired_margin_pct],
                fill=green_fill
            )
            
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_rule = CellIsRule(
                operator="lessThan",
                formula=[min_margin_pct],
                fill=red_fill
            )
            
            margin_col_letter = ws.cell(row=1, column=margin_col_idx).column_letter
            data_range = f"{margin_col_letter}2:{margin_col_letter}{len(df) + 1}"
            ws.conditional_formatting.add(data_range, green_rule)
            ws.conditional_formatting.add(data_range, red_rule)
        
        # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –¥–ª—è —Ç–µ–∫—É—â–µ–π —Ü–µ–Ω—ã: –∫—Ä–∞—Å–Ω—ã–π < –º–∏–Ω, –∑–µ–ª—ë–Ω—ã–π >= –º–∏–Ω, –±–æ–ª–µ–µ –∑–µ–ª—ë–Ω—ã–π >= –∂–µ–ª–∞—Ç–µ–ª—å–Ω–æ–π
        current_price_col_idx = None
        min_price_col_idx = None
        desired_price_col_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == COL_CURRENT_PRICE:
                current_price_col_idx = col_idx
            elif cell.value == COL_MIN_PRICE:
                min_price_col_idx = col_idx
            elif cell.value == COL_DESIRED_PRICE:
                desired_price_col_idx = col_idx
        
        if current_price_col_idx and min_price_col_idx:
            current_price_col_letter = ws.cell(row=1, column=current_price_col_idx).column_letter
            min_price_col_letter = ws.cell(row=1, column=min_price_col_idx).column_letter
            data_range = f"{current_price_col_letter}2:{current_price_col_letter}{len(df) + 1}"
            
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_formula = f"AND({current_price_col_letter}2<>\"\", {current_price_col_letter}2>0, {current_price_col_letter}2<{min_price_col_letter}2)"
            red_rule = FormulaRule(formula=[red_formula], fill=red_fill, stopIfTrue=False)
            
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            green_formula = f"AND({current_price_col_letter}2<>\"\", {current_price_col_letter}2>0, {current_price_col_letter}2>={min_price_col_letter}2)"
            green_rule = FormulaRule(formula=[green_formula], fill=green_fill, stopIfTrue=False)
            
            ws.conditional_formatting.add(data_range, red_rule)
            ws.conditional_formatting.add(data_range, green_rule)
            
            # –¶–µ–Ω–∞ –≤—ã—à–µ –¥–∏–∞–ø–∞–∑–æ–Ω–∞ (>= –∂–µ–ª–∞—Ç–µ–ª—å–Ω–æ–π) ‚Äî –±–æ–ª–µ–µ –Ω–∞—Å—ã—â–µ–Ω–Ω—ã–π –∑–µ–ª—ë–Ω—ã–π (–ø—Ä–∏–º–µ–Ω—è–µ—Ç—Å—è –ø–æ–≤–µ—Ä—Ö –æ–±—ã—á–Ω–æ–≥–æ –∑–µ–ª—ë–Ω–æ–≥–æ)
            if desired_price_col_idx:
                desired_price_col_letter = ws.cell(row=1, column=desired_price_col_idx).column_letter
                dark_green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                dark_green_formula = f"AND({current_price_col_letter}2<>\"\", {current_price_col_letter}2>0, {current_price_col_letter}2>={desired_price_col_letter}2)"
                dark_green_rule = FormulaRule(formula=[dark_green_formula], fill=dark_green_fill, stopIfTrue=True)
                ws.conditional_formatting.add(data_range, dark_green_rule)
        
        # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –¥–ª—è —Ü–µ–Ω—ã —Å –∞–∫—Ü–∏—è–º–∏
        marketing_price_col_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == COL_MARKETING_PRICE:
                marketing_price_col_idx = col_idx
                break
        
        if marketing_price_col_idx and min_price_col_idx:
            marketing_price_col_letter = ws.cell(row=1, column=marketing_price_col_idx).column_letter
            
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            green_formula = f"AND({marketing_price_col_letter}2<>\"\", {marketing_price_col_letter}2>0, {marketing_price_col_letter}2>={min_price_col_letter}2)"
            green_rule = FormulaRule(formula=[green_formula], fill=green_fill, stopIfTrue=False)
            
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_formula = f"AND({marketing_price_col_letter}2<>\"\", {marketing_price_col_letter}2>0, {marketing_price_col_letter}2<{min_price_col_letter}2)"
            red_rule = FormulaRule(formula=[red_formula], fill=red_fill, stopIfTrue=False)
            
            data_range = f"{marketing_price_col_letter}2:{marketing_price_col_letter}{len(df) + 1}"
            ws.conditional_formatting.add(data_range, red_rule)
            ws.conditional_formatting.add(data_range, green_rule)
        
        wb.save(costs_path)
        print("‚úÖ –ü—Ä–∏–º–µ–Ω–µ–Ω–æ —É—Å–ª–æ–≤–Ω–æ–µ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –∫ –∫–æ–ª–æ–Ω–∫–∞–º.")
    except Exception as e:
        print(f"‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –ø—Ä–∏–º–µ–Ω–∏—Ç—å —É—Å–ª–æ–≤–Ω–æ–µ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ: {e}")
    
    print("‚úÖ –¢–µ–∫—É—â–∏–µ —Ü–µ–Ω—ã –ø–æ–ª—É—á–µ–Ω—ã –∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã.")
    return True


def action_get_active_actions(repo_root: Path) -> bool:
    """
    –î–µ–π—Å—Ç–≤–∏–µ 4: –£–∑–Ω–∞—Ç—å –∞–∫—Ç–∏–≤–Ω—ã–µ –∞–∫—Ü–∏–∏.
    –ü—Ä–æ—Å–º–æ—Ç—Ä –≤ –∫–∞–∫–∏—Ö –∞–∫—Ü–∏—è—Ö —É—á–∞—Å—Ç–≤—É–µ—Ç —Ç–æ–≤–∞—Ä –∏ –ø–æ –∫–∞–∫–æ–π —Ü–µ–Ω–µ.
    """
    print_step("–£–∑–Ω–∞—Ç—å –∞–∫—Ç–∏–≤–Ω—ã–µ –∞–∫—Ü–∏–∏")
    
    costs_path = repo_root / COSTS_FILENAME
    
    if not costs_path.exists():
        print(f"‚ùå –§–∞–π–ª {COSTS_FILENAME} –Ω–µ –Ω–∞–π–¥–µ–Ω.")
        return False
    
    df, key_col, cost_col = load_costs_df(costs_path)
    print(f"–ó–∞–≥—Ä—É–∂–µ–Ω —Ñ–∞–π–ª —Å–µ–±–µ—Å—Ç–æ–∏–º–æ—Å—Ç–∏: {len(df)} –∑–∞–ø–∏—Å–µ–π.")
    
    # –ü–æ–ª—É—á–∞–µ–º –∞—Ä—Ç–∏–∫—É–ª—ã
    offer_ids_list = []
    for _, row in df.iterrows():
        art = _artikul_normalize(row.get(key_col))
        if art:
            offer_ids_list.append(art)
    
    # –ü–æ–ª—É—á–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ–± –∞–∫—Ü–∏—è—Ö
    actions_map, actions_info_list, _ = get_actions_for_products(offer_ids_list)
    
    if not actions_info_list:
        print("‚ö†Ô∏è –ê–∫—Ç–∏–≤–Ω—ã–µ –∞–∫—Ü–∏–∏ –Ω–µ –Ω–∞–π–¥–µ–Ω—ã.")
        return False
    
    log_verbose(f"–ù–∞–π–¥–µ–Ω–æ –∞–∫—Ü–∏–π: {len(actions_info_list)}")
    
    # –°–æ–∑–¥–∞—ë–º DataFrame –¥–ª—è –ª–∏—Å—Ç–∞ –∞–∫—Ü–∏–π
    actions_df_data = {}
    actions_df_data[key_col] = df[key_col].values
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π —Ü–µ–Ω—ã
    if COL_MIN_PRICE not in df.columns:
        print(f"‚ö†Ô∏è –ö–æ–ª–æ–Ω–∫–∞ ¬´{COL_MIN_PRICE}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω–∞.")
        if prompt_yes_no("–†–∞—Å—Å—á–∏—Ç–∞—Ç—å –æ–ø—Ç–∏–º–∞–ª—å–Ω—ã–µ —Ü–µ–Ω—ã —Å–µ–π—á–∞—Å?", default_yes=True):
            if not action_calculate_optimal_prices(repo_root):
                return False
            df, key_col, cost_col = load_costs_df(costs_path)
        else:
            print("‚ö†Ô∏è –ü—Ä–æ–¥–æ–ª–∂–∞–µ–º –±–µ–∑ –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π —Ü–µ–Ω—ã.")
    
    if COL_MIN_PRICE in df.columns:
        actions_df_data[COL_MIN_PRICE] = df[COL_MIN_PRICE].values
    
    # –ü–æ–ª—É—á–∞–µ–º —Ü–µ–Ω—ã –≤ –∞–∫—Ü–∏—è—Ö –¥–ª—è –∫–∞–∂–¥–æ–≥–æ —Ç–æ–≤–∞—Ä–∞
    action_prices_dicts = {}
    for action_info in actions_info_list:
        action_name = action_info["name"]
        action_prices_dicts[action_name] = []
    
    for _, row in df.iterrows():
        art = _artikul_normalize(row.get(key_col))
        if art:
            art_normalized = _normalize_offer_id(art)
            art_actions = actions_map.get(art) or actions_map.get(art_normalized) or {}
        else:
            art_actions = {}
        
        for action_info in actions_info_list:
            action_name = action_info["name"]
            action_price = art_actions.get(action_name)
            if action_price is not None:
                action_prices_dicts[action_name].append(round(action_price))
            else:
                action_prices_dicts[action_name].append(None)
    
    # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫–∏ –∞–∫—Ü–∏–π
    for action_info in actions_info_list:
        action_name = action_info["name"]
        actions_df_data[action_name] = action_prices_dicts[action_name]
    
    actions_df = pd.DataFrame(actions_df_data)
    
    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤ Excel
    try:
        wb = load_workbook(costs_path)
        
        if 'Sheet1' in wb.sheetnames:
            wb['Sheet1'].title = '–û—Å–Ω–æ–≤–Ω–æ–π'
        
        if "–ê–∫—Ü–∏–∏" in wb.sheetnames:
            wb.remove(wb["–ê–∫—Ü–∏–∏"])
        
        ws_actions = wb.create_sheet("–ê–∫—Ü–∏–∏")
        
        # –ó–∞–ø–∏—Å—ã–≤–∞–µ–º –∑–∞–≥–æ–ª–æ–≤–∫–∏
        for c_idx, col_name in enumerate(actions_df.columns, start=1):
            ws_actions.cell(row=1, column=c_idx, value=col_name)
        
        # –ó–∞–ø–∏—Å—ã–≤–∞–µ–º –¥–∞–Ω–Ω—ã–µ
        for r_idx, row in enumerate(actions_df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws_actions.cell(row=r_idx, column=c_idx, value=value)
        
        # –ü—Ä–∏–º–µ–Ω—è–µ–º —É—Å–ª–æ–≤–Ω–æ–µ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ
        if COL_MIN_PRICE in actions_df.columns:
            min_price_col_idx = None
            for col_idx, col_name in enumerate(actions_df.columns, start=1):
                if col_name == COL_MIN_PRICE:
                    min_price_col_idx = col_idx
                    break
            
            if min_price_col_idx:
                min_price_col_letter = get_column_letter(min_price_col_idx)
                
                for col_idx, col_name in enumerate(actions_df.columns, start=1):
                    if col_name == key_col or col_name == COL_MIN_PRICE:
                        continue
                    
                    action_col_letter = get_column_letter(col_idx)
                    
                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    green_formula = f"AND({action_col_letter}2<>\"\", {action_col_letter}2>0, {action_col_letter}2>={min_price_col_letter}2)"
                    green_rule = FormulaRule(formula=[green_formula], fill=green_fill, stopIfTrue=False)
                    
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    red_formula = f"AND({action_col_letter}2<>\"\", {action_col_letter}2>0, {action_col_letter}2<{min_price_col_letter}2)"
                    red_rule = FormulaRule(formula=[red_formula], fill=red_fill, stopIfTrue=False)
                    
                    data_range = f"{action_col_letter}2:{action_col_letter}{len(actions_df) + 1}"
                    ws_actions.conditional_formatting.add(data_range, red_rule)
                    ws_actions.conditional_formatting.add(data_range, green_rule)
        
        wb.save(costs_path)
        print(f"‚úÖ –°–æ–∑–¥–∞–Ω –ª–∏—Å—Ç ¬´–ê–∫—Ü–∏–∏¬ª —Å {len(actions_info_list)} –∫–æ–ª–æ–Ω–∫–∞–º–∏ –∞–∫—Ü–∏–π.")
    except Exception as e:
        print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ –ª–∏—Å—Ç–∞ ¬´–ê–∫—Ü–∏–∏¬ª: {e}")
        import traceback
        print(traceback.format_exc())
        return False
    
    return True


def action_remove_unprofitable_actions(repo_root: Path) -> bool:
    """
    –î–µ–π—Å—Ç–≤–∏–µ 5: –£–¥–∞–ª–∏—Ç—å –Ω–µ–≤—ã–≥–æ–¥–Ω—ã–µ –∞–∫—Ü–∏–∏.
    –£–¥–∞–ª–µ–Ω–∏–µ —Ç–æ–≤–∞—Ä–æ–≤ –∏–∑ –∞–∫—Ü–∏–π, –≥–¥–µ —Ü–µ–Ω–∞ –º–µ–Ω—å—à–µ –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π.
    """
    print_step("–£–¥–∞–ª–∏—Ç—å –Ω–µ–≤—ã–≥–æ–¥–Ω—ã–µ –∞–∫—Ü–∏–∏")
    
    costs_path = repo_root / COSTS_FILENAME
    
    if not costs_path.exists():
        print(f"‚ùå –§–∞–π–ª {COSTS_FILENAME} –Ω–µ –Ω–∞–π–¥–µ–Ω.")
        return False
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –ª–∏—Å—Ç–∞ "–ê–∫—Ü–∏–∏"
    try:
        wb = load_workbook(costs_path)
        if "–ê–∫—Ü–∏–∏" not in wb.sheetnames:
            print("‚ö†Ô∏è –õ–∏—Å—Ç ¬´–ê–∫—Ü–∏–∏¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω.")
            if prompt_yes_no("–ü–æ–ª—É—á–∏—Ç—å –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ–± –∞–∫—Ç–∏–≤–Ω—ã—Ö –∞–∫—Ü–∏—è—Ö —Å–µ–π—á–∞—Å?", default_yes=True):
                if not action_get_active_actions(repo_root):
                    return False
                wb = load_workbook(costs_path)
            else:
                print("‚ùå –ù–µ–≤–æ–∑–º–æ–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å –∏–∑ –∞–∫—Ü–∏–π –±–µ–∑ –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–∏ –æ–± –∞–∫—Ü–∏—è—Ö.")
                return False
    except Exception as e:
        print(f"‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ –æ—Ç–∫—Ä—ã—Ç–∏–∏ —Ñ–∞–π–ª–∞: {e}")
        return False
    
    df, key_col, cost_col = load_costs_df(costs_path)
    
    # –ü–æ–ª—É—á–∞–µ–º –∞—Ä—Ç–∏–∫—É–ª—ã –∏ –º–∞–ø–ø–∏–Ω–≥
    offer_ids_list = []
    for _, row in df.iterrows():
        art = _artikul_normalize(row.get(key_col))
        if art:
            offer_ids_list.append(art)
    
    _, actions_info_list, offer_id_to_product_id = get_actions_for_products(offer_ids_list)
    
    if not actions_info_list:
        print("‚ö†Ô∏è –ê–∫—Ç–∏–≤–Ω—ã–µ –∞–∫—Ü–∏–∏ –Ω–µ –Ω–∞–π–¥–µ–Ω—ã.")
        return False
    
    ws_actions = wb['–ê–∫—Ü–∏–∏']
    action_name_to_id = {a["name"]: a["id"] for a in actions_info_list}
    
    if not offer_id_to_product_id or not action_name_to_id:
        print("‚ö†Ô∏è –ù–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ –¥–∞–Ω–Ω—ã—Ö –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è –∏–∑ –∞–∫—Ü–∏–π.")
        return False
    
    candidates = collect_deactivation_candidates_from_sheet(
        ws_actions,
        key_col,
        COL_MIN_PRICE,
        action_name_to_id,
        offer_id_to_product_id,
    )
    
    if not candidates:
        print("‚úÖ –¢–æ–≤–∞—Ä—ã —Å —Ü–µ–Ω–æ–π –Ω–∏–∂–µ –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π –Ω–µ –Ω–∞–π–¥–µ–Ω—ã.")
        return True
    
    total_to_remove = sum(len(ids) for ids in candidates.values())
    print(f"–ù–∞–π–¥–µ–Ω–æ {total_to_remove} —Ç–æ–≤–∞—Ä–æ–≤ –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è –∏–∑ {len(candidates)} –∞–∫—Ü–∏–π.")
    
    if not prompt_yes_no("–ü—Ä–æ–¥–æ–ª–∂–∏—Ç—å —É–¥–∞–ª–µ–Ω–∏–µ?", default_yes=False):
        print("‚ùå –£–¥–∞–ª–µ–Ω–∏–µ –æ—Ç–º–µ–Ω–µ–Ω–æ.")
        return False
    
    log_verbose("–£–¥–∞–ª–µ–Ω–∏–µ —Ç–æ–≤–∞—Ä–æ–≤ –∏–∑ –∞–∫—Ü–∏–π...")
    for action_id, product_ids in candidates.items():
        result = deactivate_products_in_action(action_id, product_ids)
        removed = result.get("product_ids", []) or []
        rejected = result.get("rejected", []) or []
        log_verbose(f"–ê–∫—Ü–∏—è {action_id}: —É–¥–∞–ª–µ–Ω–æ {len(removed)}, –Ω–µ —É–¥–∞–ª–µ–Ω–æ {len(rejected)}")
    print("‚úÖ –£–¥–∞–ª–µ–Ω–∏–µ –∏–∑ –∞–∫—Ü–∏–π –∑–∞–≤–µ—Ä—à–µ–Ω–æ.")
    return True


def action_add_to_actions(repo_root: Path) -> bool:
    """
    –î–µ–π—Å—Ç–≤–∏–µ 6: –î–æ–±–∞–≤–∏—Ç—å —Ç–æ–≤–∞—Ä—ã –≤ –∞–∫—Ü–∏–∏.
    –î–æ–±–∞–≤–ª–µ–Ω–∏–µ —Ç–æ–≤–∞—Ä–æ–≤, –µ—Å–ª–∏ –¥–æ–ø—É—Å—Ç–∏–º–∞—è —Ü–µ–Ω–∞ –≤ –¥–∏–∞–ø–∞–∑–æ–Ω–µ –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π –∏ –∂–µ–ª–∞–µ–º–æ–π.
    """
    print_step("–î–æ–±–∞–≤–∏—Ç—å —Ç–æ–≤–∞—Ä—ã –≤ –∞–∫—Ü–∏–∏")
    
    costs_path = repo_root / COSTS_FILENAME
    
    if not costs_path.exists():
        print(f"‚ùå –§–∞–π–ª {COSTS_FILENAME} –Ω–µ –Ω–∞–π–¥–µ–Ω.")
        return False
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π –∏ –∂–µ–ª–∞—Ç–µ–ª—å–Ω–æ–π —Ü–µ–Ω—ã
    df, key_col, cost_col = load_costs_df(costs_path)
    
    if COL_MIN_PRICE not in df.columns or COL_DESIRED_PRICE not in df.columns:
        print(f"‚ö†Ô∏è –ö–æ–ª–æ–Ω–∫–∏ ¬´{COL_MIN_PRICE}¬ª –∏–ª–∏ ¬´{COL_DESIRED_PRICE}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω—ã.")
        if prompt_yes_no("–†–∞—Å—Å—á–∏—Ç–∞—Ç—å –æ–ø—Ç–∏–º–∞–ª—å–Ω—ã–µ —Ü–µ–Ω—ã —Å–µ–π—á–∞—Å?", default_yes=True):
            if not action_calculate_optimal_prices(repo_root):
                return False
            df, key_col, cost_col = load_costs_df(costs_path)
        else:
            print("‚ùå –ù–µ–≤–æ–∑–º–æ–∂–Ω–æ –¥–æ–±–∞–≤–∏—Ç—å –≤ –∞–∫—Ü–∏–∏ –±–µ–∑ –æ–ø—Ç–∏–º–∞–ª—å–Ω—ã—Ö —Ü–µ–Ω.")
            return False
    
    # –ü–æ–ª—É—á–∞–µ–º –∞—Ä—Ç–∏–∫—É–ª—ã –∏ –º–∞–ø–ø–∏–Ω–≥
    offer_ids_list = []
    for _, row in df.iterrows():
        art = _artikul_normalize(row.get(key_col))
        if art:
            offer_ids_list.append(art)
    
    _, actions_info_list, offer_id_to_product_id = get_actions_for_products(offer_ids_list)
    
    if not actions_info_list:
        print("‚ö†Ô∏è –ê–∫—Ç–∏–≤–Ω—ã–µ –∞–∫—Ü–∏–∏ –Ω–µ –Ω–∞–π–¥–µ–Ω—ã.")
        return False
    
    # –°–æ–∑–¥–∞—ë–º –º–∞–ø–ø–∏–Ω–≥ product_id -> offer_id
    product_id_to_offer_id = {}
    for offer_id, product_id in offer_id_to_product_id.items():
        product_id_to_offer_id[product_id] = offer_id
    
    offer_ids_set_for_candidates = set()
    for oid in offer_ids_list:
        normalized = _normalize_offer_id(oid)
        if normalized:
            offer_ids_set_for_candidates.add(normalized)
            offer_ids_set_for_candidates.add(oid)
    
    # –°–æ–∑–¥–∞—ë–º –º–∞–ø–ø–∏–Ω–≥ product_id -> (min_price, desired_price)
    product_id_to_prices: Dict[int, Tuple[float, float]] = {}
    for _, row in df.iterrows():
        art = _artikul_normalize(row.get(key_col))
        if not art:
            continue
        
        art_normalized = _normalize_offer_id(art)
        product_id = offer_id_to_product_id.get(art_normalized) or offer_id_to_product_id.get(art)
        
        if product_id:
            min_p = row.get(COL_MIN_PRICE)
            des_p = row.get(COL_DESIRED_PRICE)
            if min_p is not None and des_p is not None:
                try:
                    min_price_val = float(min_p)
                    des_price_val = float(des_p)
                    if min_price_val > 0 and des_price_val > 0:
                        product_id_to_prices[product_id] = (min_price_val, des_price_val)
                except (TypeError, ValueError):
                    pass
    
    if not product_id_to_prices:
        print("‚ö†Ô∏è –ù–µ –Ω–∞–π–¥–µ–Ω–æ —Ç–æ–≤–∞—Ä–æ–≤ —Å —Ä–∞—Å—Å—á–∏—Ç–∞–Ω–Ω—ã–º–∏ —Ü–µ–Ω–∞–º–∏.")
        return False
    
    print(f"‚úÖ –¢–æ–≤–∞—Ä–æ–≤ —Å —Ü–µ–Ω–∞–º–∏: {len(product_id_to_prices)}, –∞–∫—Ü–∏–π: {len(actions_info_list)}")
    if not prompt_yes_no("–ü—Ä–æ–¥–æ–ª–∂–∏—Ç—å –¥–æ–±–∞–≤–ª–µ–Ω–∏–µ —Ç–æ–≤–∞—Ä–æ–≤ –≤ –∞–∫—Ü–∏–∏?", default_yes=False):
        print("‚ùå –î–æ–±–∞–≤–ª–µ–Ω–∏–µ –æ—Ç–º–µ–Ω–µ–Ω–æ.")
        return False
    
    log_verbose("–ü—Ä–æ–≤–µ—Ä–∫–∞ –∫–∞–Ω–¥–∏–¥–∞—Ç–æ–≤ –¥–ª—è –¥–æ–±–∞–≤–ª–µ–Ω–∏—è –≤ –∞–∫—Ü–∏–∏...")
    total_added = 0
    for action_info in actions_info_list:
        action_id = action_info["id"]
        action_name = action_info["name"]
        candidates = get_action_candidates(action_id, product_id_to_offer_id, offer_ids_set_for_candidates)
        if not candidates:
            continue
        products_to_add = []
        for product_id, product_info in candidates.items():
            if product_id not in product_id_to_prices:
                continue
            
            min_price, desired_price = product_id_to_prices[product_id]
            
            max_action_price = product_info.get("max_action_price")
            if max_action_price is None:
                continue
            
            try:
                max_action_price_val = float(max_action_price)
            except (TypeError, ValueError):
                continue
            
            target_price = min(desired_price, max_action_price_val)
            
            if target_price >= min_price:
                current_action_price = product_info.get("action_price", 0)
                if current_action_price == 0 or current_action_price is None:
                    stock = product_info.get("stock", 0) or 0
                    products_to_add.append({
                        "product_id": product_id,
                        "action_price": int(target_price),
                        "stock": int(stock) if stock else 0
                    })
        
        if products_to_add:
            result = activate_products_in_action(action_id, products_to_add)
            added = result.get("product_ids", []) or []
            rejected = result.get("rejected", []) or []
            total_added += len(added)
            log_verbose(f"–ê–∫—Ü–∏—è {action_name}: –¥–æ–±–∞–≤–ª–µ–Ω–æ {len(added)}, –Ω–µ –¥–æ–±–∞–≤–ª–µ–Ω–æ {len(rejected)}")
    print(f"‚úÖ –î–æ–±–∞–≤–ª–µ–Ω–∏–µ –≤ –∞–∫—Ü–∏–∏ –∑–∞–≤–µ—Ä—à–µ–Ω–æ. –í—Å–µ–≥–æ –¥–æ–±–∞–≤–ª–µ–Ω–æ: {total_added} —Ç–æ–≤–∞—Ä–æ–≤.")
    return True


def action_process_discount_requests(repo_root: Path) -> bool:
    """
    –î–µ–π—Å—Ç–≤–∏–µ 7: –û–±—Ä–∞–±–æ—Ç–∞—Ç—å –∑–∞—è–≤–∫–∏ –Ω–∞ —Å–∫–∏–¥–∫—É.
    –û–¥–æ–±—Ä—è–µ—Ç –∑–∞—è–≤–∫–∏, –µ—Å–ª–∏ –∑–∞—è–≤–ª–µ–Ω–Ω–∞—è —Ü–µ–Ω–∞ >= –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π —Ü–µ–Ω—ã –ø—Ä–æ–¥–∞–∂–∏, –∏–Ω–∞—á–µ –æ—Ç–∫–ª–æ–Ω—è–µ—Ç.
    """
    print_step("–û–±—Ä–∞–±–æ—Ç–∞—Ç—å –∑–∞—è–≤–∫–∏ –Ω–∞ —Å–∫–∏–¥–∫—É")
    
    costs_path = repo_root / COSTS_FILENAME
    
    if not costs_path.exists():
        print(f"‚ùå –§–∞–π–ª {COSTS_FILENAME} –Ω–µ –Ω–∞–π–¥–µ–Ω.")
        return False
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –∫–æ–ª–æ–Ω–∫–∏ –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π —Ü–µ–Ω—ã
    df, key_col, cost_col = load_costs_df(costs_path)
    
    if COL_MIN_PRICE not in df.columns:
        print(f"‚ö†Ô∏è –ö–æ–ª–æ–Ω–∫–∞ ¬´{COL_MIN_PRICE}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω–∞.")
        if prompt_yes_no("–†–∞—Å—Å—á–∏—Ç–∞—Ç—å –æ–ø—Ç–∏–º–∞–ª—å–Ω—ã–µ —Ü–µ–Ω—ã —Å–µ–π—á–∞—Å?", default_yes=True):
            if not action_calculate_optimal_prices(repo_root):
                return False
            df, key_col, cost_col = load_costs_df(costs_path)
        else:
            print("‚ùå –ù–µ–≤–æ–∑–º–æ–∂–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞—Ç—å –∑–∞—è–≤–∫–∏ –±–µ–∑ –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π —Ü–µ–Ω—ã.")
            return False
    
    # –°–æ–∑–¥–∞—ë–º –º–∞–ø–ø–∏–Ω–≥ offer_id -> min_price
    offer_id_to_min_price: Dict[str, float] = {}
    for _, row in df.iterrows():
        art = _artikul_normalize(row.get(key_col))
        if not art:
            continue
        
        art_normalized = _normalize_offer_id(art)
        min_price = row.get(COL_MIN_PRICE)
        
        if min_price is not None:
            try:
                min_price_val = float(min_price)
                if min_price_val > 0:
                    offer_id_to_min_price[art_normalized] = min_price_val
                    # –¢–∞–∫–∂–µ –¥–æ–±–∞–≤–ª—è–µ–º –∏—Å—Ö–æ–¥–Ω—ã–π –∞—Ä—Ç–∏–∫—É–ª
                    if art != art_normalized:
                        offer_id_to_min_price[art] = min_price_val
            except (TypeError, ValueError):
                pass
    
    if not offer_id_to_min_price:
        print("‚ö†Ô∏è –ù–µ –Ω–∞–π–¥–µ–Ω–æ —Ç–æ–≤–∞—Ä–æ–≤ —Å –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π —Ü–µ–Ω–æ–π.")
        return False
    
    print(f"‚úÖ –ó–∞–≥—Ä—É–∂–µ–Ω–æ {len(offer_id_to_min_price)} —Ç–æ–≤–∞—Ä–æ–≤ —Å –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π —Ü–µ–Ω–æ–π.")
    
    # –ü–æ–ª—É—á–∞–µ–º –∑–∞—è–≤–∫–∏ –Ω–∞ —Å–∫–∏–¥–∫—É
    print("üì° –ü–æ–ª—É—á–µ–Ω–∏–µ –∑–∞—è–≤–æ–∫ –Ω–∞ —Å–∫–∏–¥–∫—É...")
    discount_tasks = get_discount_requests(status="NEW", limit=50)
    
    if not discount_tasks:
        print("‚úÖ –ù–æ–≤—ã—Ö –∑–∞—è–≤–æ–∫ –Ω–∞ —Å–∫–∏–¥–∫—É –Ω–µ –Ω–∞–π–¥–µ–Ω–æ.")
        return True
    
    print(f"‚úÖ –ù–∞–π–¥–µ–Ω–æ {len(discount_tasks)} –∑–∞—è–≤–æ–∫ –Ω–∞ —Å–∫–∏–¥–∫—É.")
    
    # Ozon –≤ –∑–∞—è–≤–∫–∞—Ö –≤–æ–∑–≤—Ä–∞—â–∞–µ—Ç SKU (product_id), –≤ costs.xlsx –∑–∞–ø–∏—Å–∞–Ω offer_id (–∞—Ä—Ç–∏–∫—É–ª).
    # –ü–æ–ª—É—á–∞–µ–º –º–∞–ø–ø–∏–Ω–≥ SKU -> offer_id —á–µ—Ä–µ–∑ API.
    skus_from_tasks = []
    for task in discount_tasks:
        sku = task.get("sku")
        if sku is not None:
            try:
                skus_from_tasks.append(int(sku))
            except (TypeError, ValueError):
                pass
    skus_unique = list(dict.fromkeys(skus_from_tasks))
    sku_to_offer_id: Dict[int, str] = {}
    if skus_unique:
        sku_to_offer_id = get_sku_to_offer_id_mapping(skus_unique)
    
    log_verbose(f"–û–±—Ä–∞–±–æ—Ç–∫–∞ {len(discount_tasks)} –∑–∞—è–≤–æ–∫...")
    
    tasks_to_approve = []
    tasks_to_decline = []
    task_id_to_sku: Dict[str, str] = {}
    
    for task in discount_tasks:
        sku = task.get("sku")
        task_id = task.get("id")
        if task_id is not None:
            task_id_to_sku[str(task_id)] = str(sku) if sku else "‚Äî"
        
        if not sku:
            continue
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º offer_id: —Å–Ω–∞—á–∞–ª–∞ –ø–æ –º–∞–ø–ø–∏–Ω–≥—É SKU -> offer_id, –∏–Ω–∞—á–µ —Å—á–∏—Ç–∞–µ–º sku –∞—Ä—Ç–∏–∫—É–ª–æ–º (offer_id)
        offer_id_raw = None
        try:
            sku_int = int(sku)
            offer_id_raw = sku_to_offer_id.get(sku_int)
        except (TypeError, ValueError):
            pass
        if not offer_id_raw:
            offer_id_raw = str(sku)
        
        offer_id_normalized = _normalize_offer_id(offer_id_raw)
        min_price = offer_id_to_min_price.get(offer_id_normalized) or offer_id_to_min_price.get(offer_id_raw)
        
        if min_price is None:
            # –ï—Å–ª–∏ –Ω–µ –Ω–∞–π–¥–µ–Ω–∞ –º–∏–Ω–∏–º–∞–ª—å–Ω–∞—è —Ü–µ–Ω–∞, –æ—Ç–∫–ª–æ–Ω—è–µ–º
            reason = "–ú–∏–Ω–∏–º–∞–ª—å–Ω–∞—è —Ü–µ–Ω–∞ –Ω–µ —Ä–∞—Å—Å—á–∏—Ç–∞–Ω–∞"
            tasks_to_decline.append({
                "id": task_id,
                "seller_comment": reason
            })
            continue
        
        requested_price = task.get("requested_price")
        if requested_price is None:
            # –ï—Å–ª–∏ –Ω–µ—Ç –∑–∞–ø—Ä–æ—à–µ–Ω–Ω–æ–π —Ü–µ–Ω—ã, –æ—Ç–∫–ª–æ–Ω—è–µ–º
            reason = "–ó–∞–ø—Ä–æ—à–µ–Ω–Ω–∞—è —Ü–µ–Ω–∞ –Ω–µ —É–∫–∞–∑–∞–Ω–∞"
            tasks_to_decline.append({
                "id": task_id,
                "seller_comment": reason
            })
            continue
        
        try:
            requested_price_val = float(requested_price)
        except (TypeError, ValueError):
            reason = "–ù–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω–∞—è –∑–∞–ø—Ä–æ—à–µ–Ω–Ω–∞—è —Ü–µ–Ω–∞"
            tasks_to_decline.append({
                "id": task_id,
                "seller_comment": reason
            })
            continue
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º —É—Å–ª–æ–≤–∏–µ: –æ–¥–æ–±—Ä—è–µ–º, –µ—Å–ª–∏ requested_price >= min_price
        if requested_price_val >= min_price:
            # –û–¥–æ–±—Ä—è–µ–º –∑–∞—è–≤–∫—É; –ø—Ä–∏—á–∏–Ω–∞ –¥–ª—è –æ—Ç–æ–±—Ä–∞–∂–µ–Ω–∏—è –∏ –¥–ª—è API (seller_comment)
            reason = f"–û–¥–æ–±—Ä–µ–Ω–æ: –∑–∞–ø—Ä–æ—à–µ–Ω–Ω–∞—è —Ü–µ–Ω–∞ {requested_price_val:.0f} ‚ÇΩ –Ω–µ –Ω–∏–∂–µ –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π {min_price:.0f} ‚ÇΩ"
            # API —Ç—Ä–µ–±—É–µ—Ç approved_quantity_min > 0
            q_min = task.get("requested_quantity_min")
            try:
                q_min = max(1, int(q_min)) if q_min is not None else 1
            except (TypeError, ValueError):
                q_min = 1
            q_max = task.get("requested_quantity_max")
            try:
                q_max = max(q_min, int(q_max)) if q_max is not None else q_min
            except (TypeError, ValueError):
                q_max = q_min
            approved_task = {
                "id": task_id,
                "approved_price": int(requested_price_val),
                "approved_quantity_min": q_min,
                "approved_quantity_max": q_max,
                "seller_comment": reason
            }
            tasks_to_approve.append(approved_task)
        else:
            # –û—Ç–∫–ª–æ–Ω—è–µ–º –∑–∞—è–≤–∫—É
            reason = f"–û—Ç–∫–ª–æ–Ω–µ–Ω–æ: –∑–∞–ø—Ä–æ—à–µ–Ω–Ω–∞—è —Ü–µ–Ω–∞ {requested_price_val:.2f} ‚ÇΩ –Ω–∏–∂–µ –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π {min_price:.2f} ‚ÇΩ"
            tasks_to_decline.append({
                "id": task_id,
                "seller_comment": reason
            })
    
    print(f"–û–¥–æ–±—Ä–∏—Ç—å: {len(tasks_to_approve)}, –æ—Ç–∫–ª–æ–Ω–∏—Ç—å: {len(tasks_to_decline)}")
    if tasks_to_approve or tasks_to_decline:
        for t in tasks_to_approve:
            sid = str(t.get("id", ""))
            sku_display = task_id_to_sku.get(sid, "‚Äî")
            print(f"   ‚úÖ –ó–∞—è–≤–∫–∞ {sid} (–∞—Ä—Ç–∏–∫—É–ª {sku_display}): {t.get('seller_comment', '–û–¥–æ–±—Ä–µ–Ω–æ')}")
        for t in tasks_to_decline:
            sid = str(t.get("id", ""))
            sku_display = task_id_to_sku.get(sid, "‚Äî")
            print(f"   ‚ùå –ó–∞—è–≤–∫–∞ {sid} (–∞—Ä—Ç–∏–∫—É–ª {sku_display}): {t.get('seller_comment', '–û—Ç–∫–ª–æ–Ω–µ–Ω–æ')}")
    
    if not tasks_to_approve and not tasks_to_decline:
        print("‚ö†Ô∏è –ù–µ—Ç –∑–∞—è–≤–æ–∫ –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏.")
        return True
    
    if not prompt_yes_no("–ü—Ä–æ–¥–æ–ª–∂–∏—Ç—å –æ–±—Ä–∞–±–æ—Ç–∫—É –∑–∞—è–≤–æ–∫?", default_yes=False):
        print("‚ùå –û–±—Ä–∞–±–æ—Ç–∫–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞.")
        return False
    
    # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º –∑–∞—è–≤–∫–∏
    log_verbose("–û–±—Ä–∞–±–æ—Ç–∫–∞ –∑–∞—è–≤–æ–∫...")
    
    if tasks_to_approve:
        approve_result = approve_discount_requests(tasks_to_approve)
        ok, fail = approve_result.get('success_count', 0), approve_result.get('fail_count', 0)
        print(f"‚úÖ –û–¥–æ–±—Ä–µ–Ω–æ: {ok}" + (f", –æ—à–∏–±–æ–∫: {fail}" if fail else ""))
        if fail:
            for detail in (approve_result.get('fail_details') or [])[:3]:
                print(f"   –ó–∞—è–≤–∫–∞ {detail.get('task_id')}: {detail.get('error_for_user', '?')}")
    if tasks_to_decline:
        decline_result = decline_discount_requests(tasks_to_decline)
        ok, fail = decline_result.get('success_count', 0), decline_result.get('fail_count', 0)
        print(f"‚ùå –û—Ç–∫–ª–æ–Ω–µ–Ω–æ: {ok}" + (f", –æ—à–∏–±–æ–∫: {fail}" if fail else ""))
        if fail:
            for detail in (decline_result.get('fail_details') or [])[:3]:
                print(f"   –ó–∞—è–≤–∫–∞ {detail.get('task_id')}: {detail.get('error_for_user', '?')}")
    
    print("‚úÖ –û–±—Ä–∞–±–æ—Ç–∫–∞ –∑–∞—è–≤–æ–∫ –∑–∞–≤–µ—Ä—à–µ–Ω–∞.")
    return True


def show_price_management_menu(repo_root: Path):
    """
    –ü–æ–∫–∞–∑—ã–≤–∞–µ—Ç –º–µ–Ω—é —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è —Ü–µ–Ω–æ–π –∏ –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ—Ç –≤—ã–±–æ—Ä –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è.
    """
    while True:
        print_step("–£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ —Ü–µ–Ω–æ–π")
        print("1. –î–∏–∞–ø–∞–∑–æ–Ω —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç–∏")
        print("2. –†–∞—Å—Å—á–∏—Ç–∞—Ç—å –æ–ø—Ç–∏–º–∞–ª—å–Ω—É—é —Ü–µ–Ω—É")
        print("3. –£–∑–Ω–∞—Ç—å —Ç–µ–∫—É—â—É—é —Ü–µ–Ω—É –ø—Ä–æ–¥–∞–∂–∏")
        print("4. –£–∑–Ω–∞—Ç—å –∞–∫—Ç–∏–≤–Ω—ã–µ –∞–∫—Ü–∏–∏")
        print("5. –£–¥–∞–ª–∏—Ç—å –Ω–µ–≤—ã–≥–æ–¥–Ω—ã–µ –∞–∫—Ü–∏–∏")
        print("6. –î–æ–±–∞–≤–∏—Ç—å —Ç–æ–≤–∞—Ä—ã –≤ –∞–∫—Ü–∏–∏")
        print("7. –û–±—Ä–∞–±–æ—Ç–∞—Ç—å –∑–∞—è–≤–∫–∏ –Ω–∞ —Å–∫–∏–¥–∫—É")
        print("8. –ù–∞–∑–∞–¥ –≤ –≥–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é")
        
        choice = input("–í—ã–±–µ—Ä–∏—Ç–µ –æ–ø—Ü–∏—é (1-8): ").strip()
        
        if choice == "1":
            action_set_margin_range(repo_root)
        elif choice == "2":
            action_calculate_optimal_prices(repo_root)
        elif choice == "3":
            action_get_current_prices(repo_root)
        elif choice == "4":
            action_get_active_actions(repo_root)
        elif choice == "5":
            action_remove_unprofitable_actions(repo_root)
        elif choice == "6":
            action_add_to_actions(repo_root)
        elif choice == "7":
            action_process_discount_requests(repo_root)
        elif choice == "8":
            break
        else:
            print("–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤—ã–±–µ—Ä–∏—Ç–µ –∫–æ—Ä—Ä–µ–∫—Ç–Ω—É—é –æ–ø—Ü–∏—é (1-8).")
        
        print()  # –ü—É—Å—Ç–∞—è —Å—Ç—Ä–æ–∫–∞ –¥–ª—è —á–∏—Ç–∞–µ–º–æ—Å—Ç–∏


def main():
    """–¢–æ—á–∫–∞ –≤—Ö–æ–¥–∞ –¥–ª—è –∑–∞–ø—É—Å–∫–∞ –º–æ–¥—É–ª—è –∫–∞–∫ —Å–∫—Ä–∏–ø—Ç–∞."""
    script_dir = Path(__file__).resolve().parent
    repo_root = script_dir.parent
    
    show_price_management_menu(repo_root)


if __name__ == "__main__":
    main()
