import os
import json
import time
import shutil
import importlib
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from decimal import Decimal
from typing import List, Dict, Any, Optional

import pandas as pd
import requests
from dateutil.parser import isoparse
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from dotenv import load_dotenv

# 🔐 Загружаем данные для авторизации из переменных окружения
load_dotenv()
CLIENT_ID = os.getenv('OZON_CLIENT_ID')
API_KEY = os.getenv('OZON_API_KEY')

if not CLIENT_ID or not API_KEY:
    raise RuntimeError("Отсутствуют переменные OZON_CLIENT_ID или OZON_API_KEY. Укажите их в .env или окружении.")

HEADERS = {
    'Client-Id': CLIENT_ID,
    'Api-Key': API_KEY,
    'Content-Type': 'application/json'
}

def create_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=5,
        backoff_factor=0.5,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(["POST", "GET"]),
        respect_retry_after_header=True,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=20)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session

def get_custom_date_range():
    while True:
        try:
            month = int(input("Введите номер месяца (1–12): ").strip())
            year = int(input("Введите год (например, 2025): ").strip())

            if 1 <= month <= 12 and 2000 <= year <= 2100:
                break
            else:
                print("⚠️ Введите корректный месяц (1–12) и год (2000–2100).")
        except ValueError:
            print("❌ Некорректный ввод. Попробуйте снова.")

    from datetime import datetime, timedelta
    from calendar import monthrange

    first_day = datetime(year, month, 1)
    last_day = datetime(year, month, monthrange(year, month)[1])
    date_from = first_day.strftime('%Y-%m-%dT00:00:00Z')
    date_to = last_day.strftime('%Y-%m-%dT23:59:59Z')
    return date_from, date_to, month, year



def _normalize_articul_key(s: str) -> str:
    """Приводит артикул к одному виду для сопоставления (Excel даёт 12345.0, API — 12345)."""
    if not s or not isinstance(s, str):
        return (s or "").strip()
    s = s.strip()
    if s.lower() == 'nan':
        return ""
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return s
    except (ValueError, TypeError):
        return s


# 📄 Загрузка карты себестоимости из внешнего файла
def load_cost_map():
    script_dir = os.path.dirname(__file__)
    repo_root = os.path.abspath(os.path.join(script_dir, '..'))

    candidates = [
        os.path.join(repo_root, 'costs.xlsx'),
    ]

    for path in candidates:
        if os.path.exists(path):
            try:
                if path.endswith('.xlsx'):
                    df = pd.read_excel(path)
                else:
                    df = pd.read_csv(path)

                # Нормализуем имена столбцов
                lower_cols = {c.lower(): c for c in df.columns}
                # Поддерживаемые варианты названий
                key_col = None
                cost_col = None

                for variant in ['prefix', 'префикс', 'код', 'артикул', 'offer_id']:
                    if variant in lower_cols:
                        key_col = lower_cols[variant]
                        break
                for variant in ['cost', 'себестоимость', 'цена', 'стоимость']:
                    if variant in lower_cols:
                        cost_col = lower_cols[variant]
                        break

                if not key_col or not cost_col:
                    print(f"⚠️ Файл {os.path.basename(path)} найден, но столбцы не распознаны. \nОжидаются столбцы: 'prefix'/'префикс'/'код'/'артикул' и 'cost'/'себестоимость'.")
                    continue

                mapping = {}
                for _, row in df.iterrows():
                    raw = row.get(key_col, '')
                    key = _normalize_articul_key(str(raw).strip() if raw is not None else '')
                    if not key:
                        continue
                    try:
                        value = float(row.get(cost_col, 0) or 0)
                    except Exception:
                        continue
                    mapping[key] = value

                print(f"🧾 Загружена карта себестоимости из {os.path.basename(path)}: {len(mapping)} записей")
                return mapping
            except Exception as e:
                print(f"⚠️ Не удалось прочитать {os.path.basename(path)}: {e}")

    print("ℹ️ Файл себестоимости не найден (costs.xlsx). Будет использовано значение 0.")
    return {}

# Импорт функций для работы с Performance API
try:
    from scripts.performance_api import get_cpc_campaigns_for_month, get_campaigns_data_for_excel  # type: ignore
except ImportError:
    import sys
    from pathlib import Path
    sys.path.append(str(Path(__file__).resolve().parent))
    from performance_api import get_cpc_campaigns_for_month, get_campaigns_data_for_excel  # type: ignore

# 📥 Получаем список заказов FBS (Fulfillment by Seller)
def _fetch_fbs_page(session: requests.Session, date_from: str, date_to: str, status: str, limit: int, offset: int) -> List[Dict[str, Any]]:
    url = 'https://api-seller.ozon.ru/v3/posting/fbs/list'
    payload = {
        "filter": {
            "since": date_from,
            "to": date_to,
            "status": status
        },
        "limit": limit,
        "offset": offset,
        "with": {
            "analytics_data": True,
            "financial_data": True
        }
    }
    resp = session.post(url, headers=HEADERS, json=payload)
    resp.raise_for_status()
    data = resp.json()
    postings = data.get("result", {}).get("postings", [])
    for p in postings:
        p["__schema"] = "FBS"
    return postings

# 📥 Получаем список заказов FBS (Fulfillment by Seller)
def get_orders(date_from, date_to, session: Optional[requests.Session] = None):
    url = 'https://api-seller.ozon.ru/v3/posting/fbs/list'
    result = []
    limit = 100
    session = session or create_session()

    # Статусы заказов, которые необходимо получить
    STATUSES = ["awaiting_packaging", "awaiting_deliver", "delivering", "delivered", "cancelled"]

    for status in STATUSES:
        offset = 0
        max_workers = 8
        while True:
            # Пакетная параллельная выборка страниц
            futures = {}
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                for i in range(max_workers):
                    page_offset = offset + i * limit
                    futures[executor.submit(_fetch_fbs_page, session, date_from, date_to, status, limit, page_offset)] = page_offset
                empty_hit = False
                # Сохраняем результаты по возрастанию offset
                page_results = []
                for fut in as_completed(futures):
                    page_offset = futures[fut]
                    try:
                        postings = fut.result()
                    except Exception as e:
                        # В случае ошибки прекращаем пакет
                        postings = []
                    page_results.append((page_offset, postings))
                page_results.sort(key=lambda x: x[0])
                for _, postings in page_results:
                    if not postings:
                        empty_hit = True
                        break
                    result.extend(postings)
            if empty_hit:
                break
            offset += max_workers * limit

    return result

# 📥 Получаем список заказов FBO (Fulfillment by Ozon)
def _fetch_fbo_page(session: requests.Session, date_from: str, date_to: str, status: str, limit: int, offset: int) -> List[Dict[str, Any]]:
    url = 'https://api-seller.ozon.ru/v2/posting/fbo/list'
    payload = {
        "dir": "ASC",
        "filter": {
            "since": date_from,
            "to": date_to,
            "status": status
        },
        "limit": limit,
        "offset": offset,
        "with": {
            "analytics_data": True,
            "financial_data": True
        }
    }
    resp = session.post(url, headers=HEADERS, json=payload)
    resp.raise_for_status()
    data = resp.json()
    if isinstance(data, list):
        postings = data
    elif isinstance(data, dict) and "result" in data:
        postings = data["result"]
    else:
        postings = []
    for p in postings:
        if isinstance(p, dict):
            p["__schema"] = "FBO"
    return postings

def get_fbo_orders(date_from, date_to, session: Optional[requests.Session] = None):

    url = 'https://api-seller.ozon.ru/v2/posting/fbo/list'
    result = []
    limit = 100
    session = session or create_session()

    STATUSES = ["awaiting_deliver", "delivering", "delivered", "cancelled"]

    for status in STATUSES:
        offset = 0
        max_workers = 8
        while True:
            futures = {}
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                for i in range(max_workers):
                    page_offset = offset + i * limit
                    futures[executor.submit(_fetch_fbo_page, session, date_from, date_to, status, limit, page_offset)] = page_offset
                empty_hit = False
                page_results = []
                for fut in as_completed(futures):
                    page_offset = futures[fut]
                    try:
                        postings = fut.result()
                    except Exception:
                        postings = []
                    page_results.append((page_offset, postings))
                page_results.sort(key=lambda x: x[0])
                for _, postings in page_results:
                    if not postings:
                        empty_hit = True
                        break
                    result.extend(postings)
            if empty_hit:
                break
            offset += max_workers * limit

    return result

# 💳 Получаем финансовые транзакции по заказу
def get_transactions(posting_number, date_from, date_to, session: Optional[requests.Session] = None):
    url = "https://api-seller.ozon.ru/v3/finance/transaction/list"
    session = session or create_session()

    payload = {
        "filter": {
            "date": {
                "from": date_from,
                "to": date_to
            },
            "posting_number": posting_number
        },
        "page_size": 100,
        "page": 1
    }

    all_operations = []

    while True:
        response = session.post(url, headers=HEADERS, json=payload)

        if response.status_code != 200:
            print(f"❌ Ошибка при запросе транзакций для {posting_number}: {response.status_code}")
            print(response.text)
            return []

        data = response.json().get("result", {})
        operations = data.get("operations", [])
        all_operations.extend(operations)

        if len(operations) < payload["page_size"]:
            break
        payload["page"] += 1

    return all_operations

# 📊 Преобразуем данные в Excel
def _ensure_reports_dir_and_check_space(reports_dir: str, min_free_mb: int = 20) -> None:
    os.makedirs(reports_dir, exist_ok=True)
    try:
        usage = shutil.disk_usage(reports_dir)
        free_mb = usage.free // (1024 * 1024)
        if free_mb < min_free_mb:
            raise RuntimeError(f"Недостаточно места на диске: доступно {free_mb} МБ, требуется ≥ {min_free_mb} МБ")
    except Exception:
        # Если не удалось определить, продолжаем без жёсткой блокировки
        pass

def _artikul_to_number(v):
    """Преобразует значение артикула в число, если возможно; иначе возвращает как есть."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return v
    s = str(v).strip()
    if not s:
        return v
    try:
        n = float(s.replace(",", "."))
        return int(n) if n == int(n) else n
    except (ValueError, TypeError):
        return v


def _safe_save_excel(df: pd.DataFrame, output_file: str, sheet_name: str = "Sheet1") -> str:
    # Пишем во временный файл и затем атомарно заменяем
    base_dir = os.path.dirname(output_file)
    tmp_path = os.path.join(base_dir, f"~tmp_{int(time.time())}.xlsx")
    try:
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            if "Артикул" in df.columns:
                col_idx = list(df.columns).index("Артикул") + 1
                ws = writer.sheets[sheet_name]
                for row in range(2, len(df) + 2):
                    ws.cell(row=row, column=col_idx).number_format = "0"
        # Пытаемся заменить целевой файл
        try:
            if os.path.exists(output_file):
                os.remove(output_file)
        except PermissionError:
            raise RuntimeError(f"Файл занят другим процессом: {output_file}. Закройте его и повторите.")
        os.replace(tmp_path, output_file)
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass
    return output_file

# 📊 Преобразуем данные в Excel
def to_excel(postings, date_from, date_to, month, year, output_file=None, session: Optional[requests.Session] = None):
    from datetime import datetime
    import pandas as pd
    session = session or create_session()

    rows = []
    total_posts = max(len(postings or []), 1)

    # Название месяца на русском в родительном падеже (Сентябрь → сентября)
    months = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    ]
    month_name = months[month-1]

    # Формируем путь и имя файла в папке ../reports относительно этого скрипта
    if not output_file:
        script_dir = os.path.dirname(__file__)
        reports_dir = os.path.abspath(os.path.join(script_dir, '..', 'reports'))
        _ensure_reports_dir_and_check_space(reports_dir)
        output_file = os.path.join(reports_dir, f"{month_name} {year}.xlsx")


    # карта себестоимости: ключ может быть точным offer_id или префиксом
    cost_map = load_cost_map()

    for idx, post in enumerate(postings, start=1):
        posting_number = post.get("posting_number", "")
        status = post.get("status", "")                         # Статус
        schema = post.get("__schema", "")
        
        # Дата отгрузки - для FBS используется shipment_date, для FBO может быть другое поле
        if schema == "FBO":
            # Для FBO заказов пробуем различные поля с датами (в порядке приоритета)
            date = (post.get("in_process_at") or 
                   post.get("shipment_date") or 
                   post.get("created_at") or 
                   post.get("date") or
                   post.get("in_process_at_date") or
                   post.get("shipment_date_time") or "")
        else:
            # Для FBS используем shipment_date
            date = post.get("shipment_date", "")
        
        # Оставляем только дату без времени (YYYY-MM-DD)
        if date and isinstance(date, str):
            if "T" in date:
                date = date.split("T")[0]
            elif " " in date:
                date = date.split(" ")[0]
        
        items = post.get("products", []) or []

        # Если в заказе нет товаров — пропускаем
        if not items:
            continue

        # Количество (сумма по позициям)
        quantity_total = sum(int(it.get("quantity", 0) or 0) for it in items)

        # Заголовок строки — первая позиция
        head = items[0]
        name = str(head.get("name", ""))
        # Все артикулы заказа (без дублей, в исходном порядке)
        seen = set()
        offer_ids_list = []
        for it in items:
            oid = str(it.get("offer_id", ""))
            if oid and oid not in seen:
                seen.add(oid)
                offer_ids_list.append(oid)
        offer_ids_joined = ", ".join(offer_ids_list)

        # Себестоимость (по всем товарам, со знаком минус)
        cost_price = 0.0
        for it in items:
            oid = str(it.get("offer_id", "") or "").strip()
            q = int(it.get("quantity", 0) or 0)

            # Совпадение по offer_id (ключ нормализован: 12345 и 12345.0 из Excel → один ключ)
            oid_norm = _normalize_articul_key(oid)
            unit_cost = float(cost_map.get(oid_norm, 0) or 0) if oid_norm else 0.0
            cost_price -= unit_cost * q

        # Агрегация транзакций по заказу (без дублей, без эквайринга)
        amount = 0.0
        sale_commission = 0.0
        price = 0.0

        transactions = get_transactions(posting_number, date_from, date_to, session=session)
        for trans in transactions or []:
            amount += float(trans.get("amount") or 0)
            sale_commission += float(trans.get("sale_commission") or 0)
            price += float(trans.get("accruals_for_sale") or 0)

        # Формируем значения в зависимости от статуса
        if status == "delivering":
            amount_cell = amount
            sale_commission_cell = "-"
            delivery_cost_cell = "-"
            profit_cell = "-"
            cost_price = 0.0   # себестоимость 0 — заказ ещё в доставке
        elif status == "awaiting_packaging":
            amount_cell = "-"
            sale_commission_cell = "-"
            delivery_cost_cell = "-"
            profit_cell = "-"
            cost_price = 0.0   # себестоимость 0 — заказ ожидает сборки
        elif status == "cancelled":
            amount_cell = amount
            sale_commission_cell = "-"
            delivery_cost_cell = "-"
            profit_cell = amount
            cost_price = 0.0   # ← себестоимость обнуляем при отмене
        elif status == "delivered":
            amount_cell = amount
            sale_commission_cell = sale_commission
            delivery_cost_cell = - amount + price + sale_commission
            profit_cell = amount + cost_price
            # Если при доставленном заказе прибыль получилась отрицательной —
            # считаем, что заказ по сути возврат: убыток = минус стоимость логистики,
            # себестоимость = 0, статус меняем на returned.
            if profit_cell < 0:
                status = "returned"
                cost_price = 0.0
                # Итоговая прибыль при возврате — всегда со знаком минус (убыток)
                profit_cell = -abs(delivery_cost_cell)
        else:
            amount_cell = "-"
            sale_commission_cell = "-"
            delivery_cost_cell = "-"
            profit_cell = "-"

        artikul_val = _artikul_to_number(offer_ids_joined) if len(offer_ids_list) == 1 else offer_ids_joined
        rows.append({
            "Статус": status,
            "Номер заказа": posting_number,
            "Название товара": name,
            "Артикул": artikul_val,
            "Количество шт.": quantity_total,
            "Цена продажи": price,
            "Комиссия за продажу Ozon": sale_commission_cell,
            "Логистика (Включает операционные ошибки продавца)": delivery_cost_cell,
            "Сумма начисления": amount_cell,
            "Себестоимость": cost_price,
            "Прибыль": profit_cell,
            "Дата отгрузки": date,
            "Схема": post.get("__schema", "")
        })

        # выводим прогресс каждые 5 записей и на финише
        if idx % 5 == 0 or idx == total_posts:
            percent = int(idx * 100 / total_posts)
            print(f"\r⚙️ Обработка заказов: {percent}%", end="", flush=True)

    df = pd.DataFrame(rows)
    if "Артикул" in df.columns:
        df["Артикул"] = df["Артикул"].apply(_artikul_to_number)
    output_file = _safe_save_excel(df, output_file, sheet_name="Заказы")
    print("\r✅ Обработка заказов: 100%")
    print(f"✅ Отчёт сохранён: {output_file}")
    return output_file

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def create_campaigns_sheet(filename: str, session: Optional[requests.Session] = None,
                           date_from: Optional[str] = None, date_to: Optional[str] = None):
    """
    Создаёт лист Excel с данными обо всех рекламных кампаниях за период (активные и неактивные).
    """
    if not session or not date_from or not date_to:
        return
    
    print("📊 Получаем данные о рекламных кампаниях за период...")
    
    campaigns_data = get_campaigns_data_for_excel(session, date_from, date_to)
    
    if campaigns_data is None:
        print("ℹ️ Не настроены переменные для Performance API. Пропускаем создание листа кампаний.")
        return
    
    if not campaigns_data:
        print("ℹ️ Не найдено кампаний за указанный период.")
        return
    
    try:
        # Открываем Excel-файл
        wb = load_workbook(filename)
        
        # Удаляем лист "Кампании", если он уже существует
        if "Кампании" in wb.sheetnames:
            wb.remove(wb["Кампании"])
        
        # Создаём новый лист
        ws_campaigns = wb.create_sheet("Кампании")
        
        # Заголовки столбцов
        headers = [
            "ID кампании", "Название кампании", "Состояние", "Тип оплаты", "Тип объекта",
            "Бюджет (руб.)", "Дневной бюджет (руб.)", "Недельный бюджет (руб.)",
            "Расход за период (руб.)", "Показы", "Клики", "CTR (%)",
            "Средняя цена клика (руб.)", "Заказы (шт.)", "Заказы (руб.)", "ДРР (%)"
        ]
        
        # Записываем заголовки
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_campaigns.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Записываем данные
        for row_idx, campaign in enumerate(campaigns_data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_campaigns.cell(row=row_idx, column=col_idx)
                value = campaign.get(header, "")
                
                # Форматируем числовые значения
                if isinstance(value, (int, float)):
                    cell.value = value
                    if "руб." in header or "ДРР" in header or "CTR" in header:
                        cell.number_format = "#,##0.00"
                    elif "Показы" in header or "Клики" in header or "Заказы (шт.)" in header:
                        cell.number_format = "#,##0"
                else:
                    cell.value = value
                
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Автоматически подбираем ширину столбцов
        for col_idx, header in enumerate(headers, start=1):
            max_length = len(str(header))
            for row in ws_campaigns.iter_rows(min_row=2, max_row=ws_campaigns.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws_campaigns.column_dimensions[ws_campaigns.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        # Сохраняем изменения
        wb.save(filename)
        print(f"✅ Лист 'Кампании' создан: {len(campaigns_data)} кампаний")
        
    except Exception as e:
        print(f"⚠️ Ошибка при создании листа кампаний: {str(e)}")


def calc_business_indicators(filename, session: Optional[requests.Session] = None, 
                            date_from: Optional[str] = None, date_to: Optional[str] = None):
    print("💲 Рассчёт бизнес показателей")
    
    # Пытаемся получить затраты на продвижение Ozon из Performance API
    ozon_promotion_cost = 0.0
    if session and date_from and date_to:
        perf_stats = get_cpc_campaigns_for_month(session, date_from, date_to)
        ozon_promotion_cost = perf_stats.get("total_cost", 0.0)
        if ozon_promotion_cost > 0:
            print(f"💰 Затраты на продвижение Ozon (CPC) из API: {ozon_promotion_cost:.2f} ₽")
    
    # Если не получилось из API или сумма 0 - спрашиваем у пользователя
    if ozon_promotion_cost == 0.0:
        print("Введите сумму затрат на продвижение Ozon за месяц (или Enter для 0):")
        try:
            user_input = input().strip()
            if user_input:
                ozon_promotion_cost = abs(float(user_input.replace(",", ".")))
            else:
                ozon_promotion_cost = 0.0
        except ValueError:
            print("❌ Некорректное число. Используем 0.")
            ozon_promotion_cost = 0.0
    
    # Запрашиваем затраты на внешний маркетинг (кампании не на Ozon)
    external_marketing_cost = 0.0
    print("Введите сумму затрат на внешний маркетинг за месяц (кампании не на Ozon, или Enter для 0):")
    try:
        user_input = input().strip()
        if user_input:
            external_marketing_cost = abs(float(user_input.replace(",", ".")))
        else:
            external_marketing_cost = 0.0
    except ValueError:
        print("❌ Некорректное число. Используем 0.")
        external_marketing_cost = 0.0
    
    
    # Открываем Excel-файл; лист «Заказы»: A=Статус, F=Цена продажи, G=Комиссия Ozon, H=Логистика, J=Себестоимость, K=Прибыль
    wb = load_workbook(filename)
    ws = wb["Заказы"] if "Заказы" in wb.sheetnames else wb.active

    # Считаем Общую выручку, Чистую прибыль, Себестоимость
    sales_revenue = 0
    for cell in ws["F"][1:]:
        if isinstance(cell.value, (int, float)):
            sales_revenue += cell.value

    net_profit = 0
    for cell in ws["K"][1:]:
        if isinstance(cell.value, (int, float)):
            net_profit += cell.value

    cost_price = 0
    for cell in ws["J"][1:]:
        if isinstance(cell.value, (int, float)):
            cost_price += cell.value

    # Новые показатели по строкам заказов: статус, средний чек, отменённые/доставленные, средние доли комиссии и логистики
    total_orders = max(0, ws.max_row - 1)
    delivered_count = 0
    cancelled_returned_count = 0
    ratios_commission_pct = []   # Комиссия Ozon / Цена продажи, %
    ratios_logistics_pct = []   # Логистика / Цена продажи, %
    revenue_for_avg_check = 0.0
    orders_nonzero_price = 0

    for row in range(2, ws.max_row + 1):
        status_val = ws.cell(row=row, column=1).value
        status = str(status_val).strip().lower() if status_val is not None else ""
        if status == "delivered":
            delivered_count += 1
        if status in ("cancelled", "returned"):
            cancelled_returned_count += 1

        price_val = ws.cell(row=row, column=6).value
        comm_val = ws.cell(row=row, column=7).value
        log_val = ws.cell(row=row, column=8).value

        try:
            price = float(price_val) if price_val is not None and str(price_val).strip() not in ("-", "") else None
        except (TypeError, ValueError):
            price = None
        try:
            comm = float(comm_val) if comm_val is not None and str(comm_val).strip() not in ("-", "") else None
        except (TypeError, ValueError):
            comm = None
        try:
            log = float(log_val) if log_val is not None and str(log_val).strip() not in ("-", "") else None
        except (TypeError, ValueError):
            log = None

        if price is not None and price != 0:
            revenue_for_avg_check += price
            orders_nonzero_price += 1
            if comm is not None:
                ratios_commission_pct.append(abs((comm / price) * 100))
            if log is not None:
                ratios_logistics_pct.append((log / price) * 100)

    average_check = (revenue_for_avg_check / orders_nonzero_price) if orders_nonzero_price > 0 else 0
    avg_commission_pct = (sum(ratios_commission_pct) / len(ratios_commission_pct)) if ratios_commission_pct else 0
    avg_logistics_pct = (sum(ratios_logistics_pct) / len(ratios_logistics_pct)) if ratios_logistics_pct else 0

    # Вычитаем затраты на продвижение Ozon и внешний маркетинг из чистой прибыли
    total_marketing_cost = ozon_promotion_cost + external_marketing_cost
    net_profit = net_profit - total_marketing_cost
    net_profit_margin = (net_profit / sales_revenue) * 100 if sales_revenue > 0 else 0
    cogs = sales_revenue + cost_price
    gross_profit_margin = (cogs / sales_revenue) * 100 if sales_revenue > 0 else 0
    operating_expenses = cogs - net_profit

    # Записываем результат
    ws["P1"] = "Общая выручка"
    ws["Q1"] = sales_revenue
    ws["P2"] = "Чистая прибыль"
    ws["Q2"] = net_profit
    ws["P3"] = "Итоговая себестоимость"
    ws["Q3"] = cost_price
    ws["P4"] = "Рентабельность по чистой прибыли (Net Profit Margin) %"
    ws["Q4"] = net_profit_margin
    ws["P5"] = "COGS (валовая прибыль)"
    ws["Q5"] = cogs
    ws["P6"] = "Gross Profit Margin Рентабельность по валовой прибыли %"
    ws["Q6"] = gross_profit_margin
    ws["P7"] = "Операционные расходы"
    ws["Q7"] = operating_expenses
    ws["P8"] = "Продвижение Ozon"
    ws["Q8"] = ozon_promotion_cost
    ws["P9"] = "Внешний маркетинг"
    ws["Q9"] = external_marketing_cost

    ws["P10"] = "Средний чек"
    ws["Q10"] = average_check
    ws["P11"] = "Общее количество заказов"
    ws["Q11"] = total_orders
    ws["P12"] = "Количество отменённых заказов"
    ws["Q12"] = cancelled_returned_count
    ws["P13"] = "Количество доставленных заказов"
    ws["Q13"] = delivered_count
    ws["P14"] = "Комиссии Ozon %"
    ws["Q14"] = avg_commission_pct
    ws["P15"] = "Логистика %"
    ws["Q15"] = avg_logistics_pct

    # Сохраняем изменения
    wb.save(filename)
    print(f"✅ Бизнес показатели добавлены в отчёт")
    
    # Создаём лист с данными о кампаниях
    create_campaigns_sheet(filename, session=session, date_from=date_from, date_to=date_to)

# 🚀 Точка входа
def date_range_for_month(month: int, year: int):
    """Возвращает (date_from, date_to) в формате API для заданных месяца и года."""
    from calendar import monthrange
    first_day = datetime(year, month, 1)
    last_day = datetime(year, month, monthrange(year, month)[1])
    return first_day.strftime('%Y-%m-%dT00:00:00Z'), last_day.strftime('%Y-%m-%dT23:59:59Z')


def main(argv=None):
    import argparse
    parser = argparse.ArgumentParser(description="Месячный отчёт по продажам Ozon.")
    parser.add_argument("--month", type=int, default=None, help="Номер месяца (1–12), для неинтерактивного запуска")
    parser.add_argument("--year", type=int, default=None, help="Год (например 2025), для неинтерактивного запуска")
    args = parser.parse_args(argv)

    if args.month is not None and args.year is not None:
        if not (1 <= args.month <= 12 and 2000 <= args.year <= 2100):
            raise ValueError("Укажите месяц 1–12 и год 2000–2100")
        month, year = args.month, args.year
        date_from, date_to = date_range_for_month(month, year)
    else:
        date_from, date_to, month, year = get_custom_date_range()

    print("📦 Получаем список заказов за месяц...")
    session = create_session()
    fbs_orders = get_orders(date_from, date_to, session=session)
    fbo_orders = get_fbo_orders(date_from, date_to, session=session)

    all_orders = fbs_orders + fbo_orders
    print(f"🔢 Найдено заказов: {len(all_orders)}")
    
    # Имя файла формируется внутри to_excel как "<Месяц> <Год>.xlsx"
    start_ts = time.time()
    output_file = to_excel(all_orders, date_from, date_to, month, year, session=session)
    duration_s = time.time() - start_ts
    
    calc_business_indicators(output_file, session=session, date_from=date_from, date_to=date_to)
    # Краткий итог
    print(f"⏱ Время формирования: {duration_s:.1f} с")


if __name__ == "__main__":
    main()