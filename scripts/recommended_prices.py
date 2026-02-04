# -*- coding: utf-8 -*-
"""
Рассчёт рекомендуемых цен продажи по артикулам.
Использует данные из предыдущего месячного отчёта (Комиссия Ozon %, Логистика %)
и себестоимость из costs.xlsx. Добавляет в costs.xlsx две колонки:
«Минимальная цена продажи» (маржа = ввод пользователя, по умолчанию 0.25)
и «Желательная цена продажи» (маржа = 0.30).

Формула: (x - (Комиссия% + Логистика%)*x - Себестоимость) / x = маржа
=> x = Себестоимость / (1 - (Комиссия% + Логистика%) - маржа)
"""

import os
import sys
import subprocess
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional, List, Any, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Загружаем переменные окружения для API Ozon
load_dotenv()
OZON_CLIENT_ID = os.getenv('OZON_CLIENT_ID')
OZON_API_KEY = os.getenv('OZON_API_KEY')

try:
    from utils import log_verbose, VERBOSE
except ImportError:
    def log_verbose(_: str) -> None:
        pass
    VERBOSE = False

OZON_HEADERS = {
    'Client-Id': OZON_CLIENT_ID or '',
    'Api-Key': OZON_API_KEY or '',
    'Content-Type': 'application/json'
}

# Названия месяцев на русском (для имени файла отчёта)
MONTHS_RU = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
]

REPORTS_DIR_NAME = "reports"
COSTS_FILENAME = "costs.xlsx"
ORDER_SHEET = "Заказы"
MARGIN_SETTINGS_FILE = "margin_settings.json"

# Колонки в листе «Заказы» месячного отчёта (1-based): Статус=1, Артикул=4, Цена продажи=6, Комиссия=7, Логистика=8
COL_STATUS = 1
COL_ARTIKUL = 4
COL_PRICE = 6
COL_COMMISSION = 7
COL_LOGISTICS = 8

DESIRED_MARGIN_DEFAULT = 0.30
MIN_MARGIN_DEFAULT = 0.25

COL_MIN_PRICE = "Минимальная цена продажи"
COL_DESIRED_PRICE = "Желательная цена продажи"
COL_CURRENT_PRICE = "Текущая цена на Ozon"
COL_MARKETING_PRICE = "Цена с учётом акций и скидок"
COL_CURRENT_MARGIN = "Текущая ожидаемая рентабельность"
# COL_ACTIONS больше не используется - вместо неё будут отдельные колонки для каждой акции


def _artikul_normalize(v):
    """Нормализация артикула для сопоставления (строка, без пробелов по краям)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def _normalize_offer_id(offer_id: str) -> str:
    """
    Преобразует артикул в строку без .0 (например, '1101.0' -> '1101').
    """
    try:
        # Пробуем преобразовать в float, затем в int, затем в строку
        return str(int(float(offer_id)))
    except (ValueError, TypeError):
        # Если не получается, возвращаем как есть, убрав пробелы
        return str(offer_id).strip()


def get_product_prices_from_ozon(offer_ids: list[str]) -> tuple[Dict[str, Optional[float]], Dict[str, Optional[float]]]:
    """
    Получает текущие цены товаров с Ozon по их offer_id (артикулам).
    Использует API v5/product/info/prices с правильной структурой запроса.
    Возвращает кортеж (словарь обычных цен, словарь цен с учётом акций и скидок).
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        print("⚠️ OZON_CLIENT_ID или OZON_API_KEY не настроены. Пропускаем получение цен с Ozon.")
        return {}
    
    if not offer_ids:
        return {}
    
    # Нормализуем артикулы: убираем .0 и преобразуем в строки
    normalized_offer_ids = []
    for oid in offer_ids:
        normalized = _normalize_offer_id(oid)
        if normalized:
            normalized_offer_ids.append(normalized)
    
    if not normalized_offer_ids:
        return {}
    
    # API Ozon v5/product/info/prices
    url = "https://api-seller.ozon.ru/v5/product/info/prices"
    prices_map = {}
    marketing_prices_map = {}
    batch_size = 100
    
    # Разбиваем на батчи по 100 артикулов
    for i in range(0, len(normalized_offer_ids), batch_size):
        batch = normalized_offer_ids[i:i + batch_size]
        
        cursor = ""
        has_more = True
        
        while has_more:
            payload = {
                "cursor": cursor,
                "filter": {
                    "offer_id": batch,
                    "visibility": "ALL"
                },
                "limit": 100
            }
            
            try:
                response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
                response.raise_for_status()
                data = response.json()
                
                # В API v5 структура ответа: items, cursor, total на верхнем уровне
                items = data.get("items", [])
                total = data.get("total", 0)
                cursor = data.get("cursor", "")
                
                if len(items) == 0 and total == 0:
                    break
                
                for item in items:
                    # Получаем offer_id из ответа для сопоставления
                    offer_id_raw = item.get("offer_id", "")
                    offer_id_normalized = _normalize_offer_id(str(offer_id_raw)) if offer_id_raw else None
                    
                    # Также сохраняем маппинг product_id -> offer_id для использования в акциях
                    product_id = item.get("product_id")
                    if product_id and offer_id_normalized:
                        # Это будет использовано позже в get_actions_for_products
                        pass
                    
                    # Получаем цену из объекта price (структура: item["price"]["price"])
                    price = None
                    marketing_price = None
                    price_obj = item.get("price", {})
                    
                    if isinstance(price_obj, dict):
                        # Пробуем получить цену из price.price
                        if "price" in price_obj:
                            try:
                                price = float(price_obj["price"])
                            except (TypeError, ValueError):
                                pass
                        
                        # Получаем цену с учётом акций и скидок из price.marketing_seller_price
                        if "marketing_seller_price" in price_obj:
                            try:
                                marketing_price = float(price_obj["marketing_seller_price"])
                            except (TypeError, ValueError):
                                pass
                        
                        # Если price нет, пробуем old_price
                        if price is None and "old_price" in price_obj:
                            try:
                                price = float(price_obj["old_price"])
                            except (TypeError, ValueError):
                                pass
                    else:
                        # Старый формат (если price - это число напрямую)
                        try:
                            price = float(price_obj)
                        except (TypeError, ValueError):
                            pass
                    
                    # Сохраняем по offer_id для сопоставления с costs.xlsx
                    if offer_id_normalized:
                        # Находим соответствующий исходный артикул из списка
                        for orig_oid in offer_ids:
                            if _normalize_offer_id(orig_oid) == offer_id_normalized:
                                prices_map[orig_oid] = price
                                marketing_prices_map[orig_oid] = marketing_price
                                break
                        # Также сохраняем по нормализованному ключу для прямого доступа
                        prices_map[offer_id_normalized] = price
                        marketing_prices_map[offer_id_normalized] = marketing_price
                
                # Проверяем, есть ли ещё страницы
                has_more = bool(cursor) and len(items) >= 100
                    
            except requests.exceptions.RequestException as e:
                print(f"⚠️ Ошибка при запросе цен для артикулов {batch[:3]}...: {e}")
                if hasattr(e, 'response') and e.response is not None:
                    try:
                        error_data = e.response.json()
                        print(f"   Детали ошибки: {error_data}")
                    except:
                        print(f"   Ответ сервера: {e.response.text[:200]}")
                # Прерываем цикл пагинации для этого батча
                break
            except Exception as e:
                print(f"⚠️ Неожиданная ошибка при получении цен: {e}")
                break
    
    return prices_map, marketing_prices_map


def get_actions_list() -> List[Dict[str, Any]]:
    """
    Получает список всех акций через API /v1/actions.
    Возвращает список словарей с информацией об акциях.
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        return []
    
    url = "https://api-seller.ozon.ru/v1/actions"
    
    try:
        # API /v1/actions использует GET метод
        response = requests.get(url, headers=OZON_HEADERS, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        # Структура ответа может быть разной, проверяем разные варианты
        if isinstance(data, dict):
            if "result" in data:
                actions = data["result"]
                if isinstance(actions, list):
                    return actions
                elif isinstance(actions, dict) and "actions" in actions:
                    return actions["actions"]
            elif "actions" in data:
                return data["actions"]
        elif isinstance(data, list):
            return data
        
        return []
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 405:
            # Пробуем POST метод (если GET не работает)
            try:
                response = requests.post(url, headers=OZON_HEADERS, json={}, timeout=30)
                response.raise_for_status()
                data = response.json()
                if isinstance(data, dict):
                    if "result" in data:
                        actions = data["result"]
                        if isinstance(actions, list):
                            return actions
                        elif isinstance(actions, dict) and "actions" in actions:
                            return actions["actions"]
                    elif "actions" in data:
                        return data["actions"]
                elif isinstance(data, list):
                    return data
            except:
                pass
        print(f"⚠️ Ошибка при получении списка акций: {e}")
        if hasattr(e, 'response') and e.response is not None:
            try:
                error_data = e.response.json()
                print(f"   Детали ошибки: {error_data}")
            except:
                print(f"   Ответ сервера: {e.response.text[:200]}")
        return []
    except requests.exceptions.RequestException as e:
        print(f"⚠️ Ошибка при получении списка акций: {e}")
        return []
    except Exception as e:
        print(f"⚠️ Неожиданная ошибка при получении списка акций: {e}")
        return []


def get_product_info_by_id(product_id: int, debug: bool = False) -> Optional[str]:
    """
    Получает offer_id товара по его product_id через API /v2/product/info.
    Возвращает offer_id или None.
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        return None
    
    url = "https://api-seller.ozon.ru/v2/product/info"
    payload = {
        "product_id": product_id
    }
    
    try:
        response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        # Структура ответа может быть разной, проверяем разные варианты
        if isinstance(data, dict):
            # Вариант 1: {"result": {"offer_id": "...", ...}}
            if "result" in data:
                result_data = data["result"]
                # Проверяем разные возможные поля
                offer_id = result_data.get("offer_id") or result_data.get("offerId") or result_data.get("sku")
                if offer_id:
                    if debug:
                        print(f"      DEBUG: product_id {product_id} -> offer_id {offer_id} (найден в result)")
                    return str(offer_id)
                elif debug:
                    print(f"      DEBUG: product_id {product_id} - ключи в result: {list(result_data.keys())[:10]}")
            
            # Вариант 2: {"offer_id": "...", ...} на верхнем уровне
            offer_id = data.get("offer_id") or data.get("offerId") or data.get("sku")
            if offer_id:
                if debug:
                    print(f"      DEBUG: product_id {product_id} -> offer_id {offer_id} (найден на верхнем уровне)")
                return str(offer_id)
    except requests.exceptions.RequestException as e:
        if debug:
            if hasattr(e, 'response') and e.response is not None:
                try:
                    error_data = e.response.json()
                    print(f"      DEBUG: Ошибка API для product_id {product_id}: {error_data}")
                except:
                    print(f"      DEBUG: Ошибка для product_id {product_id}: {e.response.status_code} - {e.response.text[:200]}")
            else:
                print(f"      DEBUG: Ошибка при получении offer_id для product_id {product_id}: {e}")
        # Для первых запросов всегда показываем ошибки
        if product_id and len(str(product_id)) < 10:  # Простая проверка для первых запросов
            if hasattr(e, 'response') and e.response is not None:
                try:
                    error_data = e.response.json()
                    print(f"      ⚠️ Ошибка API для SKU {product_id}: {error_data}")
                except:
                    print(f"      ⚠️ Ошибка для SKU {product_id}: {e.response.status_code}")
    except Exception as e:
        if debug:
            print(f"      DEBUG: Неожиданная ошибка для product_id {product_id}: {e}")
    
    return None


def get_offer_ids_by_skus(skus: List[int]) -> Dict[int, str]:
    """
    Получает offer_id по списку SKU (идентификатор товара в системе Ozon) через API /v3/product/info/list.
    В заявках на скидку Ozon возвращает именно SKU — этот метод запрашивает по sku и возвращает offer_id.
    
    Args:
        skus: Список SKU из ответа API (например, из discounts-task/list).
    
    Returns:
        Словарь {sku: offer_id}
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY or not skus:
        return {}
    url = "https://api-seller.ozon.ru/v3/product/info/list"
    # В теле запроса передаём sku как массив (Ozon: "sku" — идентификаторы в системе Ozon)
    payload: Dict[str, Any] = {"sku": skus}
    mapping: Dict[int, str] = {}
    try:
        response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
        response.raise_for_status()
        data = response.json()
        result = data.get("result") if isinstance(data.get("result"), dict) else {}
        items = result.get("items", data.get("items", []))
        if not items and isinstance(data.get("result"), list):
            items = data["result"]
        skus_set = set(skus)
        for item in items:
            offer_id = item.get("offer_id") or item.get("offerId") or item.get("article")
            if not offer_id:
                continue
            offer_id_str = str(offer_id).strip()
            # Ключ в ответе может быть sku или id (product_id)
            sku_val = item.get("sku") if item.get("sku") is not None else item.get("id")
            if sku_val is not None:
                try:
                    sku_int = int(sku_val)
                    if sku_int in skus_set:
                        mapping[sku_int] = offer_id_str
                except (TypeError, ValueError):
                    pass
        # Если по sku не сопоставили, но количество совпадает — считаем порядок как в запросе
        if len(mapping) < len(skus) and len(items) == len(skus):
            for i, item in enumerate(items):
                if i >= len(skus):
                    break
                offer_id = item.get("offer_id") or item.get("offerId") or item.get("article")
                if offer_id and skus[i] not in mapping:
                    mapping[skus[i]] = str(offer_id).strip()
    except requests.exceptions.RequestException as e:
        if VERBOSE and hasattr(e, "response") and e.response is not None:
            try:
                err = e.response.json()
                log_verbose(f"v3/product/info/list: {err}")
            except Exception:
                pass
    except Exception as e:
        log_verbose(f"get_offer_ids_by_skus: {e}")
    return mapping


def get_products_in_action(action_id: int, action_name: str, offer_ids_set: set[str], product_id_to_offer_id: Dict[int, str]) -> Dict[str, float]:
    """
    Получает товары в акции через API /v1/actions/products с использованием last_id вместо offset.
    Возвращает словарь {offer_id: цена_в_акции}.
    
    Args:
        action_id: ID акции
        action_name: Название акции
        offer_ids_set: Множество нормализованных offer_id для фильтрации
        product_id_to_offer_id: Словарь маппинга product_id -> offer_id
    
    Returns:
        Словарь {offer_id: цена_в_акции} для товаров, участвующих в акции
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        return {}
    
    url = "https://api-seller.ozon.ru/v1/actions/products"
    result = {}
    last_id = None
    limit = 100
    
    while True:
        payload = {
            "action_id": action_id,
            "limit": limit
        }
        
        if last_id:
            payload["last_id"] = last_id
        
        try:
            response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            # Структура ответа: {"result": {"products": [...], "total": ..., "last_id": ...}}
            if isinstance(data, dict) and "result" in data:
                result_data = data["result"]
                products = result_data.get("products", [])
                
                if not products:
                    break
                
                # Обрабатываем товары
                found_count = 0
                total_products = len(products)
                no_offer_id_count = 0
                not_in_list_count = 0
                
                for idx, product in enumerate(products):
                    # В ответе есть только id (product_id), нужно получить offer_id
                    product_id = product.get("id")
                    if not product_id:
                        continue
                    
                    # Пробуем найти offer_id в кэше маппинга
                    offer_id = product_id_to_offer_id.get(product_id)
                    
                    # Если нет в кэше, получаем через API
                    # Для первых 3 товаров включаем отладку
                    if not offer_id:
                        offer_id = get_product_info_by_id(product_id, debug=VERBOSE)
                        if offer_id:
                            product_id_to_offer_id[product_id] = offer_id
                        else:
                            no_offer_id_count += 1
                            continue
                    
                    if offer_id:
                        offer_id_normalized = _normalize_offer_id(offer_id)
                        # Проверяем, есть ли этот артикул в нашем списке
                        if offer_id_normalized in offer_ids_set or offer_id in offer_ids_set:
                            # Получаем цену в акции (action_price)
                            action_price = product.get("action_price")
                            if action_price is not None:
                                try:
                                    action_price_val = float(action_price)
                                    # Сохраняем цену в акции для этого артикула
                                    result[offer_id_normalized] = action_price_val
                                    found_count += 1
                                except (TypeError, ValueError):
                                    pass
                        else:
                            not_in_list_count += 1
                
                if total_products > 0:
                    log_verbose(f"Акция {action_id}: обработано {total_products}, совпадений {found_count}")
                    if no_offer_id_count > 0 or not_in_list_count > 0:
                        log_verbose(f"  без offer_id: {no_offer_id_count}, не в списке: {not_in_list_count}")
                
                # Получаем last_id для следующей страницы
                last_id = result_data.get("last_id")
                if not last_id:
                    break
            else:
                break
                
        except requests.exceptions.RequestException as e:
            print(f"⚠️ Ошибка при получении товаров акции {action_id}: {e}")
            break
        except Exception as e:
            print(f"⚠️ Неожиданная ошибка при получении товаров акции {action_id}: {e}")
            break
    
    return result


def get_actions_for_products(
    offer_ids: List[str],
) -> tuple[Dict[str, Dict[str, float]], List[Dict[str, Any]], Dict[str, int]]:
    """
    Получает информацию об акциях для списка артикулов.
    Возвращает кортеж:
    - Словарь {offer_id: {название_акции: цена_в_акции}}
    - Список акций с их названиями и ID
    - Словарь {offer_id: product_id} для удаления из акций
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        print("⚠️ OZON_CLIENT_ID или OZON_API_KEY не настроены. Пропускаем получение акций.")
        return {}, [], {}
    
    if not offer_ids:
        return {}, [], {}
    
    # Нормализуем артикулы и создаём множество для быстрого поиска
    normalized_offer_ids = []
    offer_ids_set = set()
    for oid in offer_ids:
        normalized = _normalize_offer_id(oid)
        if normalized:
            normalized_offer_ids.append(normalized)
            offer_ids_set.add(normalized)
            # Также добавляем исходный артикул
            offer_ids_set.add(oid)
    
    if not offer_ids_set:
        return {}, [], {}
    
    log_verbose("Получение списка акций...")
    actions = get_actions_list()
    
    if not actions:
        print("⚠️ Не найдено активных акций.")
        return {}, [], {}
    
    print(f"✅ Найдено {len(actions)} акций")
    log_verbose("Получение товаров в акциях...")
    
    # Создаём маппинг product_id -> offer_id для кэширования
    # Сначала пытаемся получить маппинг через API /v5/product/info/prices
    # где есть и product_id и offer_id
    log_verbose("Создание маппинга product_id -> offer_id...")
    product_id_to_offer_id: Dict[int, str] = {}
    
    # Получаем маппинг через API /v5/product/info/prices
    url = "https://api-seller.ozon.ru/v5/product/info/prices"
    cursor = ""
    has_more = True
    
    while has_more:
        payload = {
            "cursor": cursor,
            "filter": {
                "offer_id": normalized_offer_ids,
                "visibility": "ALL"
            },
            "limit": 100
        }
        
        try:
            response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            items = data.get("items", [])
            cursor = data.get("cursor", "")
            
            for item in items:
                product_id = item.get("product_id")
                offer_id_raw = item.get("offer_id", "")
                if product_id and offer_id_raw:
                    offer_id_normalized = _normalize_offer_id(str(offer_id_raw))
                    product_id_to_offer_id[product_id] = offer_id_normalized
            
            has_more = bool(cursor) and len(items) >= 100
            if not has_more:
                break
        except Exception as e:
            print(f"   ⚠️ Ошибка при создании маппинга: {e}")
            break
    
    log_verbose(f"Маппинг: {len(product_id_to_offer_id)} товаров.")
    
    # Словарь для накопления результатов: {offer_id: {название_акции: цена}}
    actions_by_offer: Dict[str, Dict[str, float]] = {}
    
    # Список акций с их названиями
    actions_info_list: List[Dict[str, Any]] = []
    
    # Обрабатываем каждую акцию
    for idx, action in enumerate(actions, 1):
        action_id = action.get("id") or action.get("action_id")
        if not action_id:
            continue
        
        action_name = action.get("title") or action.get("name") or f"Акция {action_id}"
        log_verbose(f"Обработка акции {idx}/{len(actions)}: {action_name}")
        
        # Сохраняем информацию об акции
        actions_info_list.append({
            "id": action_id,
            "name": action_name
        })
        
        products_in_action = get_products_in_action(action_id, action_name, offer_ids_set, product_id_to_offer_id)
        
        # Объединяем результаты: сохраняем цену в акции для каждого артикула
        for offer_id, action_price in products_in_action.items():
            if offer_id not in actions_by_offer:
                actions_by_offer[offer_id] = {}
            actions_by_offer[offer_id][action_name] = action_price
    
    if actions_by_offer:
        print(f"✅ В акциях: {len(actions_by_offer)} артикулов")
    else:
        print("⚠️ Не найдено товаров из вашего списка в активных акциях.")
    
    # Обратный маппинг offer_id -> product_id
    offer_id_to_product_id: Dict[str, int] = {}
    for product_id, offer_id in product_id_to_offer_id.items():
        offer_id_to_product_id[_normalize_offer_id(str(offer_id))] = product_id

    return actions_by_offer, actions_info_list, offer_id_to_product_id


def deactivate_products_in_action(action_id: int, product_ids: List[int]) -> Dict[str, Any]:
    """
    Удаляет товары из акции через API /v1/actions/products/deactivate.
    Возвращает агрегированный результат по всем батчам.
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        print("⚠️ OZON_CLIENT_ID или OZON_API_KEY не настроены. Пропускаем удаление из акций.")
        return {"product_ids": [], "rejected": []}

    if not product_ids:
        return {"product_ids": [], "rejected": []}

    url = "https://api-seller.ozon.ru/v1/actions/products/deactivate"
    batch_size = 100
    removed_ids: List[int] = []
    rejected: List[Dict[str, Any]] = []

    for i in range(0, len(product_ids), batch_size):
        batch = product_ids[i:i + batch_size]
        payload = {
            "action_id": action_id,
            "product_ids": batch,
        }
        try:
            response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
            response.raise_for_status()
            data = response.json()
            result = data.get("result", {})
            removed_ids.extend(result.get("product_ids", []) or [])
            rejected.extend(result.get("rejected", []) or [])
        except requests.exceptions.RequestException as e:
            print(f"⚠️ Ошибка при удалении товаров из акции {action_id}: {e}")
            if hasattr(e, "response") and e.response is not None:
                try:
                    error_data = e.response.json()
                    print(f"   Детали ошибки: {error_data}")
                except Exception:
                    print(f"   Ответ сервера: {e.response.text[:200]}")
        except Exception as e:
            print(f"⚠️ Неожиданная ошибка при удалении из акции {action_id}: {e}")

    return {"product_ids": removed_ids, "rejected": rejected}


def get_action_candidates(action_id: int, product_id_to_offer_id: Dict[int, str], offer_ids_set: set[str]) -> Dict[int, Dict[str, Any]]:
    """
    Получает список товаров, которые могут участвовать в акции через API /v1/actions/candidates.
    Использует last_id вместо offset.
    Возвращает словарь {product_id: информация_о_товаре} для товаров из нашего списка.
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        return {}
    
    url = "https://api-seller.ozon.ru/v1/actions/candidates"
    result = {}
    last_id = None
    limit = 100
    
    while True:
        payload = {
            "action_id": action_id,
            "limit": limit
        }
        
        if last_id:
            payload["last_id"] = last_id
        
        try:
            response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            # Структура ответа: {"result": {"products": [...], "total": ..., "last_id": ...}}
            if isinstance(data, dict) and "result" in data:
                result_data = data["result"]
                products = result_data.get("products", [])
                
                if not products:
                    break
                
                # Обрабатываем товары
                for product in products:
                    product_id = product.get("id")
                    if not product_id:
                        continue
                    
                    # Проверяем, есть ли этот товар в нашем списке
                    offer_id = product_id_to_offer_id.get(product_id)
                    if offer_id:
                        offer_id_normalized = _normalize_offer_id(offer_id)
                        if offer_id_normalized in offer_ids_set or offer_id in offer_ids_set:
                            # Сохраняем информацию о товаре
                            result[product_id] = product
                
                # Получаем last_id для следующей страницы
                last_id = result_data.get("last_id")
                if not last_id:
                    break
            else:
                break
                
        except requests.exceptions.RequestException as e:
            print(f"⚠️ Ошибка при получении кандидатов акции {action_id}: {e}")
            break
        except Exception as e:
            print(f"⚠️ Неожиданная ошибка при получении кандидатов акции {action_id}: {e}")
            break
    
    return result


def activate_products_in_action(action_id: int, products: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Добавляет товары в акцию через API /v1/actions/products/activate.
    products - список словарей вида [{"product_id": int, "action_price": float, "stock": int}, ...]
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        print("⚠️ OZON_CLIENT_ID или OZON_API_KEY не настроены. Пропускаем добавление в акции.")
        return {"product_ids": [], "rejected": []}

    if not products:
        return {"product_ids": [], "rejected": []}

    url = "https://api-seller.ozon.ru/v1/actions/products/activate"
    batch_size = 1000  # Максимум 1000 товаров за запрос
    added_ids: List[int] = []
    rejected: List[Dict[str, Any]] = []

    for i in range(0, len(products), batch_size):
        batch = products[i:i + batch_size]
        payload = {
            "action_id": action_id,
            "products": batch
        }
        
        try:
            response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
            response.raise_for_status()
            data = response.json()
            result = data.get("result", {})
            added_ids.extend(result.get("product_ids", []) or [])
            rejected.extend(result.get("rejected", []) or [])
        except requests.exceptions.RequestException as e:
            print(f"⚠️ Ошибка при добавлении товаров в акцию {action_id}: {e}")
            if hasattr(e, "response") and e.response is not None:
                try:
                    error_data = e.response.json()
                    print(f"   Детали ошибки: {error_data}")
                except Exception:
                    print(f"   Ответ сервера: {e.response.text[:200]}")
        except Exception as e:
            print(f"⚠️ Неожиданная ошибка при добавлении в акцию {action_id}: {e}")

    return {"product_ids": added_ids, "rejected": rejected}


def collect_deactivation_candidates_from_sheet(
    ws_actions,
    key_col_name: str,
    min_price_col_name: str,
    action_name_to_id: Dict[str, int],
    offer_id_to_product_id: Dict[str, int],
) -> Dict[int, List[int]]:
    """
    Читает лист «Акции» и возвращает словарь {action_id: [product_id]},
    где цена в акции меньше минимальной цены продажи.
    """
    header_map: Dict[str, int] = {}
    for idx, cell in enumerate(ws_actions[1], start=1):
        if cell.value:
            header_map[str(cell.value)] = idx

    if key_col_name not in header_map or min_price_col_name not in header_map:
        return {}

    key_col_idx = header_map[key_col_name]
    min_price_col_idx = header_map[min_price_col_name]

    action_col_indices: Dict[str, int] = {}
    for action_name in action_name_to_id:
        if action_name in header_map:
            action_col_indices[action_name] = header_map[action_name]

    if not action_col_indices:
        return {}

    candidates: Dict[int, set[int]] = {}

    for row in range(2, ws_actions.max_row + 1):
        offer_id_val = ws_actions.cell(row=row, column=key_col_idx).value
        min_price_val = ws_actions.cell(row=row, column=min_price_col_idx).value

        if offer_id_val is None or min_price_val is None:
            continue

        try:
            min_price = float(min_price_val)
        except (TypeError, ValueError):
            continue

        offer_id_norm = _normalize_offer_id(str(offer_id_val))
        product_id = offer_id_to_product_id.get(offer_id_norm)
        if not product_id:
            continue

        for action_name, col_idx in action_col_indices.items():
            action_price_val = ws_actions.cell(row=row, column=col_idx).value
            if action_price_val is None:
                continue
            try:
                action_price = float(action_price_val)
            except (TypeError, ValueError):
                continue
            if action_price < min_price:
                action_id = action_name_to_id[action_name]
                if action_id not in candidates:
                    candidates[action_id] = set()
                candidates[action_id].add(product_id)

    return {action_id: sorted(list(product_ids)) for action_id, product_ids in candidates.items()}


def load_margin_settings(repo_root: Path) -> Tuple[Optional[float], Optional[float]]:
    """
    Загружает сохранённые настройки диапазона рентабельности из файла.
    Возвращает (min_margin, desired_margin) или (None, None) если файл не найден.
    """
    settings_path = repo_root / MARGIN_SETTINGS_FILE
    if not settings_path.exists():
        return None, None
    
    try:
        with open(settings_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            min_margin = data.get("min_margin")
            desired_margin = data.get("desired_margin")
            if min_margin is not None and desired_margin is not None:
                return float(min_margin), float(desired_margin)
    except Exception as e:
        print(f"⚠️ Ошибка при загрузке настроек рентабельности: {e}")
    
    return None, None


def save_margin_settings(repo_root: Path, min_margin: float, desired_margin: float) -> None:
    """
    Сохраняет настройки диапазона рентабельности в файл.
    """
    settings_path = repo_root / MARGIN_SETTINGS_FILE
    try:
        with open(settings_path, 'w', encoding='utf-8') as f:
            json.dump({
                "min_margin": min_margin,
                "desired_margin": desired_margin
            }, f, indent=2, ensure_ascii=False)
        print(f"✅ Настройки рентабельности сохранены: минимальная {min_margin*100:.1f}%, желательная {desired_margin*100:.1f}%")
    except Exception as e:
        print(f"⚠️ Ошибка при сохранении настроек рентабельности: {e}")


def get_prev_month_year():
    """Возвращает (year, month) для предыдущего месяца."""
    now = datetime.now()
    if now.month == 1:
        return now.year - 1, 12
    return now.year, now.month - 1


def get_report_path(repo_root: Path, year: int, month: int) -> Path:
    """Путь к файлу месячного отчёта вида «Декабрь 2025.xlsx»."""
    name = f"{MONTHS_RU[month - 1]} {year}.xlsx"
    return repo_root / REPORTS_DIR_NAME / name


def load_rates_from_report(report_path: Path) -> float:
    """
    Читает значения «Комиссии Ozon %» (Q14) и «Логистика %» (Q15) из листа «Заказы»
    месячного отчёта и возвращает их сумму в виде десятичной доли (например, 0.15 для 15%).
    """
    if not report_path.exists():
        raise FileNotFoundError(f"Файл отчёта не найден: {report_path}")

    # Пробуем сначала с data_only=True (для вычисленных значений)
    # Если не получится, попробуем без него (для формул)
    wb = None
    try:
        wb = load_workbook(report_path, data_only=True)
    except Exception:
        pass
    
    if wb is None:
        try:
            wb = load_workbook(report_path, data_only=False)
        except Exception as e:
            raise ValueError(f"Не удалось открыть файл отчёта: {e}")
    
    if ORDER_SHEET not in wb.sheetnames:
        raise ValueError(f"В файле отчёта нет листа «{ORDER_SHEET}».")

    ws = wb[ORDER_SHEET]

    # Читаем значения из ячеек Q14 (Комиссии Ozon %) и Q15 (Логистика %)
    commission_pct = ws["Q14"].value
    logistics_pct = ws["Q15"].value
    
    # Если значения не прочитались, пробуем альтернативный способ через pandas
    if commission_pct is None or logistics_pct is None:
        try:
            df_summary = pd.read_excel(report_path, sheet_name=ORDER_SHEET, header=None, usecols="P:Q", skiprows=13, nrows=2)
            if len(df_summary) >= 2:
                if pd.notna(df_summary.iloc[0, 1]):
                    commission_pct = df_summary.iloc[0, 1]
                if pd.notna(df_summary.iloc[1, 1]):
                    logistics_pct = df_summary.iloc[1, 1]
        except Exception:
            pass

    def to_float(val):
        """Преобразует значение в float, обрабатывая None и строки."""
        if val is None:
            return 0.0
        if isinstance(val, str):
            val = val.strip().replace(",", ".")
            if val in ("-", ""):
                return 0.0
        try:
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    commission_pct_raw = commission_pct
    logistics_pct_raw = logistics_pct
    
    commission_pct = to_float(commission_pct)
    logistics_pct = to_float(logistics_pct)

    # Значения в процентах, преобразуем в десятичную долю (15% -> 0.15)
    commission_rate = commission_pct / 100.0
    logistics_rate = logistics_pct / 100.0

    # Общая доля комиссии + логистики
    total_rate = commission_rate + logistics_rate

    # Проверка: если комиссия и логистика равны 0 или очень малы, значит отчёт неполный
    if total_rate < 0.01:  # Меньше 1% - подозрительно мало
        # Пытаемся определить месяц и год из имени файла
        report_name = report_path.stem  # без расширения
        period_info = report_name  # например "Декабрь 2025"
        
        error_msg = (
            f"\n❌ ОШИБКА: Не удалось прочитать значения комиссии и логистики из отчёта.\n"
            f"   Файл: {report_path.name}\n"
            f"   Комиссия Ozon % (Q14): {commission_pct_raw}\n"
            f"   Логистика % (Q15): {logistics_pct_raw}\n\n"
            f"   Возможные причины:\n"
            f"   1. Отчёт не был сохранён после генерации (формулы не вычислены)\n"
            f"   2. В отчёте отсутствуют данные о комиссиях и логистике\n"
            f"   3. Отчёт был повреждён или изменён вручную\n\n"
            f"   РЕШЕНИЕ: Запустите генерацию месячного отчёта заново для периода\n"
            f"   «{period_info}» и убедитесь, что отчёт сохранён корректно.\n"
            f"   После генерации откройте файл отчёта в Excel и сохраните его,\n"
            f"   чтобы формулы вычислились, затем повторите расчёт цен.\n"
        )
        raise ValueError(error_msg)

    return total_rate


def load_costs_df(costs_path: Path) -> pd.DataFrame:
    """Читает costs.xlsx, возвращает DataFrame с колонками артикул и себестоимость (оригинальные имена)."""
    if not costs_path.exists():
        raise FileNotFoundError(f"Файл себестоимости не найден: {costs_path}")

    df = pd.read_excel(costs_path)
    lower_cols = {c.lower(): c for c in df.columns}
    key_col = None
    cost_col = None
    for v in ["prefix", "префикс", "код", "артикул", "offer_id"]:
        if v in lower_cols:
            key_col = lower_cols[v]
            break
    for v in ["cost", "себестоимость", "цена", "стоимость"]:
        if v in lower_cols:
            cost_col = lower_cols[v]
            break
    if not key_col or not cost_col:
        raise ValueError(
            "В costs.xlsx не найдены столбцы артикула и себестоимости. "
            "Ожидаются: 'артикул' (или prefix/код) и 'себестоимость' (или cost)."
        )
    return df, key_col, cost_col


def compute_prices(
    cost: float,
    rate: float,
    min_margin: float,
    desired_margin: float,
) -> tuple[float | None, float | None]:
    """
    Вычисляет рекомендуемые цены по формуле:
    (x - rate*x - cost) / x = margin
    => x - rate*x - cost = margin*x
    => x*(1 - rate - margin) = cost
    => x = cost / (1 - rate - margin)
    
    Пример:
    - rate = 0.6 (60% комиссии+логистики из отчёта)
    - cost = 250 (себестоимость)
    - margin = 0.3 (30% желательная рентабельность)
    - x = 250 / (1 - 0.6 - 0.3) = 250 / 0.1 = 2500
    
    Проверка: (2500 - 0.6*2500 - 250) / 2500 = 750 / 2500 = 0.3 ✓
    
    Если знаменатель <= 0 или cost <= 0, возвращаем (None, None).
    """
    if cost is None or cost <= 0:
        return None, None
    denom_min = 1.0 - rate - min_margin
    denom_desired = 1.0 - rate - desired_margin
    if denom_min <= 0 or denom_desired <= 0:
        return None, None
    min_price = round(cost / denom_min)
    desired_price = round(cost / denom_desired)
    return min_price, desired_price


def compute_current_margin(price: float | None, cost: float, rate: float) -> float | None:
    """
    Вычисляет текущую ожидаемую рентабельность по формуле:
    margin = (price - rate*price - cost) / price
    
    Пример:
    - price = 2500 (текущая цена на Ozon)
    - rate = 0.6 (60% комиссии+логистики)
    - cost = 250 (себестоимость)
    - margin = (2500 - 0.6*2500 - 250) / 2500 = 750 / 2500 = 0.3 (30%)
    
    Если price <= 0 или None, возвращает None.
    """
    if price is None or price <= 0:
        return None
    try:
        margin = (price - rate * price - cost) / price
        return margin
    except (TypeError, ValueError, ZeroDivisionError):
        return None


def generate_monthly_report(repo_root: Path, month: int, year: int) -> Path:
    """
    Генерирует месячный отчёт за указанный месяц и год.
    Возвращает путь к созданному файлу отчёта.
    """
    print(f"\n📊 Генерация отчёта за {MONTHS_RU[month - 1]} {year}...")
    
    # Убеждаемся, что папка reports существует
    reports_dir = repo_root / REPORTS_DIR_NAME
    reports_dir.mkdir(exist_ok=True)
    
    # Определяем путь к Python из виртуального окружения или системный
    venv_python = repo_root / ".venv" / "Scripts" / "python.exe"
    if not venv_python.exists():
        venv_python = repo_root / ".venv" / "bin" / "python"
    if not venv_python.exists():
        venv_python = Path(sys.executable)  # Используем текущий интерпретатор
    
    report_script = repo_root / "scripts" / "Monthly_sales_report.py"
    if not report_script.exists():
        raise FileNotFoundError(f"Скрипт генерации отчёта не найден: {report_script}")
    
    # Запускаем генерацию отчёта с параметрами месяца и года
    print("⏳ Запуск генерации отчёта...")
    result = subprocess.run(
        [str(venv_python), str(report_script), "--month", str(month), "--year", str(year)],
        cwd=repo_root,
    )
    
    if result.returncode != 0:
        raise RuntimeError(f"Ошибка при генерации отчёта за {MONTHS_RU[month - 1]} {year}. Код возврата: {result.returncode}")
    
    # Проверяем, что файл создан
    report_path = get_report_path(repo_root, year, month)
    if not report_path.exists():
        raise FileNotFoundError(f"Отчёт не был создан: {report_path}")
    
    print(f"✅ Отчёт успешно создан: {report_path.name}")
    return report_path


def run(
    repo_root: Path,
    min_margin: float = MIN_MARGIN_DEFAULT,
    desired_margin: float = DESIRED_MARGIN_DEFAULT,
) -> None:
    prev_year, prev_month = get_prev_month_year()
    report_path = get_report_path(repo_root, prev_year, prev_month)
    costs_path = repo_root / COSTS_FILENAME

    print(f"Используется отчёт за предыдущий месяц: {MONTHS_RU[prev_month - 1]} {prev_year}")
    
    # Если файл отчёта не существует, генерируем его
    if not report_path.exists():
        print(f"⚠ Файл отчёта не найден: {report_path.name}")
        try:
            report_path = generate_monthly_report(repo_root, prev_month, prev_year)
        except Exception as e:
            raise RuntimeError(
                f"Не удалось сгенерировать отчёт за {MONTHS_RU[prev_month - 1]} {prev_year}.\n"
                f"Ошибка: {e}\n\n"
                f"Попробуйте запустить генерацию отчёта вручную через меню программы."
            )
    
    log_verbose(f"Отчёт: {report_path}")
    total_rate = load_rates_from_report(report_path)
    log_verbose(f"Комиссия+логистика: {total_rate*100:.2f}%")
    df, key_col, cost_col = load_costs_df(costs_path)
    log_verbose(f"Загружено записей: {len(df)}")

    # Получаем текущие цены с Ozon
    log_verbose("Получение цен с Ozon...")
    offer_ids_list = []
    for _, row in df.iterrows():
        art = _artikul_normalize(row.get(key_col))
        if art:
            offer_ids_list.append(art)
    
    prices_map, marketing_prices_map = get_product_prices_from_ozon(offer_ids_list)
    if prices_map:
        print(f"✅ Получено цен для {len(prices_map)} артикулов")
    else:
        print("⚠️ Не удалось получить цены с Ozon (возможно, не настроены API ключи).")
    
    # Получаем информацию об акциях для артикулов
    actions_map, actions_info_list, offer_id_to_product_id = get_actions_for_products(offer_ids_list)
    if actions_map:
        log_verbose(f"Акции: {len(actions_map)} артикулов")
    else:
        print("⚠️ Не удалось получить информацию об акциях.")
        actions_info_list = []
        offer_id_to_product_id = {}
    
    # Создаём маппинг product_id -> offer_id для использования в добавлении товаров
    product_id_to_offer_id: Dict[int, str] = {}
    for offer_id, product_id in offer_id_to_product_id.items():
        product_id_to_offer_id[product_id] = offer_id
    
    # Создаём множество нормализованных offer_id для проверки кандидатов
    offer_ids_set_for_candidates = set()
    for oid in offer_ids_list:
        normalized = _normalize_offer_id(oid)
        if normalized:
            offer_ids_set_for_candidates.add(normalized)
            offer_ids_set_for_candidates.add(oid)

    min_prices = []
    desired_prices = []
    current_prices = []
    marketing_prices = []
    current_margins = []
    
    # Создаём словари для цен в каждой акции: {название_акции: [список цен]}
    action_prices_dicts = {}
    for action_info in actions_info_list:
        action_name = action_info["name"]
        action_prices_dicts[action_name] = []
    for _, row in df.iterrows():
        try:
            cost_val = float(row.get(cost_col, 0) or 0)
        except (TypeError, ValueError):
            cost_val = 0.0
        min_p, des_p = compute_prices(cost_val, total_rate, min_margin, desired_margin)
        min_prices.append(min_p)
        desired_prices.append(des_p)
        
        # Получаем текущую цену с Ozon
        art = _artikul_normalize(row.get(key_col))
        if art:
            # Пробуем найти по исходному артикулу и по нормализованному
            art_normalized = _normalize_offer_id(art)
            current_price = prices_map.get(art) or prices_map.get(art_normalized)
            marketing_price = marketing_prices_map.get(art) or marketing_prices_map.get(art_normalized)
        else:
            current_price = None
            marketing_price = None
        current_prices.append(round(current_price) if current_price is not None else None)
        marketing_prices.append(round(marketing_price) if marketing_price is not None else None)
        
        # Рассчитываем текущую ожидаемую рентабельность
        # Используем цену с акциями, если она есть, иначе обычную цену
        price_for_margin = marketing_price if marketing_price is not None else current_price
        margin = compute_current_margin(price_for_margin, cost_val, total_rate)
        # Преобразуем в проценты и округляем до 2 знаков
        current_margins.append(round(margin * 100, 2) if margin is not None else None)
        
        # Получаем информацию об акциях для этого артикула
        if art:
            art_normalized = _normalize_offer_id(art)
            art_actions = actions_map.get(art) or actions_map.get(art_normalized) or {}
        else:
            art_actions = {}
        
        # Заполняем цены для каждой акции
        for action_info in actions_info_list:
            action_name = action_info["name"]
            action_price = art_actions.get(action_name)
            if action_price is not None:
                action_prices_dicts[action_name].append(round(action_price))
            else:
                action_prices_dicts[action_name].append(None)

    # Удаляем старые колонки с такими именами, если есть
    cols_to_remove = [COL_MIN_PRICE, COL_DESIRED_PRICE, COL_CURRENT_PRICE, COL_MARKETING_PRICE, COL_CURRENT_MARGIN]
    # Также удаляем старые колонки акций (если они есть)
    for action_info in actions_info_list:
        cols_to_remove.append(action_info["name"])
    
    for c in cols_to_remove:
        if c in df.columns:
            df = df.drop(columns=[c])
    
    df[COL_MIN_PRICE] = min_prices
    df[COL_DESIRED_PRICE] = desired_prices
    df[COL_CURRENT_PRICE] = current_prices
    df[COL_MARKETING_PRICE] = marketing_prices
    df[COL_CURRENT_MARGIN] = current_margins
    
    # Сохраняем основной лист без колонок акций
    # Используем ExcelWriter для создания нескольких листов
    if actions_info_list:
        print("\n📋 Создание листа «Акции»...")
        try:
            # Создаём DataFrame для листа акций
            actions_df_data = {}
            
            # Добавляем колонку с артикулами
            actions_df_data[key_col] = df[key_col].values
            
            # Добавляем колонку с минимальной ценой
            actions_df_data[COL_MIN_PRICE] = df[COL_MIN_PRICE].values
            
            # Добавляем колонки для каждой акции
            for action_info in actions_info_list:
                action_name = action_info["name"]
                actions_df_data[action_name] = action_prices_dicts[action_name]
            
            # Создаём DataFrame для листа акций
            actions_df = pd.DataFrame(actions_df_data)
            
            # Сохраняем оба листа через ExcelWriter
            with pd.ExcelWriter(costs_path, engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, sheet_name='Sheet1', index=False)
                actions_df.to_excel(writer, sheet_name='Акции', index=False)
            
            # Применяем условное форматирование
            wb = load_workbook(costs_path)
            
            # Переименовываем Sheet1 в более понятное имя (если нужно)
            if 'Sheet1' in wb.sheetnames:
                wb['Sheet1'].title = 'Основной'
            
            ws_actions = wb['Акции']
            
            # Находим колонки для условного форматирования
            min_price_col_idx = None
            for col_idx, col_name in enumerate(actions_df.columns, start=1):
                if col_name == COL_MIN_PRICE:
                    min_price_col_idx = col_idx
                    break
            
            # Применяем условное форматирование к колонкам акций
            if min_price_col_idx:
                min_price_col_letter = get_column_letter(min_price_col_idx)
                
                for col_idx, col_name in enumerate(actions_df.columns, start=1):
                    # Пропускаем колонки артикула и минимальной цены
                    if col_name == key_col or col_name == COL_MIN_PRICE:
                        continue
                    
                    action_col_letter = get_column_letter(col_idx)
                    
                    # Зелёная заливка: цена в акции >= минимальной цены
                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    green_formula = f"AND({action_col_letter}2<>\"\", {action_col_letter}2>0, {action_col_letter}2>={min_price_col_letter}2)"
                    green_rule = FormulaRule(formula=[green_formula], fill=green_fill, stopIfTrue=False)
                    
                    # Красная заливка: цена в акции < минимальной цены (и не пустая)
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    red_formula = f"AND({action_col_letter}2<>\"\", {action_col_letter}2>0, {action_col_letter}2<{min_price_col_letter}2)"
                    red_rule = FormulaRule(formula=[red_formula], fill=red_fill, stopIfTrue=False)
                    
                    # Применяем правила к диапазону колонки (начиная со 2-й строки, где данные)
                    data_range = f"{action_col_letter}2:{action_col_letter}{len(actions_df) + 1}"
                    
                    # Применяем правила условного форматирования
                    ws_actions.conditional_formatting.add(data_range, red_rule)
                    ws_actions.conditional_formatting.add(data_range, green_rule)
            
            # Удаляем товары из акций, где цена ниже минимальной
            action_name_to_id = {a["name"]: a["id"] for a in actions_info_list}
            if offer_id_to_product_id and action_name_to_id:
                candidates = collect_deactivation_candidates_from_sheet(
                    ws_actions,
                    key_col,
                    COL_MIN_PRICE,
                    action_name_to_id,
                    offer_id_to_product_id,
                )
                if candidates:
                    print("\n🧹 Удаление товаров из акций с ценой ниже минимальной...")
                    for action_id, product_ids in candidates.items():
                        print(f"   Акция {action_id}: удаляем {len(product_ids)} товаров")
                        result = deactivate_products_in_action(action_id, product_ids)
                        removed = result.get("product_ids", []) or []
                        rejected = result.get("rejected", []) or []
                        if removed:
                            print(f"      ✅ Удалено: {len(removed)}")
                        if rejected:
                            print(f"      ⚠️ Не удалено: {len(rejected)}")
                else:
                    print("✅ Товары с ценой ниже минимальной не найдены.")
            else:
                print("⚠️ Недостаточно данных для удаления из акций.")
            
            # Добавляем товары в акции, если можно установить цену в диапазоне [min_price, desired_price]
            print("\n➕ Проверка кандидатов для добавления в акции...")
            
            # Создаём маппинг product_id -> (min_price, desired_price, offer_id)
            product_id_to_prices: Dict[int, tuple[float, float, str]] = {}
            for _, row in df.iterrows():
                art = _artikul_normalize(row.get(key_col))
                if not art:
                    continue
                
                art_normalized = _normalize_offer_id(art)
                product_id = offer_id_to_product_id.get(art_normalized) or offer_id_to_product_id.get(art)
                
                if product_id:
                    min_p = row.get(COL_MIN_PRICE)
                    des_p = row.get(COL_DESIRED_PRICE)
                    if min_p is not None and des_p is not None:
                        try:
                            min_price_val = float(min_p)
                            des_price_val = float(des_p)
                            if min_price_val > 0 and des_price_val > 0:
                                product_id_to_prices[product_id] = (min_price_val, des_price_val, art_normalized)
                        except (TypeError, ValueError):
                            pass
            
            # Для каждой акции проверяем кандидатов
            for action_info in actions_info_list:
                action_id = action_info["id"]
                action_name = action_info["name"]
                
                print(f"   Проверка акции: {action_name} (ID: {action_id})...")
                
                # Получаем кандидатов для этой акции
                candidates = get_action_candidates(action_id, product_id_to_offer_id, offer_ids_set_for_candidates)
                
                if not candidates:
                    continue
                
                # Фильтруем кандидатов: проверяем, можно ли установить цену в диапазоне [min_price, desired_price]
                products_to_add = []
                
                for product_id, product_info in candidates.items():
                    if product_id not in product_id_to_prices:
                        continue
                    
                    min_price, desired_price, offer_id = product_id_to_prices[product_id]
                    
                    # Получаем максимальную цену в акции
                    max_action_price = product_info.get("max_action_price")
                    if max_action_price is None:
                        continue
                    
                    try:
                        max_action_price_val = float(max_action_price)
                    except (TypeError, ValueError):
                        continue
                    
                    # Проверяем, можно ли установить цену в диапазоне [min_price, desired_price]
                    # Цена должна быть не меньше min_price и не больше min(desired_price, max_action_price)
                    target_price = min(desired_price, max_action_price_val)
                    
                    if target_price >= min_price:
                        # Проверяем, не участвует ли уже товар в акции
                        current_action_price = product_info.get("action_price", 0)
                        if current_action_price == 0 or current_action_price is None:
                            # Товар не участвует в акции, добавляем
                            stock = product_info.get("stock", 0) or 0
                            products_to_add.append({
                                "product_id": product_id,
                                "action_price": int(target_price),  # Используем целевую цену
                                "stock": int(stock) if stock else 0
                            })
                
                # Добавляем товары в акцию
                if products_to_add:
                    print(f"      Найдено {len(products_to_add)} товаров для добавления")
                    result = activate_products_in_action(action_id, products_to_add)
                    added = result.get("product_ids", []) or []
                    rejected = result.get("rejected", []) or []
                    if added:
                        print(f"      ✅ Добавлено: {len(added)} товаров")
                    if rejected:
                        print(f"      ⚠️ Не добавлено: {len(rejected)} товаров")
                else:
                    print(f"      Товары для добавления не найдены")

            wb.save(costs_path)
            print(f"✅ Создан лист «Акции» с {len(actions_info_list)} колонками акций")
            print(f"   Применено условное форматирование: 🟢 зелёный если >= минимальной цены, 🔴 красный если < минимальной цены")
        except Exception as e:
            print(f"⚠️ Не удалось создать лист «Акции»: {e}")
            import traceback
            print(traceback.format_exc())
            # Если не удалось создать лист акций, сохраняем хотя бы основной лист
            df.to_excel(costs_path, index=False)
    else:
        # Если нет акций, просто сохраняем основной лист
        df.to_excel(costs_path, index=False)
    
    # Применяем условное форматирование к колонке рентабельности
    try:
        wb = load_workbook(costs_path)
        ws = wb.active
        
        # Находим колонку с рентабельностью
        margin_col_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == COL_CURRENT_MARGIN:
                margin_col_idx = col_idx
                break
        
        if margin_col_idx:
            # Минимальная и желательная рентабельность в процентах
            min_margin_pct = min_margin * 100
            desired_margin_pct = desired_margin * 100
            
            # Зелёная заливка: значение >= min_margin и <= desired_margin
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            green_rule = CellIsRule(
                operator="between",
                formula=[min_margin_pct, desired_margin_pct],
                fill=green_fill
            )
            
            # Красная заливка: значение < min_margin
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_rule = CellIsRule(
                operator="lessThan",
                formula=[min_margin_pct],
                fill=red_fill
            )
            
            # Применяем правила к диапазону колонки (начиная со 2-й строки, где данные)
            margin_col_letter = ws.cell(row=1, column=margin_col_idx).column_letter
            data_range = f"{margin_col_letter}2:{margin_col_letter}{len(df) + 1}"
            
            # Применяем правила условного форматирования к диапазону
            ws.conditional_formatting.add(data_range, green_rule)
            ws.conditional_formatting.add(data_range, red_rule)
            
            wb.save(costs_path)
            print(f"✅ Применено условное форматирование к колонке «{COL_CURRENT_MARGIN}»:")
            print(f"   🟢 Зелёный: рентабельность от {min_margin_pct:.1f}% до {desired_margin_pct:.1f}%")
            print(f"   🔴 Красный: рентабельность меньше {min_margin_pct:.1f}%")
    except Exception as e:
        print(f"⚠️ Не удалось применить условное форматирование: {e}")
    
    # Применяем условное форматирование к колонке "Текущая цена"
    try:
        wb = load_workbook(costs_path)
        ws = wb.active
        
        # Находим колонки с текущей ценой и минимальной ценой
        current_price_col_idx = None
        min_price_col_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == COL_CURRENT_PRICE:
                current_price_col_idx = col_idx
            elif cell.value == COL_MIN_PRICE:
                min_price_col_idx = col_idx
        
        if current_price_col_idx and min_price_col_idx:
            current_price_col_letter = ws.cell(row=1, column=current_price_col_idx).column_letter
            min_price_col_letter = ws.cell(row=1, column=min_price_col_idx).column_letter
            
            # Зелёная заливка: текущая цена >= минимальной цены
            # Формула использует относительные ссылки: для каждой строки сравнивается текущая цена с минимальной ценой в той же строке
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            # Формула написана относительно первой ячейки диапазона (строка 2)
            # Excel автоматически применит её ко всем строкам с относительными ссылками
            green_formula = f"AND({current_price_col_letter}2<>\"\", {current_price_col_letter}2>0, {current_price_col_letter}2>={min_price_col_letter}2)"
            green_rule = FormulaRule(formula=[green_formula], fill=green_fill, stopIfTrue=False)
            
            # Красная заливка: текущая цена < минимальной цены (и не пустая)
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_formula = f"AND({current_price_col_letter}2<>\"\", {current_price_col_letter}2>0, {current_price_col_letter}2<{min_price_col_letter}2)"
            red_rule = FormulaRule(formula=[red_formula], fill=red_fill, stopIfTrue=False)
            
            # Применяем правила к диапазону колонки (начиная со 2-й строки, где данные)
            data_range = f"{current_price_col_letter}2:{current_price_col_letter}{len(df) + 1}"
            
            # Применяем правила условного форматирования к диапазону
            # Красное правило применяем первым, чтобы оно имело приоритет
            ws.conditional_formatting.add(data_range, red_rule)
            ws.conditional_formatting.add(data_range, green_rule)
            
            wb.save(costs_path)
            print(f"✅ Применено условное форматирование к колонке «{COL_CURRENT_PRICE}»:")
            print(f"   🟢 Зелёный: текущая цена >= минимальной цены")
            print(f"   🔴 Красный: текущая цена < минимальной цены")
    except Exception as e:
        print(f"⚠️ Не удалось применить условное форматирование к колонке «{COL_CURRENT_PRICE}»: {e}")
    
    # Применяем условное форматирование к колонке "Цена с учётом акций и скидок"
    try:
        wb = load_workbook(costs_path)
        ws = wb.active
        
        # Находим колонки с ценой с акциями и минимальной ценой
        marketing_price_col_idx = None
        min_price_col_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == COL_MARKETING_PRICE:
                marketing_price_col_idx = col_idx
            elif cell.value == COL_MIN_PRICE:
                min_price_col_idx = col_idx
        
        if marketing_price_col_idx and min_price_col_idx:
            marketing_price_col_letter = ws.cell(row=1, column=marketing_price_col_idx).column_letter
            min_price_col_letter = ws.cell(row=1, column=min_price_col_idx).column_letter
            
            # Зелёная заливка: цена с акциями >= минимальной цены
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            green_formula = f"AND({marketing_price_col_letter}2<>\"\", {marketing_price_col_letter}2>0, {marketing_price_col_letter}2>={min_price_col_letter}2)"
            green_rule = FormulaRule(formula=[green_formula], fill=green_fill, stopIfTrue=False)
            
            # Красная заливка: цена с акциями < минимальной цены (и не пустая)
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_formula = f"AND({marketing_price_col_letter}2<>\"\", {marketing_price_col_letter}2>0, {marketing_price_col_letter}2<{min_price_col_letter}2)"
            red_rule = FormulaRule(formula=[red_formula], fill=red_fill, stopIfTrue=False)
            
            # Применяем правила к диапазону колонки (начиная со 2-й строки, где данные)
            data_range = f"{marketing_price_col_letter}2:{marketing_price_col_letter}{len(df) + 1}"
            
            # Применяем правила условного форматирования к диапазону
            # Красное правило применяем первым, чтобы оно имело приоритет
            ws.conditional_formatting.add(data_range, red_rule)
            ws.conditional_formatting.add(data_range, green_rule)
            
            wb.save(costs_path)
            print(f"✅ Применено условное форматирование к колонке «{COL_MARKETING_PRICE}»:")
            print(f"   🟢 Зелёный: цена с акциями >= минимальной цены")
            print(f"   🔴 Красный: цена с акциями < минимальной цены")
    except Exception as e:
        print(f"⚠️ Не удалось применить условное форматирование к колонке «{COL_MARKETING_PRICE}»: {e}")
    
    action_names_str = ", ".join([f"«{a['name']}»" for a in actions_info_list])
    if action_names_str:
        print(f"В файл {COSTS_FILENAME} добавлены колонки «{COL_MIN_PRICE}», «{COL_DESIRED_PRICE}», «{COL_CURRENT_PRICE}», «{COL_MARKETING_PRICE}», «{COL_CURRENT_MARGIN}» и колонки акций: {action_names_str}.")
    else:
        print(f"В файл {COSTS_FILENAME} добавлены колонки «{COL_MIN_PRICE}», «{COL_DESIRED_PRICE}», «{COL_CURRENT_PRICE}», «{COL_MARKETING_PRICE}» и «{COL_CURRENT_MARGIN}».")
    print("Готово.")


def get_discount_requests(status: str = "ALL", limit: int = 50) -> List[Dict[str, Any]]:
    """
    Получает список заявок на скидку через API /v2/actions/discounts-task/list.
    
    Args:
        status: Статус заявки ("ALL", "NEW", "APPROVED", "DECLINED")
        limit: Максимальное количество заявок на странице (до 50)
    
    Returns:
        Список заявок на скидку
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        print("⚠️ OZON_CLIENT_ID или OZON_API_KEY не настроены.")
        return []
    
    url = "https://api-seller.ozon.ru/v2/actions/discounts-task/list"
    all_tasks = []
    last_id = None
    
    while True:
        payload = {
            "status": status,
            "limit": min(limit, 50)  # API ограничивает до 50
        }
        
        if last_id is not None:
            payload["last_id"] = last_id
        
        try:
            response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            tasks = data.get("tasks", [])
            if not tasks:
                break
            
            all_tasks.extend(tasks)
            
            # Проверяем, есть ли ещё страницы
            # Если получили меньше запрошенного лимита, значит это последняя страница
            if len(tasks) < limit:
                break
            
            # Используем ID последней заявки как last_id для следующей страницы
            if tasks:
                last_id = tasks[-1].get("id")
                if last_id is None:
                    break
            else:
                break
            
        except requests.exceptions.RequestException as e:
            print(f"⚠️ Ошибка при получении заявок на скидку: {e}")
            if hasattr(e, 'response') and e.response is not None:
                try:
                    error_data = e.response.json()
                    print(f"   Детали ошибки: {error_data}")
                except:
                    print(f"   Ответ сервера: {e.response.text[:200]}")
            break
        except Exception as e:
            print(f"⚠️ Неожиданная ошибка при получении заявок на скидку: {e}")
            break
    
    return all_tasks


def approve_discount_requests(tasks: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Одобряет заявки на скидку через API /v1/actions/discounts-task/approve.
    
    Args:
        tasks: Список заявок для одобрения. Каждая заявка должна содержать:
            - id: идентификатор заявки
            - approved_price: одобренная цена
            - approved_quantity_min: минимальное одобренное количество (опционально)
            - approved_quantity_max: максимальное одобренное количество (опционально)
            - seller_comment: комментарий продавца (опционально)
    
    Returns:
        Результат обработки с полями success_count, fail_count, fail_details
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        return {"success_count": 0, "fail_count": len(tasks), "fail_details": []}
    
    if not tasks:
        return {"success_count": 0, "fail_count": 0, "fail_details": []}
    
    url = "https://api-seller.ozon.ru/v1/actions/discounts-task/approve"
    payload = {"tasks": tasks}
    
    try:
        response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        result = data.get("result", {})
        return {
            "success_count": result.get("success_count", 0),
            "fail_count": result.get("fail_count", 0),
            "fail_details": result.get("fail_details", [])
        }
    except requests.exceptions.RequestException as e:
        print(f"⚠️ Ошибка при одобрении заявок на скидку: {e}")
        if hasattr(e, 'response') and e.response is not None:
            try:
                error_data = e.response.json()
                print(f"   Детали ошибки: {error_data}")
            except:
                print(f"   Ответ сервера: {e.response.text[:200]}")
        return {"success_count": 0, "fail_count": len(tasks), "fail_details": []}
    except Exception as e:
        print(f"⚠️ Неожиданная ошибка при одобрении заявок на скидку: {e}")
        return {"success_count": 0, "fail_count": len(tasks), "fail_details": []}


def decline_discount_requests(tasks: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Отклоняет заявки на скидку через API /v1/actions/discounts-task/decline.
    
    Args:
        tasks: Список заявок для отклонения. Каждая заявка должна содержать:
            - id: идентификатор заявки
            - seller_comment: комментарий продавца (опционально)
    
    Returns:
        Результат обработки с полями success_count, fail_count, fail_details
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        return {"success_count": 0, "fail_count": len(tasks), "fail_details": []}
    
    if not tasks:
        return {"success_count": 0, "fail_count": 0, "fail_details": []}
    
    url = "https://api-seller.ozon.ru/v1/actions/discounts-task/decline"
    payload = {"tasks": tasks}
    
    try:
        response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        result = data.get("result", {})
        return {
            "success_count": result.get("success_count", 0),
            "fail_count": result.get("fail_count", 0),
            "fail_details": result.get("fail_details", [])
        }
    except requests.exceptions.RequestException as e:
        print(f"⚠️ Ошибка при отклонении заявок на скидку: {e}")
        if hasattr(e, 'response') and e.response is not None:
            try:
                error_data = e.response.json()
                print(f"   Детали ошибки: {error_data}")
            except:
                print(f"   Ответ сервера: {e.response.text[:200]}")
        return {"success_count": 0, "fail_count": len(tasks), "fail_details": []}
    except Exception as e:
        print(f"⚠️ Неожиданная ошибка при отклонении заявок на скидку: {e}")
        return {"success_count": 0, "fail_count": len(tasks), "fail_details": []}


def get_sku_to_offer_id_mapping(skus: List[int]) -> Dict[int, str]:
    """
    Создаёт маппинг SKU (идентификатор товара в системе Ozon) -> offer_id.
    Сначала использует /v3/product/info/list с параметром sku (как в заявках на скидку),
    затем при необходимости — /v2/product/info по product_id и v5/product/info/prices.
    
    Args:
        skus: Список SKU из API (например, из discounts-task/list).
    
    Returns:
        Словарь {sku: offer_id}
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        print("⚠️ OZON_CLIENT_ID или OZON_API_KEY не настроены.")
        return {}
    
    if not skus:
        return {}
    
    # 1) Запрос по SKU — как возвращает API заявок на скидку (v3/product/info/list)
    log_verbose(f"Запрос offer_id по SKU через v3/product/info/list ({len(skus)} шт.)...")
    mapping = get_offer_ids_by_skus(skus)
    if len(mapping) == len(skus):
        return mapping
    
    # 2) Для недостающих пробуем v2/product/info (sku как product_id)
    missing = [s for s in skus if s not in mapping]
    log_verbose(f"Обработка {len(missing)} SKU через v2/product/info...")
    for sku in missing:
        offer_id = get_product_info_by_id(sku, debug=VERBOSE)
        if offer_id:
            mapping[sku] = offer_id
            log_verbose(f"  SKU {sku} -> {offer_id}")
    
    if len(mapping) < len(skus):
        failed_count = len(skus) - len(mapping)
        print(f"   ⚠️ Не удалось получить offer_id для {failed_count} из {len(skus)} SKU")
    
    # 3) Альтернатива: v5/product/info/prices — перебор товаров по product_id
    if len(mapping) < len(skus):
        missing_skus = [sku for sku in skus if sku not in mapping]
        print(f"   Попытка альтернативного способа для {len(missing_skus)} SKU...")
        print(f"   Ищем SKU: {missing_skus}")
        
        try:
            url = "https://api-seller.ozon.ru/v5/product/info/prices"
            cursor = ""
            initial_mapping_count = len(mapping)
            # Создаём множества для быстрого поиска (в разных типах)
            skus_set_int = set()
            skus_set_str = set()
            for sku in missing_skus:
                try:
                    skus_set_int.add(int(sku))
                except (TypeError, ValueError):
                    pass
                skus_set_str.add(str(sku))
            
            max_iterations = 100  # Ограничение на количество итераций для безопасности
            iteration = 0
            total_items_checked = 0
            found_product_ids = []
            
            # Получаем товары порциями и ищем нужные SKU
            while len(mapping) < len(skus) and iteration < max_iterations:
                iteration += 1
                payload = {
                    "cursor": cursor,
                    "filter": {
                        "visibility": "ALL"
                    },
                    "limit": 1000
                }
                
                response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
                response.raise_for_status()
                data = response.json()
                
                items = data.get("items", [])
                cursor = data.get("cursor", "")
                
                if not items:
                    break
                
                total_items_checked += len(items)
                
                for item in items:
                    product_id = item.get("product_id")
                    offer_id_raw = item.get("offer_id", "")
                    
                    if not product_id or not offer_id_raw:
                        continue
                    
                    # Приводим product_id к int для сравнения
                    product_id_int = None
                    try:
                        product_id_int = int(product_id) if product_id is not None else None
                    except (TypeError, ValueError):
                        pass
                    
                    # Ищем соответствующий SKU из исходного списка
                    matching_sku = None
                    
                    # Проверяем по int значению
                    if product_id_int is not None and product_id_int in skus_set_int:
                        # Находим исходный SKU с таким значением
                        for sku in missing_skus:
                            try:
                                if int(sku) == product_id_int:
                                    matching_sku = sku
                                    break
                            except (TypeError, ValueError):
                                pass
                    
                    # Если не нашли, пробуем по строковому значению
                    if matching_sku is None and str(product_id) in skus_set_str:
                        for sku in missing_skus:
                            if str(sku) == str(product_id):
                                matching_sku = sku
                                break
                    
                    # Если нашли совпадение и ещё не добавили в маппинг
                    if matching_sku is not None and matching_sku not in mapping:
                        offer_id_normalized = _normalize_offer_id(str(offer_id_raw))
                        mapping[matching_sku] = offer_id_normalized
                        found_product_ids.append((matching_sku, product_id, offer_id_normalized))
                        if len(mapping) >= len(skus):
                            break
                
                # Если нашли все нужные SKU или нет курсора для следующей страницы
                if len(mapping) >= len(skus) or not cursor:
                    break
            
            newly_found = len(mapping) - initial_mapping_count
            if newly_found > 0:
                print(f"   ✅ Альтернативный способ: найдено ещё {newly_found} из {len(missing_skus)} SKU")
                if len(missing_skus) <= 5:  # Показываем детали для небольшого количества
                    for sku, pid, oid in found_product_ids:
                        print(f"      SKU {sku} -> product_id {pid} -> offer_id {oid}")
            elif len(mapping) == 0:
                print(f"   ⚠️ Альтернативный способ: не удалось найти ни одного SKU")
                print(f"      Проверено товаров: {total_items_checked}")
                print(f"      Искали SKU: {missing_skus}")
                if total_items_checked > 0 and len(missing_skus) <= 5:
                    print(f"      Возможно, эти товары отсутствуют в вашем каталоге или имеют другой product_id")
        except Exception as e:
            print(f"   ⚠️ Альтернативный способ не сработал: {e}")
            import traceback
            if len(skus) <= 3:  # Показываем полный traceback только для небольшого количества SKU
                print(traceback.format_exc())
    
    return mapping


def get_offer_id_to_sku_mapping(offer_ids: List[str]) -> Dict[str, int]:
    """
    Получает SKU (product_id в Ozon) по списку offer_id через API v3/product/list.
    Нужно для запроса /v1/analytics/stocks, который принимает skus.
    
    Args:
        offer_ids: Список артикулов (offer_id).
    
    Returns:
        Словарь {offer_id_normalized: sku (int)}
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY or not offer_ids:
        return {}
    url = "https://api-seller.ozon.ru/v3/product/list"
    mapping: Dict[str, int] = {}
    seen = set()
    normalized_list = []
    for oid in offer_ids:
        n = _normalize_offer_id(str(oid).strip()) if oid else ""
        if n and n not in seen:
            seen.add(n)
            normalized_list.append(n)
    if not normalized_list:
        return {}
    for i in range(0, len(normalized_list), 100):
        batch = normalized_list[i : i + 100]
        try:
            payload = {
                "filter": {"offer_id": batch, "visibility": "ALL"},
                "limit": 1000,
            }
            response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
            response.raise_for_status()
            data = response.json()
            items = data.get("result", {}).get("items", data.get("items", []))
            if not items and isinstance(data.get("result"), list):
                items = data["result"]
            for item in items:
                offer_id_raw = item.get("offer_id") or item.get("offerId")
                # v3/product/list возвращает product_id; для /v1/analytics/stocks используем его как sku
                sku_val = item.get("product_id") if item.get("product_id") is not None else item.get("sku")
                if offer_id_raw is not None and sku_val is not None:
                    try:
                        oid_n = _normalize_offer_id(str(offer_id_raw))
                        sku_int = int(sku_val)
                        if oid_n:
                            mapping[oid_n] = sku_int
                    except (TypeError, ValueError):
                        pass
        except requests.exceptions.RequestException as e:
            log_verbose(f"get_offer_id_to_sku_mapping: {e}")
        except Exception as e:
            log_verbose(f"get_offer_id_to_sku_mapping: {e}")
    return mapping


def get_fbo_stocks_analytics(skus: List[int]) -> Dict[int, int]:
    """
    Получает аналитику по остаткам FBO через POST /v1/analytics/stocks.
    Возвращает доступный остаток к продаже (available_stock_count) по SKU.
    При нескольких складах/кластерах суммирует available_stock_count по sku.
    
    Args:
        skus: Список SKU (до 100 за запрос).
    
    Returns:
        Словарь {sku: available_stock_count}
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY or not skus:
        return {}
    url = "https://api-seller.ozon.ru/v1/analytics/stocks"
    result: Dict[int, int] = {}
    for i in range(0, len(skus), 100):
        batch = [int(s) for s in skus[i : i + 100]]
        try:
            payload = {"skus": [str(s) for s in batch]}
            response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
            response.raise_for_status()
            data = response.json()
            items = data.get("result", {}).get("items", data.get("items", []))
            if not items and isinstance(data.get("result"), list):
                items = data["result"]
            for item in items:
                sku_val = item.get("sku")
                if sku_val is None:
                    sku_val = item.get("product_id")
                avail = item.get("available_stock_count")
                if sku_val is not None:
                    try:
                        sku_int = int(sku_val)
                        result[sku_int] = result.get(sku_int, 0) + int(avail or 0)
                    except (TypeError, ValueError):
                        pass
        except requests.exceptions.RequestException as e:
            log_verbose(f"get_fbo_stocks_analytics: {e}")
        except Exception as e:
            log_verbose(f"get_fbo_stocks_analytics: {e}")
    return result


def get_fbo_warehouse_ids() -> set:
    """
    Возвращает только ID складов Ozon FBO (Fulfillment by Ozon).
    Используется только /v1/warehouse/fbo/list — общий /v1/warehouse/list не вызываем,
    т.к. там и FBO, и FBS (ваш склад), нужен только FBO.
    """
    ids: set = set()
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        return ids
    try:
        url_fbo = "https://api-seller.ozon.ru/v1/warehouse/fbo/list"
        for search in ("Ozon", "FBO ", "склад", "Склад", "Kazan", "Москва"):
            r = requests.post(url_fbo, headers=OZON_HEADERS, json={"search": search}, timeout=15)
            r.raise_for_status()
            data = r.json()
            result = data.get("result", data.get("search", data.get("warehouses", [])))
            if isinstance(result, list):
                for w in result:
                    wid = w.get("warehouse_id") or w.get("id") or w.get("warehouseId")
                    if wid is not None:
                        try:
                            ids.add(int(wid))
                        except (TypeError, ValueError):
                            pass
            if ids:
                break
    except Exception as e:
        log_verbose(f"get_fbo_warehouse_ids: {e}")
    return ids


def get_fbo_stocks_by_offer_ids(offer_ids: List[str]) -> Dict[str, int]:
    """
    Получает только активный сток (доступно к продаже) на складах Ozon FBO
    через POST /v4/product/info/stocks. Учитываются только склады FBO; по каждому
    складу берётся «активный» остаток (present / free_to_sell_amount).
    Ключи результата — нормализованные offer_id.
    
    Args:
        offer_ids: Список артикулов (offer_id).
    
    Returns:
        Словарь {offer_id_normalized: total_active_stock_fbo}
    """
    if not OZON_CLIENT_ID or not OZON_API_KEY or not offer_ids:
        return {}
    fbo_warehouse_ids = get_fbo_warehouse_ids()
    url = "https://api-seller.ozon.ru/v4/product/info/stocks"
    result: Dict[str, int] = {}
    seen = set()
    normalized_list = []
    for oid in offer_ids:
        n = _normalize_offer_id(str(oid).strip()) if oid else ""
        if n and n not in seen:
            seen.add(n)
            normalized_list.append(n)
    if not normalized_list:
        return {}
    for i in range(0, len(normalized_list), 1000):
        batch = normalized_list[i : i + 1000]
        last_id = ""
        try:
            while True:
                payload = {
                    "filter": {"offer_id": batch, "visibility": "ALL"},
                    "limit": 1000,
                }
                if last_id:
                    payload["last_id"] = last_id
                response = requests.post(url, headers=OZON_HEADERS, json=payload, timeout=30)
                response.raise_for_status()
                data = response.json()
                res = data.get("result", data)
                items = res.get("items", []) if isinstance(res, dict) else []
                last_id = res.get("last_id", "") if isinstance(res, dict) else ""
                # Учитываем только склады FBO; если список FBO пустой — остатки не суммируем
                if not fbo_warehouse_ids:
                    break
                for item in items:
                    offer_id_raw = item.get("offer_id") or item.get("offerId")
                    if offer_id_raw is None:
                        continue
                    oid_n = _normalize_offer_id(str(offer_id_raw))
                    if not oid_n:
                        continue
                    # Только активный сток на складах из fbo_warehouse_ids (только FBO, не FBS)
                    q = 0
                    stocks_arr = item.get("stocks")
                    if isinstance(stocks_arr, list):
                        for s in stocks_arr:
                            if not isinstance(s, dict):
                                continue
                            s_wh = s.get("warehouse_id") or s.get("warehouseId")
                            if s_wh is not None:
                                try:
                                    if int(s_wh) not in fbo_warehouse_ids:
                                        continue
                                except (TypeError, ValueError):
                                    continue
                            q += int(s.get("present") or s.get("free_to_sell_amount") or s.get("available") or 0)
                    else:
                        item_wh = item.get("warehouse_id") or item.get("warehouseId")
                        if item_wh is not None:
                            try:
                                if int(item_wh) not in fbo_warehouse_ids:
                                    continue
                            except (TypeError, ValueError):
                                continue
                        stock_val = item.get("stock") or item.get("present") or item.get("free_to_sell_amount") or item.get("available")
                        try:
                            q = int(stock_val or 0)
                        except (TypeError, ValueError):
                            q = 0
                    if q > 0:
                        result[oid_n] = result.get(oid_n, 0) + q
                if not last_id:
                    break
        except requests.exceptions.RequestException as e:
            log_verbose(f"get_fbo_stocks_by_offer_ids: {e}")
            break
        except Exception as e:
            log_verbose(f"get_fbo_stocks_by_offer_ids: {e}")
            break
    return result


def main():
    import argparse
    script_dir = Path(__file__).resolve().parent
    repo_root = script_dir.parent

    parser = argparse.ArgumentParser(description="Рассчёт рекомендуемых цен продажи по артикулам (минимальная и желательная).")
    args = parser.parse_args()

    # Минимальная маржа захардкожена на 0.25
    min_margin = MIN_MARGIN_DEFAULT
    desired_margin = DESIRED_MARGIN_DEFAULT

    try:
        run(repo_root, min_margin=min_margin, desired_margin=desired_margin)
    except FileNotFoundError as e:
        print(f"Ошибка: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"Ошибка: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
