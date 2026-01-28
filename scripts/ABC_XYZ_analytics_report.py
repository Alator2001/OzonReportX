# merge_excel_columns.py
import math
import os
import re
import sys
import argparse
import warnings
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning)

# Названия месяцев на русском (для парсинга имён файлов вида "Октябрь 2025.xlsx")
MONTHS_RU = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
]

# Канонические названия столбцов
CANON = {
    "artikul": "Артикул",
    "tsena_prodazhi": "Цена продажи",
    "kolichestvo_sht": "Количество шт.",
    "pribyl": "Прибыль",
    "data_otgruzki": "Дата отгрузки",
}

# Возможные варианты написания (регистр/пробелы/точки/синонимы)
ALIASES: Dict[str, List[str]] = {
    "artikul": ["артикул", "артикулы", "sku", "код", "код товара", "код/артикул"],
    "tsena_prodazhi": ["цена продажи", "цена", "продажная цена", "стоимость", "sale price"],
    "kolichestvo_sht": ["количество шт.", "кол-во", "количество", "шт", "шт.", "кол-во шт.", "qty"],
    "pribyl": ["прибыль", "маржа", "доход", "profit"],
    "data_otgruzki": ["дата отгрузки", "отгрузка", "дата поставки", "ship date", "дата"],
}

# Расширения Excel
EXCEL_EXT = {".xlsx", ".xlsm", ".xls"}

# Листы с заказами — только их читаем. Листы "Кампании", итоговые показатели и т.п. игнорируем.
ORDER_SHEET_NAMES = ("Заказы", "Orders")

# Папка для отчётов ABC&XYZ (имя файла = "Ранний месяц год-Поздний месяц год.xlsx")
ABC_XYZ_OUTPUT_DIR = "ABC&XYZ reports"

# Только эти колонки попадают в итоговый файл ABC&XYZ
OUTPUT_COLUMNS = [
    "Артикул", "Цена продажи", "Количество шт.", "Прибыль", "Дата отгрузки"
]

# В объединённый файл попадают только заказы с этим статусом
DELIVERED_STATUS = "delivered"


def parse_filename_to_month_year(stem: str) -> Optional[Tuple[int, int]]:
    """
    Парсит имя файла без расширения вида «Октябрь 2025» → (2025, 10).
    Возвращает (year, month) или None, если не удалось распознать.
    """
    stem = (stem or "").strip()
    for month_num, month_name in enumerate(MONTHS_RU, start=1):
        # «Октябрь 2025» или «Октябрь  2025»
        m = re.match(r"^" + re.escape(month_name) + r"\s+(\d{4})\s*$", stem, re.IGNORECASE)
        if m:
            try:
                year = int(m.group(1))
                if 2000 <= year <= 2100:
                    return (year, month_num)
            except ValueError:
                pass
    return None


def format_month_year(year: int, month: int) -> str:
    """Форматирует (year, month) в строку «Месяц год», например «Октябрь 2025»."""
    if 1 <= month <= 12:
        return f"{MONTHS_RU[month - 1]} {year}"
    return f"{year}-{month:02d}"


def _artikul_to_number(v):
    """Преобразует значение артикула в число, если возможно; иначе возвращает как есть."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return v
    s = str(v).strip()
    if not s:
        return v
    try:
        n = float(s.replace(",", "."))
        return int(n) if n == int(n) else n
    except (ValueError, TypeError):
        return v


def norm(s: str) -> str:
    """Нормализация заголовка: нижний регистр, убираем пробелы по краям и двойные пробелы, точки."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = " ".join(s.split())       # сжать повторные пробелы
    s = s.replace(".", "")        # убрать точки (часто пишут "шт.")
    return s

def build_reverse_map() -> Dict[str, str]:
    """Карта из нормализованного псевдонима к ключу CANON."""
    rmap = {}
    for key, variants in ALIASES.items():
        for v in variants:
            rmap[norm(v)] = key
    # добавим сами канонические названия
    for key, title in CANON.items():
        rmap[norm(title)] = key
    return rmap

REV = build_reverse_map()

def find_columns(df: pd.DataFrame) -> Dict[str, str]:
    """
    Возвращает сопоставление {канонический_ключ -> реальное_имя_столбца_в_df}
    Например: {"artikul": "Артикул", ...}
    """
    mapping = {}
    # Если вдруг MultiIndex в колонках — сплющим
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [" ".join([str(x) for x in tup if pd.notna(x)]).strip() for tup in df.columns]

    for col in df.columns:
        k = REV.get(norm(col))
        if k and k not in mapping:
            mapping[k] = col
    return mapping

def read_all_sheets(path: str) -> Dict[str, pd.DataFrame]:
    """Читает ВСЕ листы книги в dict {sheet_name: df}. Для .xls нужна библиотека xlrd."""
    try:
        xls = pd.ExcelFile(path, engine=None)  # pandas сам подберёт движок (openpyxl/xlrd)
        dfs = {sheet: xls.parse(sheet) for sheet in xls.sheet_names}
        return dfs
    except Exception as e:
        raise RuntimeError(f"Не удалось открыть файл '{path}': {e}")

def merge_folder(
    input_dir: str,
    output_path: Optional[str] = None,
    output_dir: Optional[str] = None,
    from_month: Optional[int] = None,
    from_year: Optional[int] = None,
    to_month: Optional[int] = None,
    to_year: Optional[int] = None,
) -> None:
    """
    Объединяет заказы из всех Excel-отчётов в папке в один файл.
    Берёт только листы заказов (ORDER_SHEET_NAMES); листы «Кампании», итоговые показатели и т.п. игнорируются.
    В объединённый файл попадают только заказы со статусом «delivered».
    В итоговом файле только колонки: Артикул, Цена продажи, Количество шт., Прибыль, Дата отгрузки.
    Если задан output_dir (например «ABC&XYZ reports»), файл сохраняется туда с именем
    «Ранний месяц год-Поздний месяц год.xlsx» по диапазону месяцев из имён исходных файлов.
    Если заданы from_month, from_year, to_month, to_year — обрабатываются только файлы с именами
    вида «Месяц Год.xlsx», попадающие в этот диапазон включительно.
    """
    rows = []
    report_missing = []   # листы, где нет всех нужных колонок
    report_partial = []   # листы, где нашли часть колонок

    all_files = [
        f for f in os.listdir(input_dir)
        if os.path.isfile(os.path.join(input_dir, f))
        and os.path.splitext(f)[1].lower() in EXCEL_EXT
        and not f.startswith("~$")
    ]
    # Фильтр по диапазону месяцев: только файлы «Месяц Год.xlsx» в [from_..to_]
    if from_month is not None and from_year is not None and to_month is not None and to_year is not None:
        from_ym = (from_year, from_month)
        to_ym = (to_year, to_month)
        files = sorted(
            f for f in all_files
            if (parsed := parse_filename_to_month_year(os.path.splitext(f)[0])) is not None
            and from_ym <= parsed <= to_ym
        )
    else:
        files = sorted(all_files)

    if not files:
        print("В папке не найдено Excel-файлов.")
        return

    # Если указана папка ABC&XYZ — создаём её и формируем имя файла по диапазону месяцев
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        date_range: List[Tuple[int, int]] = []
        for f in files:
            stem = os.path.splitext(f)[0]
            parsed = parse_filename_to_month_year(stem)
            if parsed:
                date_range.append(parsed)
        if date_range:
            early_y, early_m = min(date_range, key=lambda x: (x[0], x[1]))
            late_y, late_m = max(date_range, key=lambda x: (x[0], x[1]))
            name = f"{format_month_year(early_y, early_m)}-{format_month_year(late_y, late_m)}.xlsx"
        else:
            name = "Объединённые заказы.xlsx"
        output_path = os.path.join(output_dir, name)
    elif not output_path:
        output_path = "merged.xlsx"

    print(f"📂 Найдено файлов: {len(files)}. Объединяем только листы «Заказы», только заказы со статусом «delivered».")
    for fname in files:
        fpath = os.path.join(input_dir, fname)
        try:
            xls = pd.ExcelFile(fpath, engine=None)
        except Exception as e:
            print(f"⚠️ Не удалось открыть файл '{fname}': {e}")
            continue

        # Обрабатываем только листы заказов; «Кампании» и прочие листы пропускаем
        for sheet_name in (s for s in xls.sheet_names if s in ORDER_SHEET_NAMES):
            try:
                df = xls.parse(sheet_name)
            except Exception as e:
                report_missing.append((fname, sheet_name, str(e)))
                continue
            if df is None or df.empty:
                report_missing.append((fname, sheet_name, "лист пустой"))
                continue

            # Оставляем только заказы со статусом «delivered»
            status_col = None
            for c in df.columns:
                if norm(str(c)) in ("статус", "status"):
                    status_col = c
                    break
            if status_col is not None:
                df = df[
                    df[status_col].astype(str).str.strip().str.lower() == DELIVERED_STATUS
                ].copy()
            if df.empty:
                continue

            col_map = find_columns(df)
            found_keys = set(col_map.keys())
            required_keys = set(CANON.keys())

            if not found_keys:
                report_missing.append((fname, sheet_name, "ни один столбец не найден"))
                continue

            if found_keys != required_keys:
                missing = required_keys - found_keys
                # Если вообще ничего не найдено — уже учтено выше; здесь частичное совпадение
                if missing:
                    report_partial.append((fname, sheet_name, f"нет столбцов: {', '.join(CANON[k] for k in missing)}"))

            # Берём только найденные столбцы, переименовываем в канон
            use_cols = {col_map[k]: CANON[k] for k in found_keys}
            sub = df[list(use_cols.keys())].rename(columns=use_cols)

            # Приведём типы слегка (опционально)
            # Даты
            if "Дата отгрузки" in sub.columns:
                sub["Дата отгрузки"] = pd.to_datetime(sub["Дата отгрузки"], errors="coerce").dt.date
            # Числа
            for num_col in ["Цена продажи", "Количество шт.", "Прибыль"]:
                if num_col in sub.columns:
                    sub[num_col] = pd.to_numeric(sub[num_col], errors="coerce")

            # Добавим источник
            sub["Источник файл"] = fname
            sub["Лист"] = sheet_name

            rows.append(sub)

    if not rows:
        print("Нечего объединять — нужные столбцы не найдены ни в одном листе.")
        if report_missing:
            print("\nОтчёт по пропускам:")
            for f, s, msg in report_missing:
                print(f"- {f} / {s}: {msg}")
        return

    merged = pd.concat(rows, ignore_index=True, sort=False)

    # В итоговый файл попадают только колонки для ABC&XYZ: Артикул, Цена продажи, Количество шт., Прибыль, Дата отгрузки
    export_cols = [c for c in OUTPUT_COLUMNS if c in merged.columns]
    merged_export = merged[export_cols].copy() if export_cols else merged.copy()

    # Артикулы — как число, где возможно
    if "Артикул" in merged_export.columns:
        merged_export["Артикул"] = merged_export["Артикул"].apply(_artikul_to_number)

    def _set_artikul_number_format(ws, col_a1: int = 1, data_rows: int = 0):
        """Форматирует колонку col_a1 как число (формат «0») для строк данных."""
        for r in range(2, 2 + data_rows):
            ws.cell(row=r, column=col_a1).number_format = "0"

    # Сохраняем (pivot_abc, pivot_xyz — для листа «Итог», строятся в блоках ниже)
    pivot_abc = None
    pivot_xyz = None
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        merged_export.to_excel(writer, sheet_name="Заказы", index=False)
        _set_artikul_number_format(writer.sheets["Заказы"], 1, len(merged_export))

        # Лист ABC: сводная таблица по Заказам — строки: Артикул, значения: Сумма по полю Прибыль; артикулы отсортированы по сумме прибыли по убыванию
        if "Артикул" in merged_export.columns and "Прибыль" in merged_export.columns:
            pivot_abc = merged_export.pivot_table(
                index="Артикул",
                values="Прибыль",
                aggfunc="sum",
            )
            pivot_abc = pivot_abc.sort_values("Прибыль", ascending=False)
            pivot_abc.columns = ["Сумма по полю Прибыль"]
            total_profit = pivot_abc["Сумма по полю Прибыль"].sum()
            # Колонка C: Сумма по полю Прибыль / Общая прибыль для каждого артикула
            if total_profit != 0:
                pivot_abc["Доля в общей прибыли"] = pivot_abc["Сумма по полю Прибыль"] / total_profit
            else:
                pivot_abc["Доля в общей прибыли"] = 0.0
            # Колонка D: Накопительная доля — первый артикул = своя доля, второй = своя + предыдущая, и т.д.
            pivot_abc["Накопительная доля"] = pivot_abc["Доля в общей прибыли"].cumsum()
            # Колонка E: Оценка — ЕСЛИ(Накопительная доля<=0,8;"A";ЕСЛИ(Накопительная доля<=0,95;"B";"C"))
            cumshare = pivot_abc["Накопительная доля"]
            pivot_abc["Оценка"] = "C"
            pivot_abc.loc[cumshare <= 0.95, "Оценка"] = "B"
            pivot_abc.loc[cumshare <= 0.8, "Оценка"] = "A"
            pivot_abc.to_excel(writer, sheet_name="ABC")
            _set_artikul_number_format(writer.sheets["ABC"], 1, len(pivot_abc))
            # Внизу таблицы: колонка A — надпись «Общая прибыль», колонка B — значение (сумма прибыли); колонки C–E пустые
            ws_abc = writer.sheets["ABC"]
            total_row = 2 + len(pivot_abc)  # строка 1 — заголовок, далее данные
            ws_abc.cell(row=total_row, column=1, value="Общая прибыль")
            ws_abc.cell(row=total_row, column=2, value=total_profit)
            # Колонки 3–5 для строки «Общая прибыль» пустые

            # Заливка ячеек столбца «Оценка»: A — зелёный, B — жёлтый, C — красный (только для строк с артикулами)
            col_ocenka_abc = 5
            fill_a = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # зелёный
            fill_b = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # жёлтый
            fill_c = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # красный
            for r, val in enumerate(pivot_abc["Оценка"], start=2):
                cell = ws_abc.cell(row=r, column=col_ocenka_abc)
                if val == "A":
                    cell.fill = fill_a
                elif val == "B":
                    cell.fill = fill_b
                else:
                    cell.fill = fill_c

            # Легенда ABC под таблицей (с колонки A): таблица «Оценка» | «Характеристика», ячейки оценки — цветные
            legend_col_abc = 1
            legend_start_row_abc = total_row + 3
            thin_side = Side(style="thin", color="000000")
            _fill_abc_leg = (fill_a, fill_b, fill_c)
            _text_abc_leg = ("самые значимые", "средние", "наименьшие")
            # Заголовок таблицы
            h1 = ws_abc.cell(row=legend_start_row_abc, column=legend_col_abc, value="Оценка")
            h2 = ws_abc.cell(row=legend_start_row_abc, column=legend_col_abc + 1, value="Характеристика")
            for c in (h1, h2):
                c.font = Font(bold=True)
                c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                c.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws_abc.column_dimensions[get_column_letter(legend_col_abc)].width = 12
            ws_abc.column_dimensions[get_column_letter(legend_col_abc + 1)].width = 28
            for i in range(3):
                r = legend_start_row_abc + 1 + i
                cell_code = ws_abc.cell(row=r, column=legend_col_abc, value=["A", "B", "C"][i])
                cell_code.fill = _fill_abc_leg[i]
                cell_code.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                cell_code.alignment = Alignment(horizontal="center", vertical="center")
                cell_code.font = Font(bold=True)
                cell_txt = ws_abc.cell(row=r, column=legend_col_abc + 1, value=_text_abc_leg[i])
                cell_txt.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                cell_txt.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws_abc.row_dimensions[r].height = 22

        # Лист XYZ: две таблицы — «Нерегулярный спрос» (Noreg) и «Сглаженный спрос» (reg)
        if "Артикул" in merged_export.columns and "Дата отгрузки" in merged_export.columns:
            df_xyz = merged_export.copy()
            dt = pd.to_datetime(df_xyz["Дата отгрузки"], errors="coerce")
            df_xyz["Месяцы (Дата отгрузки)"] = dt.apply(
                lambda x: format_month_year(x.year, x.month) if pd.notna(x) else ""
            )
            pivot_raw = df_xyz.pivot_table(
                index="Артикул",
                columns="Месяцы (Дата отгрузки)",
                values="Дата отгрузки",
                aggfunc="count",
                margins=False,
            )
            pivot_raw = pivot_raw.loc[:, [c for c in pivot_raw.columns if str(c).strip()]]
            month_cols = list(pivot_raw.columns)
            num_months = len(month_cols)
            regularity_threshold = round(num_months * 0.8)
            months_with_sales = (pivot_raw > 0).sum(axis=1)
            mask_noreg = months_with_sales < regularity_threshold
            mask_reg = months_with_sales >= regularity_threshold

            fill_x = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fill_y = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            fill_z = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            fill_nd = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            thin = Side(style="thin", color="000000")

            def _cv_to_rating(cv_ser):
                """Noreg: X≤0.25, Y≤0.50, иначе Z; CV=0→Недостаточно данных."""
                out = pd.Series("Z", index=cv_ser.index)
                out.loc[cv_ser <= 0.50] = "Y"
                out.loc[cv_ser <= 0.25] = "X"
                out.loc[cv_ser == 0] = "Недостаточно данных"
                return out

            def _cv_to_rating_reg(cv_ser):
                """Reg: CV≤0.20→X; 0.20–0.35→Y1; 0.35–0.50→Y2; >0.50→Y3; CV=0→Недостаточно данных."""
                out = pd.Series("Y3", index=cv_ser.index)
                out.loc[cv_ser <= 0.50] = "Y2"
                out.loc[cv_ser <= 0.35] = "Y1"
                out.loc[cv_ser <= 0.20] = "X"
                out.loc[cv_ser == 0] = "Недостаточно данных"
                return out

            # Группа Noreg: среднее по месяцам, CV, оценка. Общий итог = сумма заказов по артикулу за все месяцы периода.
            df_noreg = pivot_raw.loc[mask_noreg].copy()
            df_noreg["Общий итог"] = df_noreg[month_cols].sum(axis=1)
            df_noreg["Среднее значение"] = df_noreg[month_cols].mean(axis=1)
            df_noreg["Стандартное отклонение"] = df_noreg[month_cols].std(axis=1, ddof=0)
            mean_n = df_noreg["Среднее значение"]
            df_noreg["Коэффициент вариации"] = df_noreg["Стандартное отклонение"].div(mean_n).where(mean_n != 0, 0)
            df_noreg["Оценка"] = _cv_to_rating(df_noreg["Коэффициент вариации"])
            df_noreg = df_noreg.reset_index()
            cols_noreg = ["Артикул"] + month_cols + ["Общий итог", "Среднее значение", "Стандартное отклонение", "Коэффициент вариации", "Оценка"]
            df_noreg = df_noreg[[c for c in cols_noreg if c in df_noreg.columns]]

            # Группа reg: для среднего и Std — только месяцы со значением >= 0.3*медианы; в расчёт берём исходное значение
            # и винзоризованное в случае замены (исходное > медиана*2.5). Значения < 0.3*медианы полностью не учитываются.
            df_reg_orig = pivot_raw.loc[mask_reg].copy()
            median_ser = df_reg_orig[month_cols].median(axis=1)
            cap_ser = median_ser * 2.5
            winsorized = df_reg_orig[month_cols].clip(upper=cap_ser, axis=0)  # исходное или cap где была винзоризация
            threshold_low = median_ser * 0.3
            # Построчно: участвуют только месяцы, где значение >= 0.3*медианы; берём винзоризованное (исходное/замена)
            mean_winz = pd.Series(index=winsorized.index, dtype=float)
            std_winz = pd.Series(index=winsorized.index, dtype=float)
            for idx in winsorized.index:
                row = winsorized.loc[idx]   # исходное или винзоризованное в случае замены
                th = threshold_low.loc[idx]
                kept = row[row >= th].values  # только >= 0.3*медианы; остальные не участвуют
                if len(kept) == 0:
                    mean_winz.loc[idx] = np.nan
                    std_winz.loc[idx] = np.nan
                else:
                    mean_winz.loc[idx] = float(np.mean(kept))
                    std_winz.loc[idx] = float(np.std(kept, ddof=0))
            # Колонки: Артикул, [M1, Винз. M1, M2, Винз. M2, ...], Общий итог, Медиана, Среднее значение, Std, CV, Оценка
            build_reg = {"Артикул": df_reg_orig.index}
            for m in month_cols:
                build_reg[m] = df_reg_orig[m].values
                w = winsorized[m]
                build_reg[f"Винз. {m}"] = w.where(df_reg_orig[m] > cap_ser).values  # значение только где была винзоризация
            total_winz = winsorized.sum(axis=1)
            cv_reg = std_winz.div(mean_winz).where(mean_winz.notna() & (mean_winz != 0), 0)
            rating_reg = _cv_to_rating_reg(cv_reg)
            build_reg["Общий итог"] = total_winz.values
            build_reg["Медиана"] = median_ser.values
            build_reg["Среднее значение"] = mean_winz.values
            build_reg["Стандартное отклонение"] = std_winz.values
            build_reg["Коэффициент вариации"] = cv_reg.values
            build_reg["Оценка"] = rating_reg.values
            df_reg = pd.DataFrame(build_reg)

            pivot_xyz = pd.concat([
                df_noreg[["Артикул", "Оценка"]].set_index("Артикул"),
                df_reg[["Артикул", "Оценка"]].set_index("Артикул"),
            ])

            pd.DataFrame().to_excel(writer, sheet_name="XYZ", index=False)
            ws = writer.sheets["XYZ"]
            row_cur = 1
            ws.cell(row=row_cur, column=1, value="Нерегулярный спрос")
            ws.cell(row=row_cur, column=1).font = Font(bold=True, size=12)
            row_cur += 1
            for c, name in enumerate(df_noreg.columns, start=1):
                cell = ws.cell(row=row_cur, column=c, value=name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            row_cur += 1
            for i in range(len(df_noreg)):
                for c, name in enumerate(df_noreg.columns, start=1):
                    val = df_noreg.iloc[i][name]
                    cell = ws.cell(row=row_cur, column=c, value=val)
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    if name == "Оценка":
                        v = str(val).strip()
                        if v == "X": cell.fill = fill_x
                        elif v == "Y": cell.fill = fill_y
                        elif v == "Z": cell.fill = fill_z
                        elif v == "Недостаточно данных": cell.fill = fill_nd
                row_cur += 1
            row_cur += 2
            ws.cell(row=row_cur, column=1, value="Сглаженный спрос")
            ws.cell(row=row_cur, column=1).font = Font(bold=True, size=12)
            row_cur += 1
            for c, name in enumerate(df_reg.columns, start=1):
                cell = ws.cell(row=row_cur, column=c, value=name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            row_cur += 1
            fill_y1 = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Y1 умеренно
            fill_y2 = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")   # Y2 выраженно
            fill_y3 = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")   # Y3 очень сезонный
            for i in range(len(df_reg)):
                for c, name in enumerate(df_reg.columns, start=1):
                    val = df_reg.iloc[i][name]
                    cell = ws.cell(row=row_cur, column=c, value=val)
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    if name == "Оценка":
                        v = str(val).strip()
                        if v == "X": cell.fill = fill_x
                        elif v == "Y1": cell.fill = fill_y1
                        elif v == "Y2": cell.fill = fill_y2
                        elif v == "Y3": cell.fill = fill_y3
                        elif v == "Недостаточно данных": cell.fill = fill_nd
                row_cur += 1

            for r in range(2, 2 + len(df_noreg)):
                ws.cell(row=r, column=1).number_format = "0"
            reg_table_start = 2 + len(df_noreg) + 2 + 2
            for r in range(reg_table_start, reg_table_start + len(df_reg)):
                ws.cell(row=r, column=1).number_format = "0"

            legend_col = 1
            legend_start_row_xyz = row_cur + 3
            # Легенда XYZ: Класс | Характеристика | Управленческие рекомендации (X, Y1, Y2, Y3, Z)
            _legend_xyz_rows = [
                ("X", "Стабильный регулярный спрос", "Автопополнение запасов\nМинимальные страховые запасы\nПриоритет в планировании"),
                ("Y1", "Слабые колебания спроса", "Почти как X\nЛёгкая корректировка запасов"),
                ("Y2", "Сезонность спроса", "Планирование по сезонам\nПодготовка к пиковым периодам"),
                ("Y3", "Сильная сезонность", "Короткие циклы закупки\nПовышенный контроль остатков"),
                ("Z", "Нерегулярный / прерывистый спрос", "Закупка под заказ\nМинимальные или нулевые остатки\nКандидат на вывод"),
            ]
            _fill_xyz_leg = (fill_x, fill_y1, fill_y2, fill_y3, fill_z)
            for col, hdr in enumerate(("Класс", "Характеристика", "Управленческие рекомендации"), start=legend_col):
                cell = ws.cell(row=legend_start_row_xyz, column=col, value=hdr)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(legend_col)].width = 12
            ws.column_dimensions[get_column_letter(legend_col + 1)].width = 32
            ws.column_dimensions[get_column_letter(legend_col + 2)].width = 48
            for i, (klass, char, rec) in enumerate(_legend_xyz_rows):
                r = legend_start_row_xyz + 1 + i
                cell_code = ws.cell(row=r, column=legend_col, value=klass)
                cell_code.fill = _fill_xyz_leg[i]
                cell_code.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell_code.alignment = Alignment(horizontal="center", vertical="center")
                cell_code.font = Font(bold=True)
                c_char = ws.cell(row=r, column=legend_col + 1, value=char)
                c_char.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                c_char.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                c_rec = ws.cell(row=r, column=legend_col + 2, value=rec)
                c_rec.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                c_rec.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws.row_dimensions[r].height = 38

        # Лист «Итог»: Артикул, Оценка по ABC, Оценка по XYZ, Общая оценка ABCXYZ; артикулы в порядке убывания Общая оценка (AX вверху, CZ внизу)
        if pivot_abc is not None and pivot_xyz is not None:
            itog_abc = pivot_abc[["Оценка"]].rename(columns={"Оценка": "Оценка по ABC"})
            itog_xyz = pivot_xyz[["Оценка"]].rename(columns={"Оценка": "Оценка по XYZ"})
            itog_df = itog_abc.join(itog_xyz, how="outer")
            itog_df = itog_df.reset_index()
            itog_df["Общая оценка ABCXYZ"] = (
                itog_df["Оценка по ABC"].fillna("").astype(str)
                + itog_df["Оценка по XYZ"].fillna("").astype(str)
            )
            # При «Оценка по XYZ» = «Недостаточно данных» в «Общая оценка ABCXYZ» тоже «Недостаточно данных»
            nd_xyz = itog_df["Оценка по XYZ"].astype(str).str.strip() == "Недостаточно данных"
            itog_df.loc[nd_xyz, "Общая оценка ABCXYZ"] = "Недостаточно данных"
            # Сортировка: AX первым; XYZ: X < Y1 < Y2 < Y3 < Y < Z; «Недостаточно данных» — в конце
            order_abc = {"A": 0, "B": 1, "C": 2}
            def _xyz_order(s):
                s = str(s).strip()
                if s == "X": return 0
                if s.startswith("Y1"): return 1
                if s.startswith("Y2"): return 2
                if s.startswith("Y3"): return 3
                if s == "Y": return 1
                if s == "Z": return 4
                if s == "Недостаточно данных": return 99
                return 99
            xyz_col = itog_df["Оценка по XYZ"].astype(str).str.strip()
            obsh = itog_df["Общая оценка ABCXYZ"].astype(str).str.strip()
            itog_df["_s1"] = obsh.str[0:1].map(order_abc).fillna(99)
            itog_df["_s2"] = xyz_col.apply(_xyz_order)
            itog_df.loc[obsh == "Недостаточно данных", "_s1"] = 99
            itog_df.loc[obsh == "Недостаточно данных", "_s2"] = 99
            itog_df = itog_df.sort_values(["_s1", "_s2"]).drop(columns=["_s1", "_s2"])
            itog_df.to_excel(writer, sheet_name="Итог", index=False)
            _set_artikul_number_format(writer.sheets["Итог"], 1, len(itog_df))
            ws_itog = writer.sheets["Итог"]
            # Заливка: колонки 2–3 — по правилам ABC/XYZ (A/X=зелёный, B/Y=жёлтый, C/Z=красный)
            fill_g = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fill_y = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            fill_r = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            # Колонка 4 — Общая оценка ABCXYZ: своя заливка для каждой комбинации (AX/CZ-матрица)
            fill_ax = PatternFill(start_color="9ACD32", end_color="9ACD32", fill_type="solid")   # AX — насыщенный жёлто-зелёный
            fill_ay = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")   # AY — насыщенный жёлтый
            fill_az = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")   # AZ — насыщенный оранжевый
            fill_bx = PatternFill(start_color="ADFF2F", end_color="ADFF2F", fill_type="solid")   # BX — жёлто-зелёный
            fill_by = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # BY — жёлтый
            fill_bz = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")   # BZ — оранжевый
            fill_cx = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")   # CX — светло-жёлтый
            fill_cy = PatternFill(start_color="DAA520", end_color="DAA520", fill_type="solid")   # CY — тёмно-жёлтый / жёлто-оранжевый
            fill_cz = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")   # CZ — светло-оранжево-красный
            fill_nd = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")   # серый — недостаточно данных
            fill_abcxyz = {"AX": fill_ax, "AY": fill_ay, "AZ": fill_az, "BX": fill_bx, "BY": fill_by, "BZ": fill_bz, "CX": fill_cx, "CY": fill_cy, "CZ": fill_cz}
            for yc in ("Y1", "Y2", "Y3"):
                fill_abcxyz["A" + yc] = fill_ay
                fill_abcxyz["B" + yc] = fill_by
                fill_abcxyz["C" + yc] = fill_cy
            for r in range(2, len(itog_df) + 2):
                val_abc = str(itog_df.iloc[r - 2]["Оценка по ABC"]) if pd.notna(itog_df.iloc[r - 2]["Оценка по ABC"]) else ""
                val_xyz = str(itog_df.iloc[r - 2]["Оценка по XYZ"]) if pd.notna(itog_df.iloc[r - 2]["Оценка по XYZ"]) else ""
                val_obsh_raw = str(itog_df.iloc[r - 2]["Общая оценка ABCXYZ"]).strip()
                val_obsh = val_obsh_raw.upper()[:2]
                # Колонка 2 — Оценка по ABC
                c2 = ws_itog.cell(row=r, column=2)
                if val_abc == "A": c2.fill = fill_g
                elif val_abc == "B": c2.fill = fill_y
                elif val_abc == "C": c2.fill = fill_r
                # Колонка 3 — Оценка по XYZ
                c3 = ws_itog.cell(row=r, column=3)
                if val_xyz == "X": c3.fill = fill_g
                elif val_xyz == "Y": c3.fill = fill_y
                elif val_xyz == "Z": c3.fill = fill_r
                elif val_xyz == "Y1": c3.fill = fill_y
                elif val_xyz == "Y2": c3.fill = fill_y
                elif val_xyz == "Y3": c3.fill = fill_y
                elif val_xyz == "Недостаточно данных": c3.fill = fill_nd
                # Колонка 4 — Общая оценка ABCXYZ: код = первая буква + X/Y1/Y2/Y3/Y/Z
                c4 = ws_itog.cell(row=r, column=4)
                if val_obsh_raw == "Недостаточно данных":
                    c4.fill = fill_nd
                else:
                    code_abc = val_obsh_raw[:1] if val_obsh_raw else ""
                    code_xyz = (val_xyz.split()[0] if val_xyz else "")[:2]  # "Y1", "Y2", "Y3", "X", "Y", "Z"
                    if not code_xyz and len(val_xyz) >= 1:
                        code_xyz = val_xyz[:1]
                    code = (code_abc + code_xyz).upper()
                    if code in fill_abcxyz:
                        c4.fill = fill_abcxyz[code]

            # Легенда ABC/XYZ: Код | Интерпретация | Управленческие рекомендации
            legend_itog = [
                ("AX", "Опора бизнеса\nВысокий оборот, стабильный регулярный спрос", ["Высокий уровень сервиса", "Минимальные страховые запасы", "Приоритет в закупках и планировании", "Основа для масштабирования"]),
                ("AY1", "Высокий оборот, слабая сезонность\nРегулярный спрос с малыми колебаниями", ["Поддерживать наличие", "Планирование почти как для X", "Допустимы инвестиции"]),
                ("AY2", "Высокий оборот, сезонный\nУправляемые колебания спроса", ["Учитывать сезонность и промо", "Гибкое планирование запасов", "Контроль перед пиковыми периодами"]),
                ("AY3", "Высокий оборот, сильная сезонность\nРегулярный, но волатильный спрос", ["Короткие циклы планирования", "Минимизация излишков вне сезона", "Осторожная реклама"]),
                ("AZ", "Ценный, но нерегулярный\nВысокий оборот, прерывистый спрос", ["Частый пересмотр прогнозов", "Работа под заказ / малыми партиями", "Ограничение складских запасов"]),
                ("BX", "Рабочая лошадка\nСредний оборот, стабильный спрос", ["Стандартные правила управления", "Оптимизация запасов", "Контроль оборачиваемости"]),
                ("BY1", "Средний оборот, слабая сезонность", ["Поддерживать наличие", "Планирование близко к BX"]),
                ("BY2", "Средний оборот, сезонный\nКолебания спроса", ["Контроль сезонности", "Возможен перевод в BX или BZ"]),
                ("BY3", "Средний оборот, сильная сезонность", ["Минимизация остатков вне сезона", "Проверка рентабельности"]),
                ("BZ", "Повышенный риск\nСредний оборот, нерегулярный спрос", ["Сокращение запасов", "Проверка целесообразности наличия", "Ограничение инвестиций"]),
                ("CX", "Стабильный якорь\nНизкий оборот, регулярный спрос", ["Минимальные запасы", "Без активных инвестиций", "Поддержка ассортимента"]),
                ("CY1", "Низкий оборот, слабая сезонность", ["Строгий контроль остатков", "Проверка цены"]),
                ("CY2", "Низкий оборот, сезонный", ["Минимизация запасов", "Оптимизация SKU"]),
                ("CY3", "Низкий оборот, сильная сезонность", ["Кандидат на сокращение", "Работа только в сезон"]),
                ("CZ", "Кандидат на вывод\nНизкий оборот, нерегулярный спрос", ["Вывод из ассортимента", "Распродажа остатков", "Оставлять только по стратегии"]),
            ]
            legend_col_itog = 1
            legend_start_row_itog = 1 + len(itog_df) + 3
            thin_leg = Side(style="thin", color="000000")
            for col, hdr in enumerate(("Код", "Интерпретация", "Управленческие рекомендации"), start=legend_col_itog):
                cell = ws_itog.cell(row=legend_start_row_itog, column=col, value=hdr)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.border = Border(left=thin_leg, right=thin_leg, top=thin_leg, bottom=thin_leg)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws_itog.column_dimensions[get_column_letter(legend_col_itog)].width = 10
            ws_itog.column_dimensions[get_column_letter(legend_col_itog + 1)].width = 38
            ws_itog.column_dimensions[get_column_letter(legend_col_itog + 2)].width = 52
            for i, (code, interpretation, bullets) in enumerate(legend_itog, start=0):
                row_num = legend_start_row_itog + 1 + i
                cell_code = ws_itog.cell(row=row_num, column=legend_col_itog, value=code)
                cell_code.fill = fill_abcxyz.get(code, PatternFill())
                cell_code.border = Border(left=thin_leg, right=thin_leg, top=thin_leg, bottom=thin_leg)
                cell_code.alignment = Alignment(horizontal="center", vertical="center")
                cell_code.font = Font(bold=True)
                cell_interp = ws_itog.cell(row=row_num, column=legend_col_itog + 1, value=interpretation)
                cell_interp.border = Border(left=thin_leg, right=thin_leg, top=thin_leg, bottom=thin_leg)
                cell_interp.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                text_rec = "\n".join("• " + b for b in bullets)
                cell_rec = ws_itog.cell(row=row_num, column=legend_col_itog + 2, value=text_rec)
                cell_rec.border = Border(left=thin_leg, right=thin_leg, top=thin_leg, bottom=thin_leg)
                cell_rec.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws_itog.row_dimensions[row_num].height = 72

        # Отчёт на листе «Отчёт» (если есть пропуски)
        report_rows = []
        for f, s, msg in report_partial:
            report_rows.append({"Файл": f, "Лист": s, "Комментарий": msg, "Тип": "Частично найдено"})
        for f, s, msg in report_missing:
            report_rows.append({"Файл": f, "Лист": s, "Комментарий": msg, "Тип": "Не найдено"})
        if report_rows:
            rep_df = pd.DataFrame(report_rows)
            rep_df.to_excel(writer, sheet_name="Отчёт", index=False)

    # Лист «Итог» — первым и активным при открытии файла
    if pivot_abc is not None and pivot_xyz is not None:
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        if "Итог" in wb.sheetnames:
            idx = wb.sheetnames.index("Итог")
            wb.move_sheet("Итог", offset=-idx)
            wb.active = wb["Итог"]
            wb.save(output_path)

    print(f"Готово. Сохранено: {output_path}")

def main(argv: Optional[List[str]] = None):
    parser = argparse.ArgumentParser(description="Собрать заказы из всех Excel-отчётов папки в один файл для ABC&XYZ-анализа.")
    parser.add_argument("-i", "--input_dir", default="./reports", help="Папка с входными Excel отчётами (по умолчанию ./reports)")
    parser.add_argument("-o", "--output", default=None, help="Путь к выходному файлу (если не задан и задан --output_dir, имя берётся по диапазону месяцев)")
    parser.add_argument("--output_dir", default=None, help='Папка для сохранения (например "ABC&XYZ reports"). Имя файла: «Ранний месяц год-Поздний месяц год.xlsx»')
    parser.add_argument("--from-month", type=int, default=None, metavar="M", help="Месяц начала диапазона (1–12); вместе с --from-year/--to-month/--to-year ограничивает обрабатываемые файлы")
    parser.add_argument("--from-year", type=int, default=None, metavar="Y", help="Год начала диапазона")
    parser.add_argument("--to-month", type=int, default=None, metavar="M", help="Месяц конца диапазона (1–12)")
    parser.add_argument("--to-year", type=int, default=None, metavar="Y", help="Год конца диапазона")
    args = parser.parse_args(argv)

    if not os.path.isdir(args.input_dir):
        print(f"Папка не найдена: {args.input_dir}")
        sys.exit(1)

    merge_folder(
        args.input_dir,
        output_path=args.output,
        output_dir=args.output_dir,
        from_month=args.from_month,
        from_year=args.from_year,
        to_month=args.to_month,
        to_year=args.to_year,
    )

if __name__ == "__main__":
    main()
